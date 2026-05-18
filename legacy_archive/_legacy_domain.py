import csv
import io
import json
import logging
import mimetypes
import os
import re
import uuid
import zipfile
import hashlib
from collections import defaultdict
from datetime import date, datetime, timedelta
from difflib import SequenceMatcher
from html import unescape
from io import BytesIO
from itertools import chain
from pathlib import Path as FilePath
from types import SimpleNamespace
from typing import Dict, List, Optional, Tuple
from urllib.parse import quote
from xml.sax.saxutils import escape
import requests
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.contenttypes.models import ContentType
from django.core.cache import cache
from django.core.exceptions import PermissionDenied
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage, default_storage
from django.core.paginator import EmptyPage, PageNotAnInteger, Paginator
from django.db import IntegrityError, connections, transaction
from django.db.models import (
    Count,
    Exists,
    F,
    Max,
    OuterRef,
    Prefetch,
    Q,
    Subquery,
    Value,
    CharField,
)
from django.db.models.functions import Cast, Coalesce, Concat
from django.db.utils import ProgrammingError
from django.http import (
    FileResponse,
    Http404,
    HttpResponse,
    HttpResponseBadRequest,
    HttpResponseForbidden,
    JsonResponse,
)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.timezone import make_naive
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_GET, require_http_methods, require_POST
from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from rest_framework import filters, generics, permissions, status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from FIU.models import *
from UserManagement.models import *
from userspermissionsystem.models import *
from userspermissionsystem.views import can_create_ticket

from .forms import *
from .models import *
from .serializers import *
from .utils import *

logger = logging.getLogger(__name__)

try:
    from PyPDF2 import PdfReader
except Exception:
    PdfReader = None

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False


class NGOSearchAPIView(generics.ListAPIView):
    def list(self, request):
        search = request.query_params.get("search", "").strip()
        page = int(request.query_params.get("page", 1))
        page_size = 20  # optional pagination

        # Base queryset using secondary DB
        queryset = RegistrationsOther.objects.using('secondary').all()

        # Apply "like" filter on association_name (case-insensitive)
        if search:
            queryset = queryset.filter(association_name__icontains=search)

        total_count = queryset.count()

        # Pagination
        start = (page - 1) * page_size
        end = start + page_size
        results = queryset.order_by('association_name')[start:end]

        # Serialize results
        data = [
            {
                "id": r.id,
                "association_name": r.association_name,
                 "rcn": r.rcn,  # <-- lowercase
                # "mapStatus": r.mapStatus,
                "created_at": r.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "updated_at": r.updated_at.strftime("%Y-%m-%d %H:%M:%S")
            } for r in results
        ]
        # print(data,'--------------')
        return Response({
            "page": page,
            "page_size": page_size,
            "total_count": total_count,
            "data": data
        })

class STRSearchAPIView(generics.ListAPIView):
    queryset = FIUSTRDetails.objects.all()
    serializer_class = STRSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['case_id', 'main_entity']


class UserSearchAPIView(generics.ListAPIView):
    queryset = CustomUser.objects.all()
    serializer_class = UserSerializer
    filter_backends = [filters.SearchFilter]
    search_fields = ['first_name', 'phone_no', 'rank']


class UserByRankAPIView(generics.ListAPIView):
    serializer_class = UserSerializer

    def get_queryset(self):
        rank = self.request.query_params.get('rank')
        if rank:
            return CustomUser.objects.filter(rank=rank)
        return CustomUser.objects.none()


class TicketCreateAPIView(generics.CreateAPIView):
    queryset = Ticket.objects.all()
    serializer_class = TicketFormCreateSerializer
    permission_classes = [permissions.IsAuthenticated]


    def perform_create(self, serializer):
        print('fvd = ', self.request.user)
        serializer.save(created_by=self.request.user)     
@login_required(login_url="login")
def ticket_list(request):
    user = request.user

    params = request.GET.copy()
    params.pop("page", None)
    preserved_qs = params.urlencode()

    base_params = request.GET.copy()
    for key in [
        "page", "scope", "created_bucket", "created_sub",
        "my_bucket", "my_sub", "cat", "sec", "cat_status",
    ]:
        base_params.pop(key, None)

    base_filter_qs = base_params.urlencode()

    reassignment_model = Ticket.reassignments.rel.related_model

    report_exists_sq = ReportSummary.objects.filter(ticket_id=OuterRef("pk"))

    latest_ra_id_sq = Subquery(
        reassignment_model.objects
        .filter(ticket_id=OuterRef("pk"))
        .order_by("-created_at")
        .values("id")[:1]
    )

    action_taken_exists = TicketActionTrace.objects.filter(
        ticket_id=OuterRef("pk"),
        action_key="ACTION_TAKEN",
        is_success=True,
    )

    if user.is_superuser:
        base_qs = Ticket.objects.all()
    else:
        base_qs = Ticket.objects.filter(
            Q(created_by=user) |
            Q(assigned_users=user) |
            Q(reassignments__new_assigned_users=user)
        ).distinct()

    base_qs = (
        base_qs
        .filter(is_deleted=False)
        .select_related("created_by")
        .annotate(
            has_report=Exists(report_exists_sq),
            latest_ra_id_annotated=latest_ra_id_sq,
            has_action_taken_db=Exists(action_taken_exists),
        )
        .order_by("-created_at", "-id")
        .distinct()
    )

    qs = base_qs

    search_query = request.GET.get("search", "").strip()
    status_filter = request.GET.get("status", "all").strip() or "all"
    module_filter = request.GET.get("module", "").strip()
    from_date_raw = request.GET.get("from_date", "").strip()
    to_date_raw = request.GET.get("to_date", "").strip()
    state_filter = request.GET.get("assoc_state", "").strip()

    category_filter = request.GET.get("cat", "").strip()
    section_filter = request.GET.get("sec", "").strip()
    cat_status = request.GET.get("cat_status", "").strip()

    scope = request.GET.get("scope", "").strip()
    created_bucket = request.GET.get("created_bucket", "").strip()
    my_bucket = request.GET.get("my_bucket", "").strip()

    from_date = parse_date(from_date_raw) if from_date_raw else None
    to_date = parse_date(to_date_raw) if to_date_raw else None

    if search_query:
        if search_query.isdigit():
            qs = qs.filter(
                Q(id=int(search_query)) |
                Q(mha_file_no__icontains=search_query) |
                Q(file_no__icontains=search_query)
            )
        else:
            search_terms = [
                term.strip()
                for term in search_query.split()
                if term.strip()
            ]

            search_filter = (
                Q(mha_file_no__icontains=search_query) |
                Q(file_no__icontains=search_query) |
                Q(models_object__icontains=search_query) |
                Q(models_name__icontains=search_query) |
                Q(status__icontains=search_query) |
                Q(created_by__first_name__icontains=search_query) |
                Q(created_by__last_name__icontains=search_query) |
                Q(assigned_users__first_name__icontains=search_query) |
                Q(assigned_users__last_name__icontains=search_query) |
                Q(assigned_users__rank__icontains=search_query) |
                Q(assigned_users__ucs_no__icontains=search_query) |
                Q(assigned_users__phone_no__icontains=search_query) |
                Q(reassignments__new_assigned_users__first_name__icontains=search_query) |
                Q(reassignments__new_assigned_users__last_name__icontains=search_query) |
                Q(reassignments__new_assigned_users__rank__icontains=search_query) |
                Q(reassignments__new_assigned_users__ucs_no__icontains=search_query) |
                Q(reassignments__new_assigned_users__phone_no__icontains=search_query)
            )

            if len(search_terms) > 1:
                multi_term_filter = Q()
                for term in search_terms:
                    multi_term_filter &= (
                        Q(created_by__first_name__icontains=term) |
                        Q(created_by__last_name__icontains=term) |
                        Q(assigned_users__first_name__icontains=term) |
                        Q(assigned_users__last_name__icontains=term) |
                        Q(reassignments__new_assigned_users__first_name__icontains=term) |
                        Q(reassignments__new_assigned_users__last_name__icontains=term)
                    )
                search_filter |= multi_term_filter

            parsed_search_date = parse_date(search_query)
            if parsed_search_date:
                search_filter |= Q(created_at__date=parsed_search_date)

            osdl_ids_for_search_state = list(
                OtherServicesDocumentsLink.objects.using("secondary")
                .filter(association_state__icontains=search_query)
                .values_list("id", flat=True)
            )

            if osdl_ids_for_search_state:
                search_filter |= Q(
                    models_id__in=[str(i) for i in osdl_ids_for_search_state]
                )

            qs = qs.filter(search_filter).distinct()

    if status_filter and status_filter != "all":
        if status_filter == "reassigned":
            qs = qs.filter(reassignments__isnull=False).distinct()
        else:
            qs = qs.filter(status=status_filter)

    if module_filter:
        qs = qs.filter(models_name__icontains=module_filter)

    if from_date:
        qs = qs.filter(created_at__date__gte=from_date)

    if to_date:
        qs = qs.filter(created_at__date__lte=to_date)

    if state_filter:
        osdl_ids_for_state = list(
            OtherServicesDocumentsLink.objects.using("secondary")
            .filter(association_state__iexact=state_filter)
            .values_list("id", flat=True)
        )
        qs = qs.filter(models_id__in=[str(i) for i in osdl_ids_for_state])

    dashboard_base_qs = qs.distinct()

    filtered_ticket_ids = filter_ticket_ids_by_scope(
        ticket_qs=dashboard_base_qs,
        scope=scope,
        created_bucket=created_bucket,
        my_bucket=my_bucket,
        user=user,
    )

    if filtered_ticket_ids is not None:
        if filtered_ticket_ids:
            qs = qs.filter(id__in=filtered_ticket_ids)
        else:
            qs = qs.none()

    if category_filter and section_filter:
        qs = (
            qs
            .filter(assigned_user_categories__section_type=section_filter)
            .filter(build_category_q(
                category_filter,
                "assigned_user_categories__category__"
            ))
            .distinct()
        )

        if cat_status == "new":
            qs = qs.exclude(status="closed").exclude(
                action_traces__action_key="ACTION_TAKEN",
                action_traces__is_success=True,
            ).distinct()

        elif cat_status == "pending":
            qs = qs.exclude(status="closed").filter(
                action_traces__action_key="ACTION_TAKEN",
                action_traces__is_success=True,
            ).distinct()

        elif cat_status == "closed":
            qs = qs.filter(status="closed")

    final_qs = qs.distinct().order_by("-created_at", "-id")

    allowed_sizes = [10, 20, 50, 100, 200]
    page_size_raw = request.GET.get("page_size", "10").strip()

    try:
        page_size = int(page_size_raw)
    except ValueError:
        page_size = 10

    if page_size not in allowed_sizes:
        page_size = 10

    page = request.GET.get("page", 1)
    paginator = Paginator(final_qs, page_size)

    try:
        tickets_page = paginator.page(page)
    except PageNotAnInteger:
        tickets_page = paginator.page(1)
    except EmptyPage:
        tickets_page = paginator.page(paginator.num_pages)

    page_ids = list(tickets_page.object_list.values_list("id", flat=True))

    page_ticket_qs = (
        Ticket.objects
        .filter(id__in=page_ids)
        .select_related("created_by")
        .annotate(has_report=Exists(report_exists_sq))
        .prefetch_related(
            "assigned_users",
            Prefetch(
                "answers",
                queryset=TicketAnswer.objects.filter(is_active=True).only(
                    "id", "ticket_id", "answered_by_id",
                    "answer", "is_active", "is_final",
                ),
            ),
            Prefetch(
                "reassignments",
                queryset=(
                    reassignment_model.objects
                    .filter(ticket__is_deleted=False)
                    .order_by("-created_at")
                    .select_related("reassigned_by")
                    .prefetch_related(
                        "new_assigned_users",
                        Prefetch(
                            "answers",
                            queryset=ReassignmentAnswer.objects.filter(
                                is_active=True
                            ).only(
                                "id", "reassignment_id", "answered_by_id",
                                "answer", "is_active", "is_final",
                            ),
                        ),
                    )
                ),
            ),
        )
    )

    page_ticket_map = {obj.id: obj for obj in page_ticket_qs}
    ordered_page_tickets = [
        page_ticket_map[ticket_id]
        for ticket_id in page_ids
        if ticket_id in page_ticket_map
    ]

    tickets_page.object_list = ordered_page_tickets

    osdl_ids = [
        t.models_id
        for t in ordered_page_tickets
        if t.models_id and str(t.models_id).isdigit()
    ]

    osdl_map = {}
    if osdl_ids:
        osdl_qs = OtherServicesDocumentsLink.objects.using("secondary").filter(
            id__in=osdl_ids
        )
        osdl_map = {str(o.id): o for o in osdl_qs}

    for ticket in ordered_page_tickets:
        ticket.osdl = osdl_map.get(str(ticket.models_id))

        summary = get_ticket_reply_summary(ticket, current_user=user)
        latest_ra = summary["latest_ra"]
        ras = list(ticket.reassignments.all())

        assignee_user_ids = set(ticket.assigned_users.values_list("id", flat=True))

        reassign_user_ids = set()
        for ra in ras:
            reassign_user_ids.update(
                ra.new_assigned_users.values_list("id", flat=True)
            )

        allowed_assignee_ids = assignee_user_ids | reassign_user_ids

        action_taken = (
            ticket.action_traces
            .filter(
                action_key="ACTION_TAKEN",
                is_success=True,
                performed_by_id__in=allowed_assignee_ids,
            )
            .order_by("-performed_at")
            .first()
        )

        ticket.has_action_taken_db = bool(action_taken)
        ticket.action_taken = action_taken
        ticket.latest_ra_id = latest_ra.id if latest_ra else None

        my_latest = None
        for ra in ras:
            ra_user_ids = {u.id for u in ra.new_assigned_users.all()}
            if user.id in ra_user_ids:
                my_latest = ra
                break

        ticket.my_latest_ra_id = my_latest.id if my_latest else None

        latest_ra_user_ids = (
            {u.id for u in latest_ra.new_assigned_users.all()}
            if latest_ra
            else set()
        )

        if latest_ra and (
            user.is_superuser or
            ticket.created_by_id == user.id or
            user.id in latest_ra_user_ids
        ):
            ticket.best_ra_id = latest_ra.id
        elif my_latest:
            ticket.best_ra_id = my_latest.id
        else:
            ticket.best_ra_id = None

        all_assignees = summary["all_assignees"]
        received_user_ids = summary["received_user_ids"]

        if user.is_superuser or ticket.created_by_id == user.id:
            ticket.current_assignees = all_assignees
            expected_count = len(all_assignees)
            received_count = len(received_user_ids)
        else:
            ticket.current_assignees = [
                u for u in all_assignees
                if u.id == user.id
            ]
            expected_count = 1 if ticket.current_assignees else 0
            received_count = 1 if summary["my_answered"] else 0

        if user.is_superuser or ticket.created_by_id == user.id:
            if expected_count > 0 and received_count >= expected_count:
                ticket.answer_status = "Received"
            elif received_count > 0:
                ticket.answer_status = f"Partial Received ({received_count}/{expected_count})"
            elif ticket.has_action_taken_db:
                ticket.answer_status = "Action Taken"
            else:
                ticket.answer_status = "Pending"
        else:
            if summary["my_answered"]:
                ticket.answer_status = "Replied"
            elif ticket.has_action_taken_db:
                ticket.answer_status = "Action Taken"
            else:
                ticket.answer_status = "Pending"

        sib_units = []
        for assignee in ticket.current_assignees:
            unit = (getattr(assignee, "unit", "") or "").strip()
            if unit and unit not in sib_units:
                sib_units.append(unit)

        ticket.sib_units = sib_units
        ticket.sib_unit_display = ", ".join(sib_units) if sib_units else "-"

        normal_reply_date = (
            TicketAnswer.objects
            .filter(ticket=ticket, is_active=True, is_final=True)
            .exclude(answer__isnull=True)
            .exclude(answer__exact="")
            .aggregate(latest=Max("updated_at"))
            ["latest"]
        )

        reassign_reply_date = (
            ReassignmentAnswer.objects
            .filter(
                reassignment__ticket=ticket,
                reassignment__is_final=True,
                is_active=True,
                is_final=True,
            )
            .exclude(answer__isnull=True)
            .exclude(answer__exact="")
            .aggregate(latest=Max("created_at"))
            ["latest"]
        )

        ticket.reply_date = max(
            [d for d in [normal_reply_date, reassign_reply_date] if d],
            default=None,
        )

        ticket.has_action_taken = bool(ticket.has_action_taken_db)
        ticket.admin_can_modify = user.is_superuser and can_admin_modify_ticket(ticket)

        creator_action_taken = (
            ticket.action_traces
            .filter(
                action_key="CREATOR_ACTION_TAKEN",
                is_success=True,
            )
            .order_by("-performed_at")
            .first()
        )

        ticket.has_action_taken_by_creator = bool(creator_action_taken)

    states = list(State.objects.order_by("id").values("id", "name"))
    questions = list(Question.objects.filter(status="active").values("id", "text"))

    ticketsDash = dashboard_base_qs.distinct()

    total_tickets = ticketsDash.count()
    open_tickets = ticketsDash.filter(status="open").count()
    pending_tickets = ticketsDash.filter(status__in=PENDING_STATUSES).count()
    closed_tickets = ticketsDash.filter(status="closed").count()
    reopened_tickets = ticketsDash.filter(status="reopened").count()
    reassigned_tickets = ticketsDash.filter(
        reassignments__isnull=False
    ).distinct().count()

    assoc_states = list(
        OtherServicesDocumentsLink.objects.using("secondary")
        .exclude(association_state__isnull=True)
        .exclude(association_state__exact="")
        .values_list("association_state", flat=True)
        .distinct()
        .order_by("association_state")
    )

    a_ngo_new, a_ngo_pending, a_ngo_closed = section_counts(ticketsDash, "3A", "NGO")
    a_ob_new, a_ob_pending, a_ob_closed = section_counts(ticketsDash, "3A", "OFFICE_BEARER")
    b_ngo_new, b_ngo_pending, b_ngo_closed = section_counts(ticketsDash, "3B", "NGO")
    b_ob_new, b_ob_pending, b_ob_closed = section_counts(ticketsDash, "3B", "OFFICE_BEARER")
    c_ngo_new, c_ngo_pending, c_ngo_closed = section_counts(ticketsDash, "3C", "NGO")
    c_ob_new, c_ob_pending, c_ob_closed = section_counts(ticketsDash, "3C", "OFFICE_BEARER")
    e_ngo_new, e_ngo_pending, e_ngo_closed = section_counts(ticketsDash, "6E", "NGO")
    e_ob_new, e_ob_pending, e_ob_closed = section_counts(ticketsDash, "6E", "OFFICE_BEARER")
    s_ngo_new, s_ngo_pending, s_ngo_closed = section_counts(ticketsDash, "7", "NGO")
    s_ob_new, s_ob_pending, s_ob_closed = section_counts(ticketsDash, "7", "OFFICE_BEARER")

    if user.is_superuser:
        created_qs = ticketsDash
    else:
        created_qs = ticketsDash.filter(created_by=user).distinct()

    created_rows = list(created_qs.values("id", "status"))

    created_reply_index = build_ticket_reply_index(
        [row["id"] for row in created_rows],
        current_user_id=user.id,
    )

    created_total = len(created_rows)
    created_pending_total = 0
    created_received_total = 0
    created_received_partial = 0
    created_reopened = 0
    created_closed = 0

    for row in created_rows:
        ticket_id = row["id"]
        status = row["status"]
        summary = created_reply_index.get(ticket_id, {})

        if status == "closed":
            created_closed += 1
            continue

        if status == "reopened":
            created_reopened += 1

        if summary.get("is_received"):
            created_received_total += 1
        elif summary.get("is_partial_received"):
            created_received_partial += 1
        else:
            created_pending_total += 1

    if user.is_superuser:
        my_qs = ticketsDash
    else:
        my_qs = ticketsDash.filter(
            Q(assigned_users=user) |
            Q(reassignments__new_assigned_users=user)
        ).distinct()

    my_rows = list(my_qs.values("id", "status"))

    my_reply_index = build_ticket_reply_index(
        [row["id"] for row in my_rows],
        current_user_id=user.id,
    )

    my_action_taken_ids = set(
        TicketActionTrace.objects.filter(
            ticket_id__in=[row["id"] for row in my_rows],
            action_key="ACTION_TAKEN",
            is_success=True,
            performed_by=user,
        ).values_list("ticket_id", flat=True)
    )

    my_total = len(my_rows)
    my_new_total = 0
    my_pending_total = 0
    my_reopened = 0
    my_closed = 0

    for row in my_rows:
        ticket_id = row["id"]
        status = row["status"]
        summary = my_reply_index.get(ticket_id, {})

        if status == "closed":
            my_closed += 1
            continue

        if status == "reopened":
            my_reopened += 1

        has_my_action_taken = ticket_id in my_action_taken_ids

        if summary.get("my_answered") or has_my_action_taken:
            my_pending_total += 1
        else:
            my_new_total += 1

    created_pending_not_answered = 0
    created_pending_partial = 0
    created_received_complete = created_received_total
    created_reopened_with_leg = 0
    created_reopened_without_leg = 0

    my_pending_no_action = my_new_total
    my_pending_action_taken = my_pending_total
    my_replied_total = my_pending_total
    my_replied_complete = 0
    my_replied_partial = 0
    my_reopened_with_leg = 0
    my_reopened_without_leg = 0

    show_created_dashboard = user.is_superuser or Ticket.objects.filter(
        created_by=user,
        is_deleted=False,
    ).exists()

    show_assignee_dashboard = (not user.is_superuser) and Ticket.objects.filter(
        Q(assigned_users=user) |
        Q(reassignments__new_assigned_users=user),
        is_deleted=False,
    ).exists()

    show_clear_dashboard = any([
        request.GET.get("scope"),
        request.GET.get("created_bucket"),
        request.GET.get("my_bucket"),
        request.GET.get("cat"),
        request.GET.get("sec"),
        request.GET.get("cat_status"),
    ])

    return render(request, "tickets/ticket_list.html", {
        "tickets": tickets_page,
        "states": states,
        "questions": json.dumps(questions),

        "assoc_states": assoc_states,
        "assoc_state": state_filter,

        "total_tickets": total_tickets,
        "open_tickets": open_tickets,
        "pending_tickets": pending_tickets,
        "closed_tickets": closed_tickets,
        "reopened_tickets": reopened_tickets,
        "reassigned_tickets": reassigned_tickets,
        "received_tickets": created_received_total,

        "search_query": search_query,
        "status_filter": status_filter,
        "module_filter": module_filter,
        "from_date": from_date_raw,
        "to_date": to_date_raw,

        "cat": category_filter,
        "sec": section_filter,
        "cat_status": cat_status,

        "page_size": page_size,
        "allowed_sizes": allowed_sizes,
        "preserved_qs": preserved_qs,
        "base_filter_qs": base_filter_qs,

        "page": "ticket_list",
        "create_ticket": can_create_ticket(request.user),
        "NGO_BACKEND_URL": getattr(settings, "NGO_BACKEND_URL", ""),
        "user": user,

        "a_ngo_new": a_ngo_new,
        "a_ngo_pending": a_ngo_pending,
        "a_ngo_closed": a_ngo_closed,
        "a_ob_new": a_ob_new,
        "a_ob_pending": a_ob_pending,
        "a_ob_closed": a_ob_closed,

        "b_ngo_new": b_ngo_new,
        "b_ngo_pending": b_ngo_pending,
        "b_ngo_closed": b_ngo_closed,
        "b_ob_new": b_ob_new,
        "b_ob_pending": b_ob_pending,
        "b_ob_closed": b_ob_closed,

        "c_ngo_new": c_ngo_new,
        "c_ngo_pending": c_ngo_pending,
        "c_ngo_closed": c_ngo_closed,
        "c_ob_new": c_ob_new,
        "c_ob_pending": c_ob_pending,
        "c_ob_closed": c_ob_closed,

        "e_ngo_new": e_ngo_new,
        "e_ngo_pending": e_ngo_pending,
        "e_ngo_closed": e_ngo_closed,
        "e_ob_new": e_ob_new,
        "e_ob_pending": e_ob_pending,
        "e_ob_closed": e_ob_closed,

        "s_ngo_new": s_ngo_new,
        "s_ngo_pending": s_ngo_pending,
        "s_ngo_closed": s_ngo_closed,
        "s_ob_new": s_ob_new,
        "s_ob_pending": s_ob_pending,
        "s_ob_closed": s_ob_closed,

        "created_total": created_total,
        "created_pending_total": created_pending_total,
        "created_received_total": created_received_total,
        "created_received_partial": created_received_partial,
        "created_reopened": created_reopened,
        "created_closed": created_closed,

        "created_pending_not_answered": created_pending_not_answered,
        "created_pending_partial": created_pending_partial,
        "created_received_complete": created_received_complete,
        "created_reopened_with_leg": created_reopened_with_leg,
        "created_reopened_without_leg": created_reopened_without_leg,

        "my_total": my_total,
        "my_new_total": my_new_total,
        "my_pending_total": my_pending_total,
        "my_reopened": my_reopened,
        "my_closed": my_closed,

        "my_pending_no_action": my_pending_no_action,
        "my_pending_action_taken": my_pending_action_taken,
        "my_replied_total": my_replied_total,
        "my_replied_complete": my_replied_complete,
        "my_replied_partial": my_replied_partial,
        "my_reopened_with_leg": my_reopened_with_leg,
        "my_reopened_without_leg": my_reopened_without_leg,

        "show_creator_dashboard": show_created_dashboard,
        "show_assignee_dashboard": show_assignee_dashboard,
        "show_clear_dashboard": show_clear_dashboard,
    })


def user_display(u):
    """Return a nice display name for a user, handling method/property cases."""
    if not u:
        return None
    val = getattr(u, "get_full_name", None)
    if callable(val):                             # classic Django method
        name = (val() or "").strip()
        if name:
            return name
    elif isinstance(val, str):                    # if you ever make it a property/field
        name = val.strip()
        if name:
            return name
    # fallback to first + last, then username
    first = getattr(u, "first_name", "") or ""
    last = getattr(u, "last_name", "") or ""
    full = f"{first} {last}".strip()
    return full or getattr(u, "username", None)


import os
from pathlib import Path

def _safe_rel_folder(folder: str) -> str:
    """
    Convert '/2025/07/05/01/0013882025/' -> '2025/07/05/01/0013882025'
    and prevent path traversal.
    """
    folder = (folder or "").replace("\\", "/").strip()
    folder = folder.lstrip("/").rstrip("/")

    # prevent ../ traversal
    parts = [p for p in folder.split("/") if p and p not in (".", "..")]
    return "/".join(parts)

# ✅ Add these imports at top of views.py (only once)


def build_assignment_context(ticket, viewer, is_creator=False, is_superuser=False):
    """
    Returns dict for template:
    - creator view: ngo_assignees + office_member_assignments (member -> users)
    - assignee view: my_ngo_assigned + my_office_members (only own members)
    """
    qs = (
        AssignUsersCategory.objects
        .filter(ticket=ticket)
        .select_related("user", "category", "member_row")
        .prefetch_related("questions")
        .order_by("section_type", "member_row_id", "user_id")
    )

    # ------------------------------
    # CREATOR / SUPERUSER: see all
    # ------------------------------
    if is_creator or is_superuser:
        ngo_assignees_set = set()
        office_member_map = defaultdict(lambda: {"member": None, "users": []})

        for a in qs:
            if a.section_type == "NGO":
                ngo_assignees_set.add(a.user)

            elif a.section_type == "OFFICE_BEARER":
                # member_row is mandatory for office bearer mapping
                if a.member_row_id:
                    office_member_map[a.member_row_id]["member"] = a.member_row
                    office_member_map[a.member_row_id]["users"].append(a.user)

        # format for template
        ngo_assignees = sorted(
            list(ngo_assignees_set),
            key=lambda u: ((u.first_name or "").lower(), (u.last_name or "").lower(), u.id)
        )

        office_member_assignments = []
        for v in office_member_map.values():
            # sort users per member row
            v["users"] = sorted(
                v["users"],
                key=lambda u: ((u.first_name or "").lower(), (u.last_name or "").lower(), u.id)
            )
            office_member_assignments.append(v)

        # sort rows by member name
        office_member_assignments.sort(
            key=lambda x: ((getattr(x["member"], "member_name", "") or "").lower(), x["member"].id if x["member"] else 0)
        )

        # Optional: overall assignees list (unique) from AssignUsersCategory (better than Ticket.assigned_users)
        all_assignees = sorted(
            list({a.user for a in qs if a.user}),
            key=lambda u: ((u.first_name or "").lower(), (u.last_name or "").lower(), u.id)
        )

        return {
            "all_assignees": all_assignees,
            "ngo_assignees": ngo_assignees,
            "office_member_assignments": office_member_assignments,
            # assignee-only keys empty
            "my_ngo_assigned": False,
            "my_office_members": [],
        }

    # ------------------------------
    # ASSIGNEE: only own assignments
    # ------------------------------
    qs = qs.filter(user=viewer)

    my_ngo_assigned = qs.filter(section_type="NGO").exists()

    # Only OFFICE_BEARER member rows
    my_members_qs = (
        qs.filter(section_type="OFFICE_BEARER", member_row__isnull=False)
          .values_list("member_row_id", flat=True)
          .distinct()
    )

    my_office_members = list(
        TicketMemberRow.objects.filter(id__in=my_members_qs).order_by("id")
    )

    return {
        "all_assignees": [],  # assignee doesn't need global list
        "ngo_assignees": [],
        "office_member_assignments": [],
        "my_ngo_assigned": my_ngo_assigned,
        "my_office_members": my_office_members,
    }

# ✅ COMPLETE WORKING CODE: list + preview + download OSDL docs using
# OtherServicesDocumentsLink(id=ticket.models_id) + DOCROOT + registration_applications_documents_location

# ---------------------------
# Add these imports (top of views.py)
# ---------------------------


# ---------------------------
# Safe helpers
# ---------------------------
def _safe_rel_folder(folder: str) -> str:
    folder = (folder or "").replace("\\", "/").strip()
    folder = folder.lstrip("/").rstrip("/")
    parts = [p for p in folder.split("/") if p and p not in (".", "..")]
    return "/".join(parts)

def _safe_filename(name: str) -> str:
    name = (name or "").replace("\\", "/").strip()
    if "/" in name or name in ("", ".", ".."):
        return ""
    return name

# ---------------------------
# ✅ Preview/Download endpoint (inline preview by default)
# URL: /tickets/<ticket_id>/osdl-docs/<filename>/
# Add ?download=1 to force download
# ---------------------------
# @login_required(login_url="login")
# from pathlib import Path

def build_folder_structure(docs_list):
    """
    Convert flat docs list into nested folder structure for tree view
    """
    folder_structure = {}
    
    for doc in docs_list:
        folder_path = doc.get('folder_path', '/')
        if folder_path == '/':
            folder_key = 'root'
        else:
            # Remove leading slash and split
            parts = folder_path.lstrip('/').split('/')
            folder_key = '/'.join(parts)
        
        if folder_key not in folder_structure:
            folder_structure[folder_key] = {
                'name': parts[-1] if folder_key != 'root' else 'Root',
                'path': folder_path,
                'files': []
            }
        
        folder_structure[folder_key]['files'].append(doc)
    
    # Convert to list and sort
    result = []
    for key, value in folder_structure.items():
        if key != 'root':
            result.append(value)
    
    # Sort folders by name
    result.sort(key=lambda x: x['name'])
    
    # Add root at the beginning if it has files
    if 'root' in folder_structure and folder_structure['root']['files']:
        result.insert(0, folder_structure['root'])
    
    return result
@login_required(login_url="login")
def osdl_doc_serve(request, ticket_id, filename):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    user = request.user

    # ✅ permissions (adjust if you want assignees too)
    is_creator = (ticket.created_by_id == user.id)
    # is_super = user.is_superuser
    # if not (is_creator or is_super):
    #     return HttpResponseForbidden("Not allowed")

    # ✅ must be NGO ticket (optional but recommended)
    if not ticket.models_name or ticket.models_name.lower() != "ngo":
        raise Http404("No NGO docs")

    # ✅ this is exactly what you asked for
    osdl = (
        OtherServicesDocumentsLink.objects.using("secondary")
        .filter(id=ticket.models_id)
        .first()
    )
    if not osdl:
        raise Http404("OSDL row not found (models_id mismatch)")

    raw_folder = getattr(osdl, "registration_applications_documents_location", None)
    if not raw_folder:
        raise Http404("OSDL docs folder is empty")

    docs_folder = _safe_rel_folder(raw_folder)
    filename = _safe_filename(filename)
    if not filename:
        return HttpResponseForbidden("Invalid filename")

    docroot = getattr(settings, "DOCROOT", None)
    if not docroot:
        raise Http404("DOCROOT not configured")

    # Check if folder parameter is provided (for files in subfolders)
    folder_param = request.GET.get("folder", "")
    if folder_param and folder_param != '.':
        # Construct path including subfolder
        subfolder = _safe_rel_folder(folder_param)
        base_dir = (Path(docroot) / docs_folder / subfolder).resolve()
    else:
        base_dir = (Path(docroot) / docs_folder).resolve()

    file_path = (base_dir / filename).resolve()

    # ✅ block traversal - check against the main docs folder
    main_docs_dir = (Path(docroot) / docs_folder).resolve()
    if main_docs_dir not in file_path.parents:
        return HttpResponseForbidden("Invalid path")

    if not file_path.exists() or not file_path.is_file():
        raise Http404("File not found")

    content_type, _ = mimetypes.guess_type(str(file_path))
    response = FileResponse(
        open(file_path, "rb"),
        content_type=content_type or "application/octet-stream"
    )

    force_download = request.GET.get("download") == "1"
    disposition = "attachment" if force_download else "inline"
    response["Content-Disposition"] = f'{disposition}; filename="{filename}"'
    return response




def build_osdl_docs_list_for_ticket(ticket, viewer):
    docs_from_osdl = []

    # ----------------------------------------------------
    # ✅ PERMISSION: creator OR superuser OR assignee
    # (includes reassignment assignees also)
    # ----------------------------------------------------
    is_creator = (ticket.created_by_id == viewer.id)
    is_super = getattr(viewer, "is_superuser", False)

    is_assignee = ticket.assigned_users.filter(id=viewer.id).exists()

    # If you want reassignment-assignees also:
    is_reassign_assignee = TicketReassignment.objects.filter(
        ticket=ticket,
        new_assigned_users=viewer
    ).exists()

    if not (is_creator or is_super or is_assignee or is_reassign_assignee):
        return docs_from_osdl

    # ----------------------------------------------------
    # ✅ only NGO tickets
    # ----------------------------------------------------
    if not ticket.models_name or ticket.models_name.lower() != "ngo":
        return docs_from_osdl

    # ----------------------------------------------------
    # ✅ read folder path from OSDL table (secondary DB)
    # ----------------------------------------------------
    osdl = (
        OtherServicesDocumentsLink.objects.using("secondary")
        .filter(id=ticket.models_id)
        .first()
    )
    if not osdl:
        return docs_from_osdl

    raw_folder = getattr(osdl, "registration_applications_documents_location", None)
    if not raw_folder:
        return docs_from_osdl

    docs_folder = _safe_rel_folder(raw_folder)

    docroot = getattr(settings, "DOCROOT", None)
    if not docroot or not docs_folder:
        return docs_from_osdl

    base_path = Path(docroot) / docs_folder
    if not base_path.exists() or not base_path.is_dir():
        return docs_from_osdl

    # ----------------------------------------------------
    # ✅ RECURSIVE FILE LISTING (including subfolders)
    # ----------------------------------------------------
    allowed_ext = (".pdf", ".jpg", ".jpeg", ".png", ".doc", ".docx", ".xls", ".xlsx")

    # Walk through all directories recursively
    for root, dirs, files in os.walk(base_path):
        # Get relative path from base_path
        rel_path = os.path.relpath(root, base_path)
        if rel_path == '.':
            folder_path = '/'
        else:
            folder_path = '/' + rel_path.replace(os.sep, '/')

        for file in sorted(files):
            if file.lower().endswith(allowed_ext):
                file_path = os.path.join(root, file)
                
                # Get file size
                try:
                    file_size = os.path.getsize(file_path)
                except:
                    file_size = 0

                docs_from_osdl.append({
                    "name": file,
                    "folder_path": folder_path,
                    "size": file_size,
                    "size_display": _format_file_size(file_size),
                    "preview_url": reverse(
                        "tickets:osdl_doc_serve",
                        kwargs={"ticket_id": ticket.id, "filename": file},
                    ) + f"?folder={rel_path}" if rel_path != '.' else reverse(
                        "tickets:osdl_doc_serve",
                        kwargs={"ticket_id": ticket.id, "filename": file},
                    ),
                    "download_url": reverse(
                        "tickets:osdl_doc_serve",
                        kwargs={"ticket_id": ticket.id, "filename": file},
                    ) + f"?download=1&folder={rel_path}" if rel_path != '.' else reverse(
                        "tickets:osdl_doc_serve",
                        kwargs={"ticket_id": ticket.id, "filename": file},
                    ) + "?download=1",
                    "full_path": os.path.join(rel_path, file) if rel_path != '.' else file,
                })

    return docs_from_osdl

def _format_file_size(size_bytes):
    """Convert bytes to human readable format"""
    if size_bytes == 0:
        return "0 B"
    size_names = ["B", "KB", "MB", "GB"]
    i = 0
    while size_bytes >= 1024 and i < len(size_names) - 1:
        size_bytes /= 1024.0
        i += 1
    return f"{size_bytes:.1f} {size_names[i]}" 
@login_required(login_url='login')
def ticket_detail(request, pk):
    ticket = (
        Ticket.objects
        .select_related("created_by")
        .prefetch_related("assigned_users")
        .get(pk=pk)
    )

    user = request.user
    User = get_user_model()
    creator = ticket.created_by
    
    userData = creator
    is_creator = (user.id == ticket.created_by_id)
    is_assignee = ticket.assigned_users.filter(id=user.id).exists()
    show_all_replies = (
        is_creator
        or user.is_superuser
        or (ticket.status and ticket.status.lower() == "closed")
    )

    status_history = TicketStatusHistory.objects.none()
    timeline_events = []

    # 🔹 Does this user have any final answers at all? (for display/labels)
    user_has_final = TicketAnswer.objects.filter(
        ticket=ticket,
        answered_by=user,
        is_final=True
    ).exists()

    # 🔹 Does this user have at least one NON-FINAL answer (latest version after reopen)?
    has_non_final = TicketAnswer.objects.filter(
        ticket=ticket,
        answered_by=user,
        is_final=False
    ).exists()

    # 🔹 Can this user still edit answers?
    can_edit_ticket = (
        (is_assignee and has_non_final and ticket.status != "closed")
        or user.is_superuser
    )

    # ======================================================
    # ================== APPLICATION DOCS / OSDL (NGO) =====
    # ======================================================
    ngo_data = None
    osdl_parent = None
    osdl_docs = []
    osdl_logs = []
    osdl_docs_count = 0
    osdl_logs_count = 0
    ngoRedirectURL = "#"

    try:
        if ticket.models_name and ticket.models_name.lower() == "ngo" and ticket.models_id:
            osdl_parent = (
                OtherServicesDocumentsLink.objects.using('secondary')
                .filter(id=ticket.models_id)
                # .values()
                .first()
            )
            # osdl_parent = OtherServicesDocumentsLink.objects.using("secondary").get(id=ticket.models_id)
            # print(osdl_parent)

            if osdl_parent:
                osdl_docs = (
                    TicketApplicationDoc.objects.using('secondary')
                    .filter(ticket_id=ticket.id)
                    .order_by('-uploaded_at')
                )
                osdl_docs_count = osdl_docs.count()

                osdl_logs = list(
                    OtherServicesDocumentsLink_logs.objects.using('secondary')
                    .filter(other_services_documents_link_id=osdl_parent.id)
                    .order_by('-id')
                )
                osdl_logs_count = len(osdl_logs)

                for lg in osdl_logs:
                    rcn = (
                        getattr(lg, 'rcn', None)
                        or getattr(osdl_parent, 'rcn', '')
                        or '-'
                    )
                    pan = (
                        getattr(lg, 'pan_number', None)
                        or getattr(lg, 'pan', None)
                        or getattr(osdl_parent, 'pan_number', '')
                        or getattr(osdl_parent, 'pan', '')
                        or '-'
                    )
                    association_name = (
                        getattr(lg, 'association_name', None)
                        or getattr(osdl_parent, 'association_name', '')
                        or '-'
                    )

                    setattr(lg, 'old_subset', {
                        'rcn': rcn,
                        'pan': pan,
                        'association_name': association_name,
                    })

    except Exception:
        osdl_parent = None
        osdl_docs = []
        osdl_logs = []
        osdl_docs_count = 0
        osdl_logs_count = 0
    # print(ticket)
    base = settings.NGO_BACKEND_URL.rstrip('/')
    if getattr(ticket, "rcn", None):
        ngoRedirectURL = f"{base}/ngo/details/{quote(str(ticket.rcn))}/"
    elif getattr(ticket, "models_id", None):
        ngoRedirectURL = f"{base}/ngo/applicationdetails/{quote(str(ticket.models_id))}/"
    else:
        ngoRedirectURL = "#"

    # ---------------------------------------
    # REASSIGNMENT LENS
    # ---------------------------------------
    reassign_id = request.GET.get("reassign")
    reassignment = None
    if reassign_id:
        try:
            reassign_id = int(reassign_id)
            reassignment = get_object_or_404(
                TicketReassignment.objects
                .prefetch_related("answers__question", "new_assigned_users", "reassigned_by"),
                id=reassign_id,
                ticket=ticket
            )
        except (ValueError, TypeError):
            reassignment = None

    # ========= REPORT SUMMARIES VISIBLE IN UI =========
    # Creator → sees all reports (ticket + all reassignments, all users)
    # Others → see only their own reports
    if is_creator:
        report_summaries = (
            ReportSummary.objects
            .filter(ticket=ticket)
            .select_related("user", "reassignment")
            .order_by("-created_at")
        )
    else:
        if reassignment:
            # In reassignment view: show this user's reports for this reassignment only
            report_summaries = (
                ReportSummary.objects
                .filter(ticket=ticket, reassignment=reassignment, user=user)
                .select_related("user", "reassignment")
                .order_by("-created_at")
            )
        else:
            # Normal ticket view: show this user's reports for this ticket
            report_summaries = (
                ReportSummary.objects
                .filter(ticket=ticket, user=user)
                .select_related("user", "reassignment")
                .order_by("-created_at")
            )

    # ================= CURRENT USER'S SINGLE REPORT (for form pre-fill) =================
    if reassignment:
        # Try reassignment-specific summary first
        my_report = (
            ReportSummary.objects
            .filter(ticket=ticket, reassignment=reassignment, user=user)
            .order_by("-id")
            .first()
        )
        # Fallback to ticket-level summary if nothing saved yet for this reassignment
        if not my_report:
            my_report = (
                ReportSummary.objects
                .filter(ticket=ticket, reassignment__isnull=True, user=user)
                .order_by("-id")
                .first()
            )
    else:
        # Normal ticket view → just ticket-level summary
        my_report = (
            ReportSummary.objects
            .filter(ticket=ticket, reassignment__isnull=True, user=user)
            .order_by("-id")
            .first()
        )

    my_report_message = my_report.summary if (my_report and my_report.summary) else ""
    my_report_file = my_report.file if (my_report and my_report.file) else None

    # ======================================================
    #                MODE: REASSIGNMENT VIEW
    # ======================================================
    if reassignment:
        is_reassign_owner = (user.id == reassignment.reassigned_by_id)
        is_reassign_assignee = reassignment.new_assigned_users.filter(id=user.id).exists()

        allowed_user_ids = {creator.id, *reassignment.new_assigned_users.values_list('id', flat=True)}
        if user.id not in allowed_user_ids and not is_reassign_owner and not user.is_superuser:
            raise PermissionDenied()

        ticket_q_ids = list(
            TicketAnswer.objects
            .filter(ticket=ticket)
            .values_list("question_id", flat=True)
            .distinct()
        )

        if ticket_q_ids:
            questions = list(Question.objects.filter(id__in=ticket_q_ids).order_by("id"))
        else:
            questions = list(Question.objects.filter(status="active").order_by("id"))

        ra_map = {}
        for ra in (
            ReassignmentAnswer.objects
            .filter(reassignment=reassignment)
            .select_related("question", "answered_by")
        ):
            ra_map[ra.question_id] = {
                "answer": ra.answer or "",
                "answered_by": user_display(ra.answered_by)
            }

        answers = [
            SimpleNamespace(
                question=q,
                answer=ra_map.get(q.id, {}).get("answer", ""),
                answered_by=ra_map.get(q.id, {}).get("answered_by")
            )
            for q in questions
        ]

        logs = (
            TicketLog.objects
            .filter(ticket=ticket, ticketReassignment=reassignment)
            .select_related("user")
            .order_by("-created_at")
        )

        can_edit_reassign = (not reassignment.is_final) and (is_reassign_assignee or user.is_superuser)

    # ======================================================
    #                MODE: NORMAL TICKET VIEW
    # ======================================================
    else:
        # if is_creator:
        if show_all_replies:
            qa_qs = (
                TicketAnswer.objects
                .filter(ticket=ticket, is_final=True)
                .exclude(answer__isnull=True)
                .exclude(answer__exact="")
                .select_related("question", "answered_by")
            )
            answers = [
                SimpleNamespace(
                    question=ta.question,
                    answer=ta.answer or "",
                    answered_by=user_display(ta.answered_by)
                )
                for ta in qa_qs
            ]

        elif is_assignee:
            from collections import defaultdict

            base_qs = (
                TicketAnswer.objects
                .filter(ticket=ticket, answered_by=user)
                .select_related("question")
                .order_by("question__id", "version")
            )

            grouped = defaultdict(list)
            for ta in base_qs:
                grouped[ta.question_id].append(ta)

            questions = Question.objects.filter(
                id__in=grouped.keys(),
                status="active"
            ).order_by("id")

            answers = []
            for q in questions:
                versions = grouped[q.id]
                latest_ta = versions[-1]
                old_versions = versions[:-1]

                answers.append(
                    SimpleNamespace(
                        question=q,
                        answer=(latest_ta.answer or ""),
                        answered_by=user_display(latest_ta.answered_by) if latest_ta.answered_by else None,
                        versions=versions,
                        old_versions=old_versions,
                        updated_at=latest_ta.updated_at or latest_ta.created_at,
                    )
                )
        else:
            answers = []

        base_logs_qs = (
            TicketLog.objects
            .filter(ticket=ticket, ticketReassignment__isnull=True)
            .select_related("user")
            .order_by("-created_at")
        )

        if is_creator:
            logs = base_logs_qs
        elif is_assignee:
            logs = base_logs_qs.filter(user_id__in=[creator.id, user.id])
        else:
            logs = TicketLog.objects.none()

        can_edit_reassign = False

    # ======================================================
    # CREATOR: USER-WISE Q&A TABS (only FINAL answers)
    # ======================================================
    qa_by_user = []
    active_user_id = None

    # if is_creator:
    if show_all_replies:
        from collections import defaultdict
        user_map = defaultdict(lambda: {"user": None, "rows": []})
        latest_ts = None

        t_answers = (
            TicketAnswer.objects
            .filter(ticket=ticket, is_final=True)
            .exclude(answer__isnull=True)
            .exclude(answer__exact="")
            .select_related("question", "answered_by")
            .order_by("answered_by_id", "question__id", "-version")
        )

        seen = set()

        for ta in t_answers:
            if not ta.answered_by:
                continue

            key = (ta.answered_by_id, ta.question_id)
            if key in seen:
                continue
            seen.add(key)

            u = ta.answered_by
            user_map[u.id]["user"] = u

            previous_qs = (
                TicketAnswer.objects
                .filter(
                    ticket=ticket,
                    answered_by=u,
                    question=ta.question,
                    is_final=True,
                    version__lt=(ta.version or 1),
                )
                .exclude(answer__isnull=True)
                .exclude(answer__exact="")
                .order_by("-version")
            )

            user_map[u.id]["rows"].append({
                "question": ta.question.text,
                "answer": ta.answer,
                "source": "Ticket",
                "previous": [
                    {
                        "answer": p.answer,
                        "is_final": p.is_final,
                        "version": p.version,
                        "updated_at": p.updated_at,
                        "created_at": p.created_at,
                    }
                    for p in previous_qs
                ],
                "updated_at": ta.updated_at,
                "created_at": ta.created_at,
            })

            if ta.updated_at:
                ts = make_naive(ta.updated_at)
                if latest_ts is None or ts > latest_ts:
                    latest_ts = ts
                    active_user_id = u.id

        ra_answers = (
            ReassignmentAnswer.objects
            .filter(reassignment__ticket=ticket, reassignment__is_final=True)
            .exclude(answer__isnull=True)
            .exclude(answer__exact="")
            .select_related("question", "answered_by", "reassignment")
            .order_by("reassignment_id", "question__id")
        )

        for ra in ra_answers:
            if not ra.answered_by:
                continue

            u = ra.answered_by
            user_map[u.id]["user"] = u
            user_map[u.id]["rows"].append({
                "question": ra.question.text,
                "answer": ra.answer,
                "source": f"Reassign #{ra.reassignment_id}",
                "previous": [],
                "updated_at": ra.created_at,
                "created_at": ra.created_at,
            })

            if ra.created_at:
                ts = make_naive(ra.created_at)
                if latest_ts is None or ts > latest_ts:
                    latest_ts = ts
                    active_user_id = u.id

        # qa_by_user = [
        #     {"user": v["user"], "rows": v["rows"]}
        #     for _, v in user_map.items()
        #     if v["user"] and v["rows"]
        # ]

        qa_by_user = []
        for _, v in user_map.items():
            if not v["user"] or not v["rows"]:
                continue

            u = v["user"]
            rows = v["rows"]

            report_obj = None

            # take source from the last row in this tab
            last_source = rows[-1].get("source", "") if rows else ""

            if last_source == "Ticket":
                report_obj = (
                    ReportSummary.objects
                    .filter(ticket=ticket, user=u, reassignment__isnull=True)
                    .order_by("-id")
                    .first()
                )

            elif last_source.startswith("Reassign #"):
                reassign_id = last_source.replace("Reassign #", "").strip()
                if reassign_id.isdigit():
                    report_obj = (
                        ReportSummary.objects
                        .filter(ticket=ticket, user=u, reassignment_id=int(reassign_id))
                        .order_by("-id")
                        .first()
                    )

            qa_by_user.append({
                "user": u,
                "rows": rows,
                "report_message": report_obj.summary if report_obj else "",
                "report_file": report_obj.file if report_obj and report_obj.file else None,
            })

    # ======================================================
    # FILES for legacy logs (logs_with_files)
    # ======================================================
    logs_with_files = []
    for log in logs:
        files = FileUpload.objects.filter(log=log)
        logs_with_files.append({"log": log, "files": files})

    latest_feedback = logs.first().message if logs else ""

    my_latest_log = (
        TicketLog.objects
        .filter(ticket=ticket, user=user, ticketReassignment__isnull=True)
        .order_by("-updated_at")
        .first()
    )
    feedback_text = (my_latest_log.message if (my_latest_log and not ticket.status == "closed" and not reassignment) else "")

    # ======================================================|
    # FEEDBACK CONVERSATIONS (creator <-> assignees) .       |
    # ======================================================|
    fb_by_user = []
    fb_active_user_id = None
    assignee_fb_rows = []

    if is_creator:
        latest_fb_ts = None
        # User = get_user_model()
        creator = ticket.created_by

        if qa_by_user:
            fb_users = [bundle["user"] for bundle in qa_by_user]
        else:
            fb_users = list(ticket.assigned_users.all())

        all_users_map = {u.id: u for u in fb_users if u}

        reassign_user_ids = (
            TicketReassignment.objects
            .filter(ticket=ticket)
            .values_list("new_assigned_users__id", flat=True)
            .distinct()
        )
        if reassign_user_ids:
            reassign_users = User.objects.filter(id__in=reassign_user_ids)
            for u in reassign_users:
                all_users_map[u.id] = u

        participant_ids = (
            TicketLog.objects
            .filter(ticket=ticket)
            .values_list("user_id", "to_user_id")
        )

        log_user_ids = set()
        for uid, to_uid in participant_ids:
            if uid and uid != creator.id:
                log_user_ids.add(uid)
            if to_uid and to_uid != creator.id:
                log_user_ids.add(to_uid)

        if log_user_ids:
            log_users = User.objects.filter(id__in=log_user_ids)
            for u in log_users:
                all_users_map[u.id] = u

        fb_users = list(all_users_map.values())

        for person in fb_users:
            if not person:
                continue

            conv_logs = (
                TicketLog.objects
                .filter(ticket=ticket)
                .filter(
                    Q(user_id=creator.id, to_user_id=person.id) |
                    Q(user_id=person.id, to_user_id=creator.id)
                )
                .select_related("user", "ticketReassignment")
                .order_by("created_at")
            )

            if not conv_logs.exists():
                fb_by_user.append({"user": person, "rows": []})
                continue

            log_ids = list(conv_logs.values_list("id", flat=True))
            files_qs = FileUpload.objects.filter(log_id__in=log_ids)

            from collections import defaultdict
            files_map = defaultdict(list)
            for f in files_qs:
                files_map[f.log_id].append(f)

            rows = []
            for lg in conv_logs:
                rows.append({
                    "message": (lg.message or "").strip(),
                    "author": lg.user,
                    "created_at": lg.created_at,
                    "source": (
                        "Ticket"
                        if lg.ticketReassignment_id is None
                        else f"Reassign #{lg.ticketReassignment_id}"
                    ),
                    "files": files_map.get(lg.id, []),
                })

                if lg.created_at:
                    ts = make_naive(lg.created_at)
                    if latest_fb_ts is None or ts > latest_fb_ts:
                        latest_fb_ts = ts
                        fb_active_user_id = person.id

            fb_by_user.append({"user": person, "rows": rows})

        fb_by_user.sort(
            key=lambda x: (
                (x["user"].first_name or "").lower(),
                (x["user"].last_name or "").lower()
            )
        )

    else:
        creator = ticket.created_by

        is_assignee = ticket.assigned_users.filter(id=user.id).exists()
        is_reassign_assignee_flag = TicketReassignment.objects.filter(
            ticket=ticket,
            new_assigned_users=user,
        ).exists()

        if is_assignee or is_reassign_assignee_flag or user.is_superuser:
            conv_logs = (
                TicketLog.objects
                .filter(ticket=ticket)
                .filter(
                    Q(user_id=user.id, to_user_id=creator.id) |
                    Q(user_id=creator.id, to_user_id=user.id)
                )
                .select_related("user", "ticketReassignment")
                .order_by("created_at")
            )

            log_ids = list(conv_logs.values_list("id", flat=True))
            files_qs = FileUpload.objects.filter(log_id__in=log_ids)

            from collections import defaultdict
            files_map = defaultdict(list)
            for f in files_qs:
                files_map[f.log_id].append(f)

            assignee_fb_rows = []
            for lg in conv_logs:
                assignee_fb_rows.append({
                    "message": (lg.message or "").strip(),
                    "author": lg.user,
                    "created_at": lg.created_at,
                    "source": (
                        "Ticket"
                        if lg.ticketReassignment_id is None
                        else f"Reassign #{lg.ticketReassignment_id}"
                    ),
                    "files": files_map.get(lg.id, []),
                })
        else:
            assignee_fb_rows = []

    # ======================================================
    # STATUS HISTORY (creator / reassignment / assignee rules)
    # Only change: visibility is user-wise
    # Creator/superuser -> all
    # Others -> only their own
    # ======================================================

    status_history = TicketStatusHistory.objects.none()

    if reassignment:
        # --- REASSIGNMENT LENS ---
        if is_creator or user.is_superuser:
            # creator sees full history even in reassignment view
            status_history = (
                TicketStatusHistory.objects
                .filter(ticket=ticket)
                .order_by("created_at", "id")   # chronological + stable
            )
        else:
            # normal reassignee sees only the window of this reassignment
            next_ra = (
                TicketReassignment.objects
                .filter(ticket=ticket, created_at__gt=reassignment.created_at)
                .order_by("created_at")
                .first()
            )

            window = {
                "ticket": ticket,
                "created_at__gte": reassignment.created_at,
            }
            if next_ra:
                window["created_at__lt"] = next_ra.created_at

            base_history = (
                TicketStatusHistory.objects
                .filter(**window)
                .order_by("created_at", "id")
            )

            # ✅ ONLY this reassignee’s actions (user-wise visibility)
            status_history = base_history.filter(changed_by=user)

    else:
        # --- NORMAL TICKET VIEW ---
        if is_creator or user.is_superuser:
            status_history = (
                TicketStatusHistory.objects
                .filter(ticket=ticket)
                .order_by("created_at", "id")
            )
        else:
            # Non-creator: show history only from the time they got involved
            latest_assign_event = (
                TicketReassignment.objects
                .filter(ticket=ticket, new_assigned_users=user)
                .order_by("-created_at")
                .first()
            )

            if latest_assign_event:
                start_ts = latest_assign_event.created_at
            elif is_assignee:
                start_ts = ticket.created_at
            else:
                start_ts = None

            if start_ts:
                next_event = (
                    TicketReassignment.objects
                    .filter(ticket=ticket, created_at__gt=start_ts)
                    .order_by("created_at")
                    .first()
                )

                window = {
                    "ticket": ticket,
                    "created_at__gte": start_ts,
                }
                if next_event:
                    window["created_at__lt"] = next_event.created_at

                base_history = (
                    TicketStatusHistory.objects
                    .filter(**window)
                    .order_by("created_at", "id")
                )

                # ✅ ONLY this user's actions (assignee-wise visibility)
                status_history = base_history.filter(changed_by=user)
            else:
                status_history = TicketStatusHistory.objects.none()

    # ======================================================
    # ALWAYS INCLUDE CLOSED (safety)
    # (only add it if it's missing in what user can see)
    # ======================================================

    history_list = list(status_history)

    if ticket.status and ticket.status.lower() == "closed":
        closed_qs = TicketStatusHistory.objects.filter(
            ticket=ticket,
            new_status__iexact="closed"
        ).order_by("created_at", "id")

        for ch in closed_qs:
            if ch not in history_list:
                # creator sees it; non-creator sees it ONLY if they changed it
                if is_creator or user.is_superuser or ch.changed_by_id == user.id:
                    history_list.append(ch)

    # final stable sort oldest -> newest
    history_list.sort(key=lambda h: (h.created_at or timezone.now(), h.id))
    status_history = history_list  # list (template-friendly)

    # ======================================================
    # TIMELINE (single source of truth = TicketStatusHistory)
    # ======================================================

    timeline_events = []
    for h in status_history:
        timeline_events.append(SimpleNamespace(
            kind="status",
            at=h.created_at,
            by=h.changed_by,
            old_status=h.old_status,
            new_status=h.new_status,
        ))

    # final sort (oldest -> newest)
    timeline_events.sort(key=lambda e: (e.at or timezone.now(),))


    OtherServicesDocumentsLinkData = (
                OtherServicesDocumentsLink.objects.using("secondary")
                .filter(id=ticket.models_id)
                .values()
                .first()
            )
    osdl = (
            OtherServicesDocumentsLink.objects.using("secondary")
            .filter(id=ticket.models_id)
            .first()
        )
    # print(osdl.rcn)
    app_id = osdl.id if osdl else None
    if app_id:
        committee_members = _pick_office_bearers(app_id,osdl)
        bank_name, bank_address, account_no = _pick_bank_details(app_id)
    else:
        committee_members = secretary = "-"
        bank_name = bank_address = account_no = "-"

    # print(committee_members)
    # --- Who is allowed to see everything? ---
    is_reassign_assignee_flag = TicketReassignment.objects.filter(
        ticket=ticket,
        new_assigned_users=user,
    ).exists()

    can_view_all_tabs = (
        is_creator
        or is_assignee
        or is_reassign_assignee_flag
        or user.is_superuser
    )
        
    docs_from_osdl = build_osdl_docs_list_for_ticket(ticket, user)

    adverse_reasons = AdverseReason.objects.filter(is_active=True).order_by("name")

    # ======================================================
    # ✅ ASSIGNMENT MAPPING (NGO + OFFICE_BEARER)
    # ======================================================
    assignment_ctx = build_assignment_context(
        ticket=ticket,
        viewer=user,
        is_creator=is_creator,
        is_superuser=user.is_superuser
    )

    # action_taken = (ticket.action_traces.filter(action_key="ACTION_TAKEN",performed_by=request.user).order_by("-performed_at").first())


    assignee_user_ids = set(ticket.assigned_users.values_list("id", flat=True))

    reassign_user_ids = set(TicketReassignment.objects.filter(ticket=ticket).values_list("new_assigned_users__id", flat=True))

    allowed_assignee_ids = assignee_user_ids | reassign_user_ids

    action_taken = (ticket.action_traces.filter(action_key="ACTION_TAKEN",is_success=True,performed_by_id__in=allowed_assignee_ids,).order_by("-performed_at").first())
    folder_structure = build_folder_structure(docs_from_osdl)
    total_files = len(docs_from_osdl)
    folder_count = len(folder_structure)
    signaters = Signater.objects.filter(status="active").order_by("id")
    my_signaters_to = TicketSignater.objects.filter(
            ticket=ticket,
            user=request.user,
            action_type="to"
        ).values_list("signater_id", flat=True).first()

    my_signaters_from = TicketSignater.objects.filter(
            ticket=ticket,
            user=request.user,
            action_type="from"
        ).values_list("signater_id", flat=True).first()

    # ======================================================
    # DOWNLOAD TAB: Reply PDF visibility
    # 1) Creator/superuser -> show only users who submitted FINAL
    # 2) Assignee -> show own reply PDF only if own FINAL exists
    # ======================================================
    reply_download_users = []
    can_download_reply_pdf = False

    if is_creator or user.is_superuser:
        added_user_ids = set()

        # Normal ticket final answers
        normal_final_user_ids = (
            TicketAnswer.objects
            .filter(ticket=ticket, is_final=True)
            .exclude(answer__isnull=True)
            .exclude(answer__exact="")
            .values_list("answered_by_id", flat=True)
            .distinct()
        )

        for uid in normal_final_user_ids:
            if uid and uid not in added_user_ids:
                u = User.objects.filter(id=uid).first()
                if u:
                    reply_download_users.append(u)
                    added_user_ids.add(uid)

        # Reassignment final answers
        reassign_final_user_ids = (
            ReassignmentAnswer.objects
            .filter(
                reassignment__ticket=ticket,
                reassignment__is_final=True,
            )
            .exclude(answer__isnull=True)
            .exclude(answer__exact="")
            .values_list("answered_by_id", flat=True)
            .distinct()
        )

        for uid in reassign_final_user_ids:
            if uid and uid not in added_user_ids:
                u = User.objects.filter(id=uid).first()
                if u:
                    reply_download_users.append(u)
                    added_user_ids.add(uid)

        reply_download_users.sort(
            key=lambda x: (
                (x.first_name or "").lower(),
                (x.last_name or "").lower(),
                x.id,
            )
        )

    else:
        my_has_final_ticket = (
            TicketAnswer.objects
            .filter(ticket=ticket, answered_by=user, is_final=True)
            .exclude(answer__isnull=True)
            .exclude(answer__exact="")
            .exists()
        )

        my_has_final_reassign = (
            ReassignmentAnswer.objects
            .filter(
                reassignment__ticket=ticket,
                answered_by=user,
                reassignment__is_final=True,
            )
            .exclude(answer__isnull=True)
            .exclude(answer__exact="")
            .exists()
        )

        can_download_reply_pdf = (my_has_final_ticket or my_has_final_reassign)

    application_docs_from_osdl = []
    supporting_docs_from_osdl = []

    for d in docs_from_osdl:
        file_name = str(d.get("name", "") or "")
        base_name = file_name.split("/")[-1] if "/" in file_name else file_name

        if base_name.startswith("Application_"):
            application_docs_from_osdl.append(d)
        else:
            supporting_docs_from_osdl.append(d)    

    uo_documents = []


    rcn_value = ""
    association_name_value = ""

    if osdl:
        rcn_value = (getattr(osdl, "rcn", "") or "").strip()
        association_name_value = (getattr(osdl, "association_name", "") or "").strip()

    uo_qs = NgoUoDocument.objects.using("secondary").filter(document_category="uo")

    # 1) First match by exact RCN, case-insensitive
    if rcn_value:
        uo_documents = list(
            uo_qs.filter(rcn__iexact=rcn_value).order_by("-id")
        )

    # 2) If no RCN match, match exact association name, case-insensitive
    # Matches: nitesh / NITESH / Nitesh
    # Does NOT match: Nitesh Kumar / Kumar Nitesh
    if not uo_documents and association_name_value:
        uo_documents = list(
            uo_qs.filter(association_name__iexact=association_name_value).order_by("-id")
        )   

    context = {
        "ticket": ticket,
        "reassignment": reassignment,
        "answers": answers,
        "logs": logs,
        "docs_from_osdl": docs_from_osdl,
        "application_docs_from_osdl": application_docs_from_osdl,
        "supporting_docs_from_osdl": supporting_docs_from_osdl,
        "folder_structure": folder_structure,
        "total_files": total_files,
        "folder_count": folder_count,
        "logs_with_files": logs_with_files,
        "status_history": status_history,
        "timeline_events": timeline_events,
        "latest_feedback": latest_feedback,
        "feedback_text": feedback_text,
        "qa_by_user": qa_by_user,
        "active_user_id": active_user_id,

        "action_taken": action_taken, 

        "is_creator": is_creator,
        "is_assignee": is_assignee,
        "is_reassign_owner": bool(reassignment and (user.id == reassignment.reassigned_by_id)),
        "is_reassign_assignee": bool(reassignment and reassignment.new_assigned_users.filter(id=user.id).exists()),
        "can_edit_reassign": bool(reassignment and can_edit_reassign),

        "can_edit_ticket": can_edit_ticket,
        "user_has_final": user_has_final,

        "userData": userData,
        "osdl_parent": osdl_parent,
        "osdl_docs": osdl_docs,
        "osdl_docs_count": osdl_docs_count,
        "osdl_logs": osdl_logs,
        "osdl_logs_count": osdl_logs_count,
        "ngoRedirectURL": ngoRedirectURL,

        "fb_by_user": fb_by_user,
        "fb_active_user_id": fb_active_user_id,
        "assignee_fb_rows": assignee_fb_rows,

        "report_summaries": report_summaries,
        "my_report_message": my_report_message,
        "my_report_file": my_report_file,
        "osdl": OtherServicesDocumentsLinkData,
        "committee_members": committee_members,
        "bank_name": bank_name,
        "bank_address": bank_address,
        "account_no": account_no,
        "can_view_all_tabs": can_view_all_tabs,
        "show_all_replies": show_all_replies,
        "adverse_reasons": adverse_reasons,

        "all_assignees": assignment_ctx["all_assignees"],
        "ngo_assignees": assignment_ctx["ngo_assignees"],
        "office_member_assignments": assignment_ctx["office_member_assignments"],
        "my_ngo_assigned": assignment_ctx["my_ngo_assigned"],
        "my_office_members": assignment_ctx["my_office_members"],
        "signaters": signaters,
        "my_signaters_to": my_signaters_to,
        "my_signaters_from": my_signaters_from,
        "reply_download_users": reply_download_users,
        "can_download_reply_pdf": can_download_reply_pdf,
        "uo_documents": uo_documents,
    }
  
    return render(request, "tickets/ticket_details_new.html", context)

# views.py


# If you have a helper for user display
try:
    from .utils import user_display
except ImportError:
    def user_display(u):
        if not u:
            return ""
        return (u.first_name or u.username or str(u))



@login_required(login_url='login') 
def reassign_ticket_detail(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    user = request.user
    all_reassignments = TicketReassignment.objects.filter(ticket=ticket).prefetch_related("new_assigned_users", "answers__question","reassigned_by")

    if user == ticket.created_by:
        reassignments = all_reassignments
    else:
        
        reassignments = all_reassignments.filter(
            Q(new_assigned_users=user) | Q(reassigned_by=user)
        ).distinct()
        # 
    ngo_data = None
    if ticket.models_id: 
        try:
            # url = f"http://124.123.18.181:8085/api/getngodetail?rcn={ticket.models_id}"
            NGO_BASE_URL=settings.NGO_BACKEND_URL
            # LOGIN_URL = NGO_BASE_URL+"/api/login/user"


            # url = NGO_BASE_URL+"/api/getngodetail?rcn={ticket.models_id}"
            # headers = {
            #     "Content-Type": "application/json",
            #     "Authorization": "Bearer 88|zG0Bk3XASNlWUsyV8lFEbVvxgRUmjNyCjE3bWYpw28531b6f",  
            # }
            # response = requests.get(url, headers=headers, timeout=10)
            # if response.status_code == 200:
            #     ngo_data = response.json()
            # else:
            #     ngo_data = {"error": f"API returned {response.status_code}"}
            ngo_obj = RegistrationsOther.objects.using('secondary').get(id=ticket.models_id)
    
            # Prepare the data in the same structure as your API response
            ngo_data = {
                "id": ngo_obj.id,
                "association_name": ngo_obj.association_name,
                "rcn": ngo_obj.rcn,
                # "mapStatus": ngo_obj.mapStatus,
                "created_at": ngo_obj.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "updated_at": ngo_obj.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
                # Add other fields if needed
            }
            # print('dddd------------------',ngo_data)
        except Exception as e:
            ngo_data = {"error": str(e)}

    return render(
        request,
        "tickets/reassign_ticket_detail.html",
        {
            "ticket": ticket,
            "reassignments": reassignments,
            "ngo_data": ngo_data,  # send NGO details to template
            # "userData":userData,
        },
    )





@login_required(login_url='login') 
def reassign_detail_view(request, ticket_id, reassign_id):
    """
    Show reassignment detail similar to ticket_detail, with scoped status history.
    """
    ticket = get_object_or_404(Ticket, id=ticket_id)
    reassignment = get_object_or_404(
        TicketReassignment.objects.prefetch_related("answers__question", "new_assigned_users"),
        id=reassign_id,
        ticket=ticket
    )

    # Answers for reassignment
    answers = reassignment.answers.all().select_related("question")

    # Feedback / Logs linked to this reassignment
    logs = TicketLog.objects.filter(ticketReassignment=reassignment).select_related('user').order_by('-created_at')

    # Get creator / assigned user data
    userData = ticket.created_by

    # --- Status history scoping logic ---

    # Creator sees everything
    if request.user == ticket.created_by:
        status_history = TicketStatusHistory.objects.filter(ticket=ticket).order_by('-created_at')
    else:
        # Reassignee window: from this reassignment's created_at up to (but not including) the next reassignment on this ticket
        next_ra = (
            TicketReassignment.objects
            .filter(ticket=ticket, created_at__gt=reassignment.created_at)
            .order_by('created_at')
            .first()
        )

        start_ts = reassignment.created_at
        end_filter = {}
        if next_ra:
            end_filter["created_at__lt"] = next_ra.created_at
        else:
            # no next reassignment → window is open-ended
            pass

        # Who is allowed within this window? Creator + the reassigned users for THIS reassignment
        reassignees = reassignment.new_assigned_users.all()
        allowed_users = [ticket.created_by_id] + list(reassignees.values_list('id', flat=True))

        # If viewer is a reassignee in THIS reassignment, show scoped/filtered history.
        if request.user.id in allowed_users and request.user != ticket.created_by:
            status_history = (
                TicketStatusHistory.objects
                .filter(ticket=ticket, created_at__gte=start_ts, **end_filter)
                .filter(changed_by_id__in=allowed_users)
                .order_by('-created_at')
            )
        else:
            # Not creator and not a reassignee on this reassignment: block or show none (choose your policy)
            raise PermissionDenied("You do not have permission to view this reassignment’s status history.")

    # --- Optional: NGO data as you had it ---
    ngo_data = None
    if ticket.models_id:
        try:
            NGO_BASE_URL = settings.NGO_BACKEND_URL  # kept for context; not used here
            ngo_obj = RegistrationsOther.objects.using('secondary').get(id=ticket.models_id)
            ngo_data = {
                "id": ngo_obj.id,
                "association_name": ngo_obj.association_name,
                "rcn": ngo_obj.rcn,
                "created_at": ngo_obj.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "updated_at": ngo_obj.updated_at.strftime("%Y-%m-%d %H:%M:%S"),
            }
        except Exception as e:
            ngo_data = {"error": str(e)}

    context = {
        "ticket": ticket,
        "reassignment": reassignment,
        "answers": answers,
        "logs": logs,
        "status_history": status_history,
        "userData": userData,
        "ngo_data": ngo_data,
    }
    return render(request, "tickets/reassign_single_detail.html", context)



@login_required(login_url='login')
def submit_reassign_form(request, reassign_id):
    reassignment = get_object_or_404(TicketReassignment, id=reassign_id)
    ticket = reassignment.ticket

    if request.method != "POST":
        return JsonResponse({
            "responseCode": 405,
            "responseMessage": "Invalid request method"
        })

    user = request.user
    action = request.POST.get("action")  # draft / final
    answers_data = {}

    signaters_to_id = request.POST.get("signaters_to")
    signaters_from_id = request.POST.get("signaters_from")

    # ------------------ SAVE REASSIGNMENT ANSWERS ------------------
    for key, value in request.POST.items():
        if not key.startswith("answer_"):
            continue

        question_id = key.replace("answer_", "")
        answer_text = (value or "").strip()

        try:
            answer_obj = reassignment.answers.get(question_id=question_id)
            answer_obj.answer = answer_text
            answer_obj.answered_by = user
            if action == "final":
                answer_obj.is_final = True
            else:
                answer_obj.is_final = False
            answer_obj.save()

            answers_data[question_id] = answer_text
        except ReassignmentAnswer.DoesNotExist:
            # optional: create if missing
            if answer_text:
                answer_obj = ReassignmentAnswer.objects.create(
                    reassignment=reassignment,
                    question_id=question_id,
                    answer=answer_text,
                    answered_by=user,
                    is_final=(action == "final"),
                )
                answers_data[question_id] = answer_obj.answer

    # ------------------ SAVE REPORT SUMMARY ------------------
    message = request.POST.get("message", "").strip()
    files = request.FILES.getlist("files")

    report_obj = (
        ReportSummary.objects
        .filter(ticket=ticket, reassignment=reassignment, user=user)
        .order_by("-id")
        .first()
    )

    if report_obj or message or files:
        if not report_obj:
            report_obj = ReportSummary(
                ticket=ticket,
                reassignment=reassignment,
                user=user,
            )

        if message != "":
            report_obj.summary = message

        if files:
            report_obj.file = files[0]

        report_obj.save()

    # ------------------ SAVE FEEDBACK IN TICKETLOG ------------------
    if message:
        TicketLog.objects.create(
            ticket=ticket,
            ticketReassignment=reassignment,
            user=user,
            message=message
        )

    # ------------------ SAVE SIGNATERS ------------------
    if signaters_to_id:
        try:
            signater_to = Signater.objects.get(id=signaters_to_id)
            TicketSignater.objects.update_or_create(
                ticket=ticket,
                user=user,
                action_type="to",
                defaults={
                    "signater": signater_to,
                }
            )
        except Signater.DoesNotExist:
            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "Selected Signaters To is invalid."
            })
    else:
        TicketSignater.objects.filter(
            ticket=ticket,
            user=user,
            action_type="to"
        ).delete()

    if signaters_from_id:
        try:
            signater_from = Signater.objects.get(id=signaters_from_id)
            TicketSignater.objects.update_or_create(
                ticket=ticket,
                user=user,
                action_type="from",
                defaults={
                    "signater": signater_from,
                }
            )
        except Signater.DoesNotExist:
            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "Selected Signaters From is invalid."
            })
    else:
        TicketSignater.objects.filter(
            ticket=ticket,
            user=user,
            action_type="from"
        ).delete()

    # ------------------ UPDATE REASSIGNMENT STATUS ------------------
    if action == "draft":
        reassignment.is_final = False
    elif action == "final":
        reassignment.is_final = True
        add_status_history(ticket, user, "submitted")

    reassignment.save()

    msg = (
        "Draft saved successfully."
        if action == "draft"
        else "Form submitted successfully. You cannot edit anymore."
    )

    return JsonResponse({
        "responseCode": 200,
        "responseMessage": msg,
        "answers": answers_data,
        "is_final": reassignment.is_final
    })


def submit_answer(request, ticket_id, question_id):
    if request.method != "POST":
        return HttpResponseBadRequest("Invalid request method")

    answer_text = request.POST.get("answer", "").strip()
    if not answer_text:
        return JsonResponse({"responseCode": 400, "responseMessage": "Answer cannot be empty."})

    try:
        ticket_answer = TicketAnswer.objects.get(ticket_id=ticket_id, question_id=question_id,answered_by=request.user)
        ticket_answer.answer = answer_text
        # ticket_answer.answered_by = request.user
        ticket_answer.save()
        return JsonResponse({
            "responseCode": 200,
            "responseMessage": "Answer submitted successfully",
            "answer": answer_text,
            # "answer": answer_text,
            "question_id": question_id
        })
    except TicketAnswer.DoesNotExist:
        return JsonResponse({"responseCode": 404, "responseMessage": "Ticket or Question not found."})





@login_required(login_url='login')
def submit_ticket_form(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)

    if request.method != "POST":
        return JsonResponse({
            "responseCode": 405,
            "responseMessage": "Invalid request method"
        })

    user = request.user
    action = (request.POST.get("action") or "draft").strip().lower()
    answers_data = {}

    signaters_to_id = (request.POST.get("signaters_to") or "").strip()
    signaters_from_id = (request.POST.get("signaters_from") or "").strip()

    # ------------------ QUESTIONS & ANSWERS ------------------
    for key, value in request.POST.items():
        if not key.startswith("answer_"):
            continue

        question_id = key.replace("answer_", "")
        answer_text = (value or "").strip()

        if answer_text == "":
            continue

        qs = TicketAnswer.objects.filter(
            ticket=ticket,
            question_id=question_id,
            answered_by=user,
        )

        editable_ta = qs.filter(is_final=False).order_by("-version").first()
        max_version = qs.aggregate(Max("version"))["version__max"] or 0
        has_final_any = qs.filter(is_final=True).exists()

        if editable_ta:
            ta = editable_ta
        else:
            if has_final_any and action == "final":
                latest_final = qs.order_by("-version").first()
                answers_data[question_id] = latest_final.answer if latest_final else ""
                continue

            ta = TicketAnswer(
                ticket=ticket,
                question_id=question_id,
                answered_by=user,
                version=max_version + 1,
            )

        ta.answer = answer_text

        if action == "final":
            ta.is_final = True
            ta.is_active = True
        else:
            ta.is_final = False
            ta.is_active = False

        ta.save()
        answers_data[question_id] = ta.answer or ""

    # ------------------ SAVE SIGNATERS ------------------
    if signaters_to_id:
        try:
            signater_to = Signater.objects.get(id=signaters_to_id)
            TicketSignater.objects.update_or_create(
                ticket=ticket,
                user=user,
                action_type="to",
                defaults={
                    "signater": signater_to,
                }
            )
        except Signater.DoesNotExist:
            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "Selected Signaters To is invalid."
            })
    else:
        TicketSignater.objects.filter(
            ticket=ticket,
            user=user,
            action_type="to"
        ).delete()

    if signaters_from_id:
        try:
            signater_from = Signater.objects.get(id=signaters_from_id)
            TicketSignater.objects.update_or_create(
                ticket=ticket,
                user=user,
                action_type="from",
                defaults={
                    "signater": signater_from,
                }
            )
        except Signater.DoesNotExist:
            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "Selected Signaters From is invalid."
            })
    else:
        TicketSignater.objects.filter(
            ticket=ticket,
            user=user,
            action_type="from"
        ).delete()

    # ------------------ REPORT SUMMARY (ONE ROW PER USER) ------------------
    message = request.POST.get("message", "").strip()
    files = request.FILES.getlist("files")

    report_obj = (
        ReportSummary.objects
        .filter(ticket=ticket, reassignment__isnull=True, user=user)
        .order_by("-id")
        .first()
    )

    if not report_obj and not message and not files:
        pass
    else:
        if not report_obj:
            report_obj = ReportSummary(
                ticket=ticket,
                reassignment=None,
                user=user,
            )

        if message != "":
            report_obj.summary = message

        if files:
            report_obj.file = files[0]

        report_obj.save()

    # ------------------ RESPONSE ------------------
    msg = "Draft saved successfully." if action == "draft" else "Form submitted successfully."
    is_final_for_user = (action == "final")

    if action == "final":
        redirect_url = reverse('tickets:ticket_success', args=[ticket.id])
        add_status_history(ticket, user, "submitted")

        return JsonResponse({
            "responseCode": 200,
            "responseMessage": msg,
            "answers": answers_data,
            "is_final": is_final_for_user,
            "redirect_url": redirect_url,
            "status": ticket.status,
        })

    return JsonResponse({
        "responseCode": 200,
        "responseMessage": msg,
        "answers": answers_data,
        "is_final": is_final_for_user,
        "status": ticket.status,
    })

@login_required(login_url='login')
def update_ticket_status(request, ticket_id):
    if request.method != "POST":
        return JsonResponse({"responseCode": 405, "responseMessage": "Invalid request method."})

    ticket = get_object_or_404(Ticket, id=ticket_id)
    new_status = (request.POST.get("status") or "").strip()
    user = request.user

    is_creator = ticket.created_by == user
    is_assignee = user in ticket.assigned_users.all()

    if is_creator:
        allowed_statuses = [status[0] for status in Ticket.STATUS_CHOICES]
    elif is_assignee:
        allowed_statuses = ['pending', 'resolved']  # keep your rule
    else:
        allowed_statuses = []

    if new_status not in allowed_statuses:
        return JsonResponse({
            "responseCode": 403,
            "responseMessage": "You are not allowed to set this status."
        })

    old_status = ticket.status

    # ✅ Read adverse fields (from modal)
    adverse_raw = (request.POST.get("adverse") or "no").lower()   # yes/no
    is_adverse = (adverse_raw == "yes")

    # this is reason_id (string)
    reason_id = (request.POST.get("adverse_option") or "").strip()

    reason_obj = None

    # ✅ Validate + fetch FK only when closing
    if new_status == "closed":
        if is_adverse:
            if not reason_id:
                return JsonResponse({
                    "responseCode": 400,
                    "responseMessage": "Please select Adverse Reason when Adverse is Yes."
                })

            reason_obj = AdverseReason.objects.filter(id=reason_id, is_active=True).first()
            if not reason_obj:
                return JsonResponse({
                    "responseCode": 400,
                    "responseMessage": "Invalid Adverse Reason selected."
                })
        else:
            reason_obj = None

    # ✅ Update ticket
    ticket.status = new_status
    ticket.closed_by = user if new_status == "closed" else None
    ticket.save(update_fields=["status", "closed_by", "updated_at"])

    # ✅ status history
    status_history_obj = TicketStatusHistory.objects.create(
        ticket=ticket,
        old_status=old_status,
        new_status=new_status,
        changed_by=user
    )

    # ✅ adverse history (only when closed)
    if new_status == "closed":
        TicketAdverseHistory.objects.create(
            ticket=ticket,
            status_history=status_history_obj,
            is_adverse=is_adverse,
            adverse_reason=reason_obj,   # ✅ FK saved here
            created_by=user
        )

    return JsonResponse({
        "responseCode": 200,
        "responseMessage": f"Ticket status updated to {new_status}.",
        "new_status": new_status
    })


@login_required(login_url='login') 
def submit_all_answers(request, ticket_id):
    if request.method == "POST":
        ticket = get_object_or_404(Ticket, id=ticket_id)
        answers_data = {}

        for key, value in request.POST.items():
            if key.startswith("answer_"):
                question_id = key.replace("answer_", "")
                answer_text = value.strip()
                if answer_text:
                    try:
                        ticket_answer = TicketAnswer.objects.get(ticket=ticket, question_id=question_id)
                        ticket_answer.answer = answer_text
                        ticket_answer.answered_by = request.user
                        ticket_answer.save()
                        answers_data[question_id] = answer_text
                    except TicketAnswer.DoesNotExist:
                        continue

        if answers_data:
            return JsonResponse({
                "responseCode": 200,
                "responseMessage": "Answers submitted successfully",
                "answers": answers_data
            })
        else:
            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "No answers submitted."
            })
    return HttpResponseBadRequest("Invalid request method")

#After Reopen ticket 
@login_required(login_url='login') 
def submit_reopened_answers(request, ticket_id):
    import json
    

    if request.method != "POST":
        return JsonResponse({"responseCode": 405, "responseMessage": "Invalid method"})

    try:
        data = json.loads(request.body)
        user = request.user
        ticket = Ticket.objects.get(id=ticket_id)

        with transaction.atomic():
            for qa in data.get("answers", []):
                qid = qa.get("question_id")
                ans_text = qa.get("answer")

                last_answer = (
                    TicketAnswer.objects.filter(ticket=ticket, question_id=qid)
                    .order_by("-version")
                    .first()
                )
                new_version = last_answer.version + 1 if last_answer else 1

                TicketAnswer.objects.create(
                    ticket=ticket,
                    question_id=qid,
                    answer=ans_text,
                    version=new_version,
                    answered_by=user,
                    is_active=True,
                )

            ticket.status = "closed"
            ticket.is_final = True
            ticket.save(update_fields=["status", "is_final", "updated_at"])

        return JsonResponse({
            "responseCode": 200,
            "responseMessage": "Answers submitted successfully with version tracking."
        })

    except Ticket.DoesNotExist:
        return JsonResponse({
            "responseCode": 404,
            "responseMessage": "Ticket not found."
        })
    except Exception as e:
        return JsonResponse({
            "responseCode": 500,
            "responseMessage": f"Error: {str(e)}"
        })

@login_required(login_url='login') 
def get_ticket_history(request, ticket_id):
    history = TicketQuestionHistory.objects.filter(ticket_answer__ticket_id=ticket_id).select_related("ticket_answer__question", "changed_by")

    data = [
        {
            "question": h.ticket_answer.question.text,
            "old_answer": h.old_answer,
            "changed_by": h.changed_by.username if h.changed_by else None,
            "reason": h.reason,
            "changed_at": h.changed_at.strftime("%Y-%m-%d %H:%M"),
        }
        for h in history
    ]

    return JsonResponse({"history": data})


@login_required(login_url='login') 
def create_ticket(request):
    if not can_create_ticket(request.user):
        return HttpResponseForbidden("You do not have permission to create tickets.")
    last_ticket = Ticket.objects.order_by('-id').first()
    next_ticket_id = (last_ticket.id + 1) if last_ticket else 1
    if request.method == 'POST':
        form = TicketForm(request.POST)
        if form.is_valid():
            ticket = form.save(commit=False)
            ticket.created_by = request.user
            ticket.save()
            form.save_m2m()  
            return redirect('ticket_detail', pk=ticket.pk)
    else:
        form = TicketForm()
        # Get query parameters from the URL
        ngo_name = request.GET.get('ngo', '')  # Get 'ngo' parameter, default to empty string
        ngo_id = request.GET.get('ngo_id', '')  # Get 'ngo_id' parameter, default to empty string

        state_ids = CustomUser.objects.values_list('state', flat=True).distinct()
        states = State.objects.filter().values('id', 'name')
        units = CustomUser.objects.exclude(unit__isnull=True).exclude(unit='').values_list('unit', flat=True).distinct()
        # questions = Question.objects.filter(status='active').order_by('id')
        questions = Question.objects.filter(status="active").order_by("id")
        # Categories = Category.objects.filter().order_by("id")
        # Categories = Category.objects.all().order_by("id")
        Categories = Category.objects.all().order_by("id").values("id", "name", "description")
        signaters = Signater.objects.filter(status="active").order_by("id")

        questions_data = list(questions.values("id", "text"))

      

    # return render(request, 'tickets/create_new_ticket.html', {'form': form,'questions':questions_data,"ngo_name":ngo_name,"ngo_id":ngo_id, "categories":list(Categories.values("id", "name")),'states': json.dumps(list(states))})
    return render(request, 'tickets/create_new_ticket.html', {'form': form,'signaters': list(signaters.values("id", "ranks")),'next_ticket_id': next_ticket_id,'questions':questions_data,"ngo_name":ngo_name,"ngo_id":ngo_id, "categories":list(Categories.values("id", "name")),'states': json.dumps(list(states))})


def get_questions_by_category(request):
    category_id = request.GET.get('category_id')
    questions = []
    if category_id:
        questions = Question.objects.filter(category_id=category_id, status='active').values('id', 'text')
    return JsonResponse({'questions': list(questions)})




def _get_payload(request):
    if request.content_type and request.content_type.startswith("multipart/form-data"):
        users_raw = request.POST.get("users") or "[]"
        try:
            users = json.loads(users_raw)
        except Exception:
            users = []
        return {
            "description": request.POST.get("description") or "",
            "module": request.POST.get("module") or "",
            "moduleData": request.POST.get("moduleData") or "",
            "ngoId": request.POST.get("ngoId") or "",
            "users": users,
            "ticket_type": request.POST.get("ticket_type") or "single",
        }
    else:
        try:
            return json.loads(request.body.decode("utf-8"))
        except Exception:
            return {}

def _norm(s):
    if s is None:
        return None
    s = str(s).strip()
    return s if s != "" else None

def _norm_pan(p):
    if p is None:
        return None
    p = str(p).strip().upper().replace(" ", "")
    return p or None


# -------------------------------------------------------------------
# Header normalize + alias mapping (THIS FIXES BLANK INSERTS)
# -------------------------------------------------------------------

def _clean_header(h):
    """
    "Bank Name" -> "bank_name"
    "Account No." -> "account_no"
    """
    if h is None:
        return ""
    h = str(h).strip().lower()
    h = re.sub(r"[^a-z0-9]+", "_", h)
    return h.strip("_")

def _normalize_row(row: dict) -> dict:
    return {_clean_header(k): v for k, v in (row or {}).items()}

# per-sheet alias maps (edit if your headers differ)
_ALIAS_COMMON = {
    "applicationid": "application_id",
    "app_id": "application_id",
    "applicaiton_id": "application_id",
}

_ALIAS_UTILIZATION = {
    "bank": "bank_name",
    "bankname": "bank_name",
    "bankaddress": "bank_address",
    "accountno": "account_no",
    "account_number": "account_no",
    "accountnumber": "account_no",
}

_ALIAS_COMMITTEE = {
    "member": "member_name",
    "membername": "member_name",
    "fatherhusbandname": "father_husband_name",
    "father_husband": "father_husband_name",
    "aadhar": "aadhaar_number",
    "aadhaar": "aadhaar_number",
    "pan": "pan_no",
    "fcra_reg_no": "fcra_registration_no",
    "fcra_registration": "fcra_registration_no",
    "fcra_registrationnumber": "fcra_registration_no",
}


def _apply_alias_map(row: dict, alias_map: dict) -> dict:
    """
    If row has key 'account_number' but DB expects 'account_no', copy it.
    """
    out = dict(row)
    # common aliases
    for src, dst in _ALIAS_COMMON.items():
        if src in out and dst not in out:
            out[dst] = out[src]
    # sheet specific aliases
    for src, dst in alias_map.items():
        if src in out and dst not in out:
            out[dst] = out[src]
    return out


# -------------------------------------------------------------------
# Timestamp fill (created_at NOT NULL FIX)
# -------------------------------------------------------------------

def _fill_created_updated(row: dict) -> dict:
    now = timezone.now()
    if row.get("created_at") in (None, "", " "):
        row["created_at"] = now
    if row.get("updated_at") in (None, "", " "):
        row["updated_at"] = now
    return row


# -------------------------------------------------------------------
# Excel read helpers
# -------------------------------------------------------------------

def _load_workbook_from_upload(uploaded_file):
    uploaded_file.seek(0)
    file_bytes = uploaded_file.read()
    # ✅ data_only=False so formulas return their text/value instead of None
    wb = load_workbook(filename=BytesIO(file_bytes), read_only=True, data_only=False)
    return wb


def _sheet_to_dict_rows(wb, sheet_name):
    # sheet name match ignoring case/spaces
    wanted = (sheet_name or "").strip().lower()
    actual = None
    for s in wb.sheetnames:
        if (s or "").strip().lower() == wanted:
            actual = s
            break
    if actual is None:
        return [], f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"

    ws = wb[actual]
    rows_iter = ws.iter_rows(values_only=True)

    headers = next(rows_iter, None)
    if not headers:
        return [], f"Sheet '{sheet_name}' is empty."

    # ✅ CLEAN headers HERE (this is the main fix)
    headers = [_clean_header(h) for h in headers]
    if not any(headers):
        return [], f"Sheet '{sheet_name}' has empty header row."

    out = []
    for r in rows_iter:
        if not r:
            continue
        if not any(cell not in (None, "", " ") for cell in r):
            continue

        d = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = r[i] if i < len(r) else None
            # normalize blanks
            if isinstance(v, str):
                v = v.strip()
            d[h] = v
        out.append(d)

    return out, None


def _first_data_row_from_sheet(wb, sheet_name):
    rows, err = _sheet_to_dict_rows(wb, sheet_name)
    if err:
        return {}, err
    if not rows:
        return {}, f"Sheet '{sheet_name}' has no data rows."
    return rows[0], None


# -------------------------------------------------------------------
# OSDL mapping + upsert (secondary DB)
# -------------------------------------------------------------------

def _row_to_osdl_kwargs(row):
    def g(k): return _norm(row.get(k))
    return {
        "rcn": g("rcn"),
        "registration_application_id": g("registration_application_id"),
        "section_file_number": g("section_file_number"),
        "form_submission_date": g("form_submission_date"),
        "darpan_id": g("darpan_id"),
        "association_name": g("association_name"),
        "registration_date": g("registration_date"),
        "association_address": g("association_address"),
        "association_state": g("association_state"),
        "association_district": g("association_district"),
        "association_official_telephone": g("association_official_telephone"),
        "email_id": g("email_id"),
        "association_official_website": g("association_official_website"),
        "association_chief_functionary_phone_number": g("association_chief_functionary_phone_number"),
        "association_chief_functionary_mobile_number": g("association_chief_functionary_mobile_number"),
        "act_registration_name": g("act_registration_name"),
        "act_registration_number": g("act_registration_number"),
        "date_of_act_registration": g("date_of_act_registration"),
        "place_of_act_registration": g("place_of_act_registration"),
        "pan_number": _norm_pan(row.get("pan_number")),
        "nature_of_association": g("nature_of_association"),
        "religion": g("religion"),
        "bank_name": g("bank_name"),
        "bank_address": g("bank_address"),
        "bank_email_id": g("bank_email_id"),
        "ifsc_code": g("ifsc_code"),
        "account_number": g("account_number"),
        "association_status": g("association_status"),
        "cancelled_suspended_date": g("cancelled_suspended_date"),
        "cancelled_suspended_remarks": g("cancelled_suspended_remarks"),
        "cancellation_reason": g("cancellation_reason"),
        "registration_certificate_location": g("registration_certificate_location"),
        "registration_applications_documents_location": g("registration_applications_documents_location"),
    }

_OSDL_COMMON_FIELD_NAMES = [
    "rcn","registration_application_id","section_file_number","form_submission_date",
    "darpan_id","association_name","registration_date","association_address",
    "association_state","association_district","association_official_telephone",
    "email_id","association_official_website",
    "association_chief_functionary_phone_number",
    "association_chief_functionary_mobile_number",
    "act_registration_name","act_registration_number",
    "date_of_act_registration","place_of_act_registration",
    "pan_number","nature_of_association","religion",
    "bank_name","bank_address","bank_email_id","ifsc_code",
    "account_number","association_status","cancelled_suspended_date",
    "cancelled_suspended_remarks","cancellation_reason",
    "registration_certificate_location",
    "registration_applications_documents_location",
]

def _create_log_snapshot(existing):
    using = "secondary"
    logs_field_names = {f.name for f in OtherServicesDocumentsLink_logs._meta.fields}

    payload = {"other_services_documents_link": existing}
    for fname in _OSDL_COMMON_FIELD_NAMES:
        if fname in logs_field_names:
            payload[fname] = getattr(existing, fname, None)

    if "action" in logs_field_names:
        payload["action"] = "updated"
    if "note" in logs_field_names:
        payload["note"] = "Pre-update snapshot"

    return OtherServicesDocumentsLink_logs.objects.using(using).create(**payload)

def _upsert_osdl_from_workbook(wb):
    sheet_name = "other_services_documents_link"
    row, err = _first_data_row_from_sheet(wb, sheet_name)
    if err:
        return None, err

    row = _normalize_row(row)  # ✅ normalize header names (PAN Number -> pan_number)

    osdl_kwargs = _row_to_osdl_kwargs(row)
    pan = osdl_kwargs.get("pan_number")
    if not pan:
        return None, "PAN number is required in other_services_documents_link sheet (column 'pan_number')."

    using = "secondary"
    with transaction.atomic(using=using):
        qs = OtherServicesDocumentsLink.objects.using(using).filter(pan_number=pan)
        match_count = qs.count()

        if match_count == 1:
            existing = qs.first()
            _create_log_snapshot(existing)

            for k, v in osdl_kwargs.items():
                if k == "pan_number":
                    continue
                setattr(existing, k, v)
            existing.save(using=using)
            return existing, None

        if match_count > 1:
            return None, f"Duplicate PAN found in DB for pan_number={pan}. Please fix duplicates first."

        new_obj = OtherServicesDocumentsLink(**osdl_kwargs)
        # if your table has created_at NOT NULL, ensure model auto_add OR set here:
        if hasattr(new_obj, "created_at") and getattr(new_obj, "created_at", None) is None:
            new_obj.created_at = timezone.now()
        if hasattr(new_obj, "updated_at") and getattr(new_obj, "updated_at", None) is None:
            new_obj.updated_at = timezone.now()

        new_obj.save(using=using)
        return new_obj, None


# -------------------------------------------------------------------
# Import committee_members_data + utilization_bnk_details (secondary DB)
# application_id = other_services_documents_link.id
# -------------------------------------------------------------------

def _import_committee_members_from_workbook(wb, osdl_obj, using="secondary"):
    rows, err = _sheet_to_dict_rows(wb, "committee_members_data")
    if err:
        return 0, err
    if not rows:
        return 0, None

    # normalize keys + aliases
    # rows = [_apply_alias_map(_normalize_row(r), _ALIAS_COMMITTEE) for r in rows]

    model_fields = {f.name for f in CommitteeMembersData._meta.fields}
    app_id = str(osdl_obj.id)

    for r in rows:
        if not r.get("application_id"):
            r["application_id"] = app_id
        _fill_created_updated(r)

    with transaction.atomic(using=using):
        # replace mode: delete old rows for this application_id
        CommitteeMembersData.objects.using(using).filter(application_id=app_id).delete()

        objs = []
        for r in rows:
            data = {k: r.get(k) for k in r.keys() if k in model_fields and k != "id"}
            objs.append(CommitteeMembersData(**data))

        CommitteeMembersData.objects.using(using).bulk_create(objs, batch_size=1000)

    return len(rows), None

def _import_utilization_bank_from_workbook(wb, osdl_obj, using="secondary"):
    rows, err = _sheet_to_dict_rows(wb, "utilization_bnk_details")
    if err:
        return 0, err
    if not rows:
        return 0, None

    # rows = [_apply_alias_map(_normalize_row(r), _ALIAS_UTILIZATION) for r in rows]

    model_fields = {f.name for f in UtilizationBnkDetails._meta.fields}
    app_id = str(osdl_obj.id)

    for r in rows:
        if not r.get("application_id"):
            r["application_id"] = app_id
        _fill_created_updated(r)

    with transaction.atomic(using=using):
        UtilizationBnkDetails.objects.using(using).filter(application_id=app_id).delete()

        objs = []
        for r in rows:
            data = {k: r.get(k) for k in r.keys() if k in model_fields and k != "id"}
            objs.append(UtilizationBnkDetails(**data))

        UtilizationBnkDetails.objects.using(using).bulk_create(objs, batch_size=1000)

    return len(rows), None


def _process_full_excel(uploaded_file):
    """
    1) Upsert OSDL from sheet other_services_documents_link
    2) Import committee_members_data
    3) Import utilization_bnk_details
    """
    wb = _load_workbook_from_upload(uploaded_file)

    try:
        osdl_obj, err = _upsert_osdl_from_workbook(wb)
        if err:
            return None, None, err

        c_count, err = _import_committee_members_from_workbook(wb, osdl_obj, using="secondary")
        if err:
            return None, None, err

        u_count, err = _import_utilization_bank_from_workbook(wb, osdl_obj, using="secondary")
        if err:
            return None, None, err

        return osdl_obj, {"committee_rows": c_count, "utilization_rows": u_count}, None

    except Exception as e:
        logger.exception("Excel processing failed")
        return None, None, str(e)





def _extract_pdf_text(uploaded_file) -> str:
    """
    Extract first 2-3 pages text and normalize spaces/newlines.
    """
    if not PdfReader:
        return ""

    uploaded_file.seek(0)
    reader = PdfReader(uploaded_file)

    text = ""
    for page in reader.pages[:3]:
        try:
            text += "\n" + (page.extract_text() or "")
        except Exception:
            pass

    # normalize
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)
    return text.strip()


def _extract_ngo_name_from_pdf(uploaded_file) -> Optional[str]:
    """
    Existing working flow (3A/3B/3C/7): extracts value after 'Name in full:'.
    """
    if not PdfReader:
        return None

    uploaded_file.seek(0)
    reader = PdfReader(uploaded_file)

    text = ""
    for page in reader.pages[:3]:
        try:
            text += "\n" + (page.extract_text() or "")
        except Exception:
            pass

    # normalize whitespace
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n+", "\n", text)

    m = re.search(r"Name in full\s*:\s*(.+)", text, flags=re.IGNORECASE)
    if not m:
        return None

    name = m.group(1).strip()
    name = re.split(r"\n|\(b\)\s*Address\s*:|Address\s*:", name, maxsplit=1)[0].strip()
    return name or None


def _detect_category_from_pdf_text(pdf_text: str) -> Optional[str]:
    """
    Detect FC category from PDF content.
    Returns: "3A" / "3B" / "3C" / "6E" / "7" / "renewal"
    """
    if not pdf_text:
        return None

    t = pdf_text.lower()

    if re.search(r"\bfc[\s\-]?3a\b", t) or "form fc-3a" in t or "form fc 3a" in t:
        return "3A"

    if re.search(r"\bfc[\s\-]?3b\b", t) or "form fc-3b" in t or "form fc 3b" in t:
        return "3B"

    if re.search(r"\bfc[\s\-]?3c\b", t) or "form fc-3c" in t or "form fc 3c" in t:
        return "3C"

    if re.search(r"\bfc[\s\-]?6e\b", t) or "form fc-6e" in t or "form fc 6e" in t:
        return "6E"

    if re.search(r"\bfc[\s\-]?7\b", t) or "form fc-7" in t or "form fc 7" in t:
        return "7"

    if "renewal" in t or "renew" in t:
        return "renewal"

    return None


def _extract_rcn_for_6e_from_pdf_text(pdf_text: str) -> Optional[str]:
    """
    ✅ For FC-6E:
    Find text like:
      'person/association: 075901010'
      'person / association : 075901010'
      'person/association 075901010'
    Then return 'R075901010'
    """
    if not pdf_text:
        return None

    t = pdf_text.strip()

    # Accept many spacing/punctuation variants
    m = re.search(
        r"\bperson\s*/\s*association\s*[:\-]?\s*([0-9]{6,20})\b",
        t,
        flags=re.IGNORECASE
    )
    if not m:
        return None

    num = (m.group(1) or "").strip()
    if not num:
        return None

    # Prefix with R
    return f"{num}R"


def _extract_file_no_from_pdf_text(pdf_text: str) -> Optional[str]:
    """
    Try to detect File No / MHA File No from PDF text.
    """
    if not pdf_text:
        return None

    patterns = [
        r"(?:MHA\s*)?File\s*No\.?\s*[:\-]?\s*([A-Za-z0-9\/\-\.\(\)]+)",
        r"F\.?\s*No\.?\s*[:\-]?\s*([A-Za-z0-9\/\-\.\(\)]+)",
        r"File\s*Number\s*[:\-]?\s*([A-Za-z0-9\/\-\.\(\)]+)",
    ]

    for p in patterns:
        m = re.search(p, pdf_text, flags=re.IGNORECASE)
        if m:
            val = (m.group(1) or "").strip()
            val = re.sub(r"[^\w\/\-\.\(\)]", "", val)
            return val or None

    return None


@require_POST
def parse_ngo_pdf(request):
    """
    ✅ For 6E: use RCN (person/association -> Rxxxx) to fetch NGO from DB.
    ✅ For other categories: use Name in full to fetch NGO from DB.
    """
    f = request.FILES.get("ngo_pdf")
    if not f:
        return JsonResponse({"ok": False, "message": "PDF not provided."}, status=400)

    pdf_text = _extract_pdf_text(f)
    if not pdf_text:
        return JsonResponse({"ok": False, "message": "Could not read text from PDF (might be scanned)."}, status=400)

    category_name = _detect_category_from_pdf_text(pdf_text) or ""
    file_no = _extract_file_no_from_pdf_text(pdf_text) or ""

    # ✅ category_id lookup from Category table (your existing approach)
    category_id = None
    if category_name:
        cat = Category.objects.filter(description__iexact=category_name).values("id", "name").first()
        if cat:
            category_id = cat["id"]

    qs = OtherServicesDocumentsLink.objects.using("secondary")

    # -----------------------------
    # ✅ FC-6E: lookup by RCN
    # -----------------------------
    if category_name == "6E":
        rcn = _extract_rcn_for_6e_from_pdf_text(pdf_text)
        if not rcn:
            return JsonResponse({"ok": False, "message": "Could not find person/association number in FC-6E PDF."}, status=400)

        exact = qs.filter(rcn__iexact=rcn).values(
            "id", "association_name", "rcn", "association_state", "section_file_number"
        ).first()

        if exact:
            if not file_no:
                file_no = exact.get("section_file_number") or ""

            return JsonResponse({
                "ok": True,
                "name_in_pdf": exact.get("association_name") or "",
                "file_no": file_no,
                "category_name": category_name,
                "category_id": category_id,
                "rcn_in_pdf": rcn,
                "ngo": exact
            })

        # optional closest fallback: contains match
        close = (
            qs.filter(rcn__icontains=rcn)
              .values("id", "association_name", "rcn", "association_state", "section_file_number")
              .order_by("-id")[:1]
              .first()
        )

        if close:
            if not file_no:
                file_no = close.get("section_file_number") or ""

            return JsonResponse({
                "ok": True,
                "name_in_pdf": close.get("association_name") or "",
                "file_no": file_no,
                "category_name": category_name,
                "category_id": category_id,
                "rcn_in_pdf": rcn,
                "ngo": close,
                "note": "Closest match by RCN selected."
            })

        return JsonResponse({
            "ok": True,
            "name_in_pdf": "",
            "file_no": file_no,
            "category_name": category_name,
            "category_id": category_id,
            "rcn_in_pdf": rcn,
            "ngo": None
        })

    # -----------------------------
    # ✅ Other categories: lookup by name
    # -----------------------------
    name = _extract_ngo_name_from_pdf(f)
    if not name:
        return JsonResponse({"ok": False, "message": "Could not find 'Name in full' in PDF."}, status=400)

    exact = qs.filter(association_name__iexact=name).values(
        "id", "association_name", "rcn", "association_state", "section_file_number"
    ).first()

    if exact:
        if not file_no:
            file_no = exact.get("section_file_number") or ""

        return JsonResponse({
            "ok": True,
            "name_in_pdf": name,
            "file_no": file_no,
            "category_name": category_name,
            "category_id": category_id,
            "ngo": exact
        })

    close = (
        qs.filter(association_name__icontains=name)
          .values("id", "association_name", "rcn", "association_state", "section_file_number")
          .order_by("-id")[:1]
          .first()
    )

    if close:
        if not file_no:
            file_no = close.get("section_file_number") or ""

        return JsonResponse({
            "ok": True,
            "name_in_pdf": name,
            "file_no": file_no,
            "category_name": category_name,
            "category_id": category_id,
            "ngo": close,
            "note": "Closest match selected."
        })

    return JsonResponse({
        "ok": True,
        "name_in_pdf": name,
        "file_no": file_no,
        "category_name": category_name,
        "category_id": category_id,
        "ngo": None
    })




DOWNLOADS_BASE_DIR = settings.DOCURL
DOCROOT = settings.DOCROOT
USING = "secondary"
ALLOWED_EXTS = {
    ".pdf", ".jpg", ".jpeg", ".png",
    ".doc", ".docx", ".xls", ".xlsx", ".txt"
}
# NOTE: keep your existing ALLOWED_EXTS wherever it is defined
# ALLOWED_EXTS = {".pdf", ".jpg", ".jpeg", ".png", ...}


# =========================================================
# Helpers
# =========================================================
def _safe_basename(name: str) -> str:
    name = os.path.basename(name or "upload")
    name = name.replace("\x00", "")
    return name


def _normalize_folder(folder: str) -> str:
    """
    "/2025/07/05/01/0013882025/" -> "2025/07/05/01/0013882025/"
    ensures trailing slash
    """
    folder = (folder or "").strip()
    folder = folder.lstrip("/")
    if folder and not folder.endswith("/"):
        folder += "/"
    return folder


def _unique_path(dest_dir: Path, filename: str) -> Path:
    base = Path(filename).stem
    ext = Path(filename).suffix
    p = dest_dir / filename
    i = 1
    while p.exists():
        p = dest_dir / f"{base}({i}){ext}"
        i += 1
    return p


def _write_uploaded_file_to_disk(uploaded_file, dest_path: Path) -> None:
    dest_path.parent.mkdir(parents=True, exist_ok=True)
    with open(dest_path, "wb") as out:
        for chunk in uploaded_file.chunks():
            out.write(chunk)


def _build_new_folder(file_no: str, dt=None) -> str:
    """
    Folder format: YYYY/MM/DD/HH/<file_no>/
    """
    dt = dt or timezone.now()
    file_no = (file_no or "").strip() or dt.strftime("%H%M%S")
    return f"{dt:%Y/%m/%d/%H}/{file_no}/"


def _abs_dir_from_rel_folder(rel_folder: str) -> Path:
    """
    Creates: <settings.DOCURL>/<rel_folder>/
    """
    rel_folder = _normalize_folder(rel_folder)
    abs_dir = DOWNLOADS_BASE_DIR / rel_folder
    abs_dir.mkdir(parents=True, exist_ok=True)
    return abs_dir


def _ensure_secondary_connection():
    connections[USING].cursor()


def _ensure_osdl_folder(osdl_id: int, file_no: str, OtherServicesDocumentsLink):
    """
    1) Read registration_applications_documents_location from DB
    2) If exists -> use it
    3) If not -> create folder and SAVE BACK to DB
    Returns: (osdl_obj, rel_folder)
    """
    _ensure_secondary_connection()

    osdl = OtherServicesDocumentsLink.objects.using(USING).filter(id=osdl_id).first()
    if not osdl:
        return None, None

    db_folder = getattr(osdl, "registration_applications_documents_location", "") or ""
    db_folder = _normalize_folder(db_folder)

    if db_folder:
        _abs_dir_from_rel_folder(db_folder)
        return osdl, db_folder

    new_folder = _build_new_folder(file_no=file_no)
    _abs_dir_from_rel_folder(new_folder)

    # keep same format: leading "/" in DB column
    osdl.registration_applications_documents_location = f"/{new_folder}"
    osdl.save(using=USING)

    return osdl, new_folder


# =========================================================
# Ensure folder exists in OSDL
# =========================================================
def _is_safe_zip_path(rel_path: str) -> bool:
    if not rel_path:
        return False
    rel_path = rel_path.replace("\\", "/").strip()

    if rel_path.startswith("/"):
        return False

    head = rel_path.split("/")[0]
    if ":" in head:  # blocks C:\ etc
        return False

    parts = [p for p in rel_path.split("/") if p]
    if any(p == ".." for p in parts):
        return False

    return True


def _allowed_extracted_file(filename: str) -> bool:
    return Path(filename).suffix.lower() in ALLOWED_EXTS

def _save_files_into_folder(
    *,
    ticket,
    request,
    rel_folder: str,
    TicketApplicationDoc,
    file_field_name: str,
    downloads_base_dir,
    osdl_obj=None,
    filename_prefix: str = "",
):
    """
    - Saves normal uploads
    - If .zip uploaded -> extracts and saves extracted files
    - Creates TicketApplicationDoc rows for each saved file
    - Returns:
        {
            "ids": [...],
            "files": [
                {"id": 1, "name": "abc.pdf", "path": "folder/abc.pdf"},
                ...
            ]
        }
    """
    result = {
        "ids": [],
        "files": [],
    }

    _ensure_secondary_connection()

    files = request.FILES.getlist(file_field_name)
    if not files:
        return result

    clean_files = [f for f in files if f and getattr(f, "size", 0) > 0]
    if not clean_files:
        return result

    rel_folder_norm = _normalize_folder(rel_folder)
    user_id = request.user.id if request.user.is_authenticated else None

    downloads_base_dir = Path(downloads_base_dir)
    downloads_base_dir.mkdir(parents=True, exist_ok=True)

    downloads_storage = FileSystemStorage(location=str(downloads_base_dir))

    def _create_doc(relative_name_for_db: str):
        doc_kwargs = dict(
            ticket_id=ticket.id,
            status="uploaded",
            uploaded_by_id=user_id,
        )

        if osdl_obj is not None:
            doc_kwargs["other_services_documents_link_id"] = osdl_obj.id

        doc = TicketApplicationDoc(**doc_kwargs)
        doc.file.name = relative_name_for_db
        doc.save(using=USING)

        result["ids"].append(doc.id)
        result["files"].append({
            "id": doc.id,
            "name": Path(relative_name_for_db).name,
            "path": relative_name_for_db,
        })

    def _save_one(storage_rel_path: str, file_obj):
        storage_rel_path = downloads_storage.get_available_name(storage_rel_path)
        saved_name = downloads_storage.save(storage_rel_path, file_obj)
        _create_doc(saved_name)

    try:
        with transaction.atomic(using=USING):
            for f in clean_files:
                original_name = _safe_basename(getattr(f, "name", "upload"))

                if filename_prefix and not original_name.lower().endswith(".zip"):
                    original_name = f"{filename_prefix}{original_name}"

                lower_name = original_name.lower()

                if lower_name.endswith(".zip"):
                    try:
                        f.seek(0)
                        zdata = io.BytesIO(f.read())

                        with zipfile.ZipFile(zdata) as zf:
                            extracted_any = False

                            for info in zf.infolist():
                                if info.is_dir():
                                    continue

                                inner_name = (info.filename or "").replace("\\", "/").strip()
                                if not _is_safe_zip_path(inner_name):
                                    logger.warning("Skipped unsafe zip path: %s", inner_name)
                                    continue

                                if not _allowed_extracted_file(inner_name):
                                    logger.warning("Skipped disallowed extracted file: %s", inner_name)
                                    continue

                                inner_path = Path(inner_name)
                                inner_dir = inner_path.parent
                                inner_file = inner_path.name

                                if filename_prefix:
                                    inner_file = f"{filename_prefix}{inner_file}"

                                new_inner_path = str(inner_dir / inner_file) if str(inner_dir) != "." else inner_file
                                extracted_rel = f"{rel_folder_norm}{new_inner_path}"

                                with zf.open(info) as src:
                                    file_bytes = src.read()

                                if not file_bytes:
                                    logger.warning("Skipped empty extracted file: %s", inner_name)
                                    continue

                                django_file = ContentFile(file_bytes, name=inner_file)
                                _save_one(extracted_rel, django_file)
                                extracted_any = True

                            if not extracted_any:
                                logger.warning("Zip uploaded but no valid file extracted: %s", original_name)

                    except zipfile.BadZipFile:
                        logger.exception("Bad zip uploaded: %s", original_name)
                        continue
                    except Exception as e:
                        logger.exception("ZIP extraction failed for %s: %s", original_name, e)
                        continue

                    continue

                relative_name_for_db = f"{rel_folder_norm}{original_name}"
                try:
                    f.seek(0)
                except Exception:
                    pass

                _save_one(relative_name_for_db, f)

        return result

    except Exception as e:
        logger.exception("atomic(using=%s) failed: %s", USING, e)

        for f in clean_files:
            try:
                original_name = _safe_basename(getattr(f, "name", "upload"))

                if filename_prefix and not original_name.lower().endswith(".zip"):
                    original_name = f"{filename_prefix}{original_name}"

                lower_name = original_name.lower()

                if lower_name.endswith(".zip"):
                    try:
                        f.seek(0)
                        zdata = io.BytesIO(f.read())

                        with zipfile.ZipFile(zdata) as zf:
                            extracted_any = False

                            for info in zf.infolist():
                                if info.is_dir():
                                    continue

                                inner_name = (info.filename or "").replace("\\", "/").strip()
                                if not _is_safe_zip_path(inner_name):
                                    logger.warning("Skipped unsafe zip path in fallback: %s", inner_name)
                                    continue

                                if not _allowed_extracted_file(inner_name):
                                    logger.warning("Skipped disallowed extracted file in fallback: %s", inner_name)
                                    continue

                                inner_path = Path(inner_name)
                                inner_dir = inner_path.parent
                                inner_file = inner_path.name

                                if filename_prefix:
                                    inner_file = f"{filename_prefix}{inner_file}"

                                new_inner_path = str(inner_dir / inner_file) if str(inner_dir) != "." else inner_file
                                extracted_rel = f"{rel_folder_norm}{new_inner_path}"

                                with zipfile.ZipFile(io.BytesIO(zdata.getvalue())) as zf2:
                                    with zf2.open(info.filename) as src:
                                        file_bytes = src.read()

                                if not file_bytes:
                                    logger.warning("Skipped empty extracted file in fallback: %s", inner_name)
                                    continue

                                django_file = ContentFile(file_bytes, name=inner_file)
                                _save_one(extracted_rel, django_file)
                                extracted_any = True

                            if not extracted_any:
                                logger.warning("Zip uploaded but no valid file extracted in fallback: %s", original_name)

                    except zipfile.BadZipFile:
                        logger.exception("Bad zip uploaded in fallback: %s", original_name)
                        continue
                    except Exception as e:
                        logger.exception("ZIP extraction failed in fallback for %s: %s", original_name, e)
                        continue

                    continue

                relative_name_for_db = f"{rel_folder_norm}{original_name}"
                try:
                    f.seek(0)
                except Exception:
                    pass

                _save_one(relative_name_for_db, f)

            except Exception as e:
                logger.exception("Non-atomic save failed: %s", e)

        return result


@login_required(login_url="login")
def store_ticket(request):
    if request.method != "POST":
        return render(request, "tickets/create_new_ticket.html")

    data = _get_payload(request)  # your existing function
    ticket_type = data.get("ticket_type") or "single"

    description = data.get("description")
    file_no = (request.POST.get("file_no") or "").strip()
    mhafile_no = (request.POST.get("mhafile_no") or "").strip()
    signater_id = (request.POST.get("signater_id") or "").strip()
    module_name = data.get("module") or "NGO"
    ngoId = data.get("ngoId")
    users = data.get("users", [])  # list of dict from JSON
    excel_import = None
    final_office_file_no = None
    office_file_no = file_no
    next_ticket_id = data.get("next_ticket_id")

    if ticket_type != "single":
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "Invalid ticket type."
        })

    if not ngoId:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "Please select a valid NGO from search."
        })

    # ✅ NEW models (do not disturb existing ones)

    def _safe_int(val):
        try:
            if val is None:
                return None
            s = str(val).strip()
            if s == "":
                return None
            return int(s)
        except Exception:
            return None

    def _normalize_upload_result(result):
        """
        Supports both old and new _save_files_into_folder return formats:
        Old: [1,2,3]
        New: {"ids":[1,2], "files":[{"id":1,"name":"a.pdf","path":"..."}]}
        """
        if isinstance(result, dict):
            return {
                "ids": result.get("ids", []) or [],
                "files": result.get("files", []) or [],
            }
        elif isinstance(result, list):
            return {
                "ids": result,
                "files": [],
            }
        return {
            "ids": [],
            "files": [],
        }

    # ensure DB folder exists and use it
    osdl, rel_folder = _ensure_osdl_folder(int(ngoId), file_no, OtherServicesDocumentsLink)
    if not osdl or not rel_folder:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "Invalid NGO selected."
        })

    models_name = "NGO" if (module_name == "NGO" or not module_name) else (module_name or "").lower()
    models_object = getattr(osdl, "association_name", "") or ""
    rcn = getattr(osdl, "rcn", None)

    if not (models_name and models_object and users):
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "Missing required fields."
        })

    ticket = None

    try:
        ticket = Ticket.objects.create(
            title=f"Ticket for {models_name}",
            description=description,
            created_by=request.user,
            models_name=models_name,
            models_id=osdl.id,
            rcn=rcn,
            file_no=file_no,
            mha_file_no=mhafile_no,
            models_object=models_object,
            request_type=ticket_type,
            status="open",
            signater_id=signater_id if signater_id else None,
        )

        final_office_file_no = f"{office_file_no}({ticket.id})" if office_file_no else ""

        TicketActionTrace.objects.create(
            ticket=ticket,
            action_key="TICKET_CREATED",
            is_success=True,
            status_text="Ticket created successfully",
            performed_by=request.user,
            office_file_no=final_office_file_no,
            meta={
                "ticket_id": ticket.id,
                "next_ticket_id": next_ticket_id,
                "office_file_input": office_file_no,
            }
        )

        # =====================================================
        # ✅ Decide sections from users JSON (source of truth)
        # =====================================================
        def _norm_section(val: str) -> str:
            return (val or "").strip().lower()

        has_user_section = any(_norm_section(u.get("section")) == "user_section" for u in users)
        has_bearer_section = any(_norm_section(u.get("section")) == "office_bearer" for u in users)

        meta_payload = {
            "ngo_id": str(ngoId),
            "ngo_name": models_object,
            "rcn": rcn,
            "application_id": (request.POST.get("application_id") or ""),
            "module": module_name,
        }

        if has_user_section:
            TicketSectionMeta.objects.get_or_create(
                ticket=ticket,
                section_type="USER_SECTION",
                defaults={"created_by": request.user, "meta": meta_payload},
            )

        if has_bearer_section:
            TicketSectionMeta.objects.get_or_create(
                ticket=ticket,
                section_type="OFFICE_BEARER",
                defaults={"created_by": request.user, "meta": meta_payload},
            )

        # =====================================================
        # ✅ Upload flow with backend validation
        # =====================================================
        doc_key = "documents" if request.FILES.getlist("documents") else "upload_documents"

        uploaded_docs_result = _save_files_into_folder(
            ticket=ticket,
            request=request,
            rel_folder=rel_folder,
            TicketApplicationDoc=TicketApplicationDoc,
            file_field_name=doc_key,
            downloads_base_dir=DOWNLOADS_BASE_DIR,
            osdl_obj=osdl,
        )
        uploaded_docs_result = _normalize_upload_result(uploaded_docs_result)
        uploaded_docs_ids = uploaded_docs_result["ids"]
        uploaded_docs_files = uploaded_docs_result["files"]

        ngo_app_result = _save_files_into_folder(
            ticket=ticket,
            request=request,
            rel_folder=rel_folder,
            TicketApplicationDoc=TicketApplicationDoc,
            file_field_name="ngo_pdf",
            downloads_base_dir=DOWNLOADS_BASE_DIR,
            osdl_obj=osdl,
            filename_prefix="Application_",
        )
        ngo_app_result = _normalize_upload_result(ngo_app_result)
        ngo_app_ids = ngo_app_result["ids"]
        ngo_app_files = ngo_app_result["files"]

        # ✅ Mandatory backend validation
        if not ngo_app_ids:
            try:
                ticket.delete()
            except Exception:
                logger.exception("Failed to rollback ticket after NGO application upload failure")

            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "Application file upload failed. Please re-upload NGO application PDF."
            })

        if not uploaded_docs_ids:
            try:
                ticket.delete()
            except Exception:
                logger.exception("Failed to rollback ticket after document upload failure")

            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "Document upload is mandatory. Please re-upload document(s)."
            })

        # =====================================================
        # ✅ Cache bearer member rows by member_key
        # =====================================================
        bearer_member_cache = {}  # key -> TicketMemberRow

        # =====================================================
        # assign users + questions
        # =====================================================
        for u in users:
            user_id = u.get("user_id")
            questions = u.get("questions", [])

            section = _norm_section(u.get("section"))  # user_section / office_bearer
            db_section = "NGO" if section == "user_section" else "OFFICE_BEARER"

            category_obj = None
            cat_id = u.get("category_id")
            if cat_id:
                try:
                    category_obj = Category.objects.filter(id=int(cat_id)).first()
                except Exception:
                    category_obj = None

            member_row_obj = None
            if db_section == "OFFICE_BEARER":
                member_id_raw = u.get("member_id")
                member_data_id = _safe_int(member_id_raw)

                member_name = (u.get("member_name") or "").strip() or None
                member_key = (u.get("member_key") or "").strip() or member_name

                if member_key:
                    if member_key in bearer_member_cache:
                        member_row_obj = bearer_member_cache[member_key]
                    else:
                        member_row_obj = TicketMemberRow.objects.create(
                            ticket=ticket,
                            section_type="OFFICE_BEARER",
                            member_data_id=member_data_id,
                            member_name=member_name,
                            created_by=request.user,
                        )
                        bearer_member_cache[member_key] = member_row_obj

            try:
                assigned_user = CustomUser.objects.get(id=user_id)
            except CustomUser.DoesNotExist:
                continue

            ticket.assigned_users.add(assigned_user)
            TicketLog.objects.create(ticket=ticket, user=assigned_user)

            assign_obj, _ = AssignUsersCategory.objects.get_or_create(
                ticket=ticket,
                section_type=db_section,
                member_row=member_row_obj if db_section == "OFFICE_BEARER" else None,
                user=assigned_user,
                category=category_obj,
            )

            for q_item in questions:
                q_obj = None
                try:
                    q_id = int(q_item)
                    q_obj = Question.objects.filter(pk=q_id).first()
                except (ValueError, TypeError):
                    q_text = str(q_item or "").strip()
                    if q_text:
                        q_obj = Question.objects.filter(text=q_text).first()

                if q_obj:
                    TicketAnswer.objects.get_or_create(
                        ticket=ticket,
                        question=q_obj,
                        answered_by=assigned_user
                    )
                    assign_obj.questions.add(q_obj)

        TicketStatusHistory.objects.create(
            ticket=ticket,
            old_status=None,
            new_status="open",
            changed_by=request.user
        )

        abs_folder = str(_abs_dir_from_rel_folder(rel_folder))

        return JsonResponse({
            "responseCode": 200,
            "responseMessage": "Ticket created successfully!",
            "ticket_id": ticket.id,
            "db_folder_value": getattr(osdl, "registration_applications_documents_location", ""),
            "folder_rel_used": f"/{_normalize_folder(rel_folder)}",
            "folder_abs_used": abs_folder,
            "uploaded_docs_ids": uploaded_docs_ids,
            "uploaded_docs_files": uploaded_docs_files,
            "ngo_application_uploaded_ids": ngo_app_ids,
            "ngo_application_files": ngo_app_files,
            "excel_import": excel_import,
            "next_ticket_id": next_ticket_id,
            "office_file_no": final_office_file_no,
            "office_file_input": office_file_no,
        })

    except Exception as e:
        logger.exception("Ticket creation failed")

        if ticket is not None:
            try:
                ticket.delete()
            except Exception:
                logger.exception("Failed to rollback ticket after exception")

        return JsonResponse({
            "responseCode": 400,
            "responseMessage": f"Error creating ticket: {str(e)}"
        })

        
@login_required(login_url='login') 
def reassign_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)

    if request.method != "POST":
        return JsonResponse({
            "responseCode": 405,
            "responseMessage": "Invalid request method"
        })

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception as e:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": f"Invalid JSON payload: {str(e)}"
        })

    user_id = data.get("user")
    reason = data.get("reason", "")

    if not user_id:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "No user selected for reassignment"
        })

    try:
        assigned_user = CustomUser.objects.get(id=user_id)

        # Create reassignment record
        reassignment = TicketReassignment.objects.create(
            ticket=ticket,
            reassigned_by=request.user,
            reason=reason
        )

        # Assign the new user
        reassignment.new_assigned_users.add(assigned_user)

        # 🔹 Fetch all questions already assigned to this ticket
        ticket_answers = TicketAnswer.objects.filter(ticket=ticket)

        # 🔹 Backup each answer into ReassignmentAnswer
        with transaction.atomic():
            for t_ans in ticket_answers:
                ReassignmentAnswer.objects.update_or_create(
                    reassignment=reassignment,
                    question=t_ans.question,
                    answered_by=t_ans.answered_by,
                    defaults={
                        # 'answer': t_ans.answer,  # uncomment if you want the text
                        'is_active': True,
                        'is_final': False,  # or whatever logic you want
                    },
                )

        # 🔹 Optional: mark old answers inactive or increment version
        # ticket_answers.update(is_active=False, version=F('version') + 1)

        # 🔹 Log ticket status change
        last_history = TicketStatusHistory.objects.filter(ticket=ticket).order_by('-id').first()
        old_status = last_history.new_status if last_history else ticket.status

        add_status_history(ticket, request.user, "reassigned")

        # TicketStatusHistory.objects.create(
        #     ticket=ticket,
        #     old_status=old_status,
        #     new_status="reassigned",
        #     changed_by=request.user
        # )

        return JsonResponse({
            "responseCode": 200,
            "responseMessage": f"Ticket reassigned to {assigned_user.first_name} successfully"
        })

    except CustomUser.DoesNotExist:
        return JsonResponse({
            "responseCode": 404,
            "responseMessage": "Selected user does not exist"
        })
    except Exception as e:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": f"Error during reassignment: {str(e)}"
        })





@login_required(login_url='login') 
def ticket_success(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)

    # Optional: you can pass ticket info to show on success page
    return render(request, 'tickets/ticket_success.html', {'ticket': ticket})

@csrf_exempt
def get_units_by_state(request):
    # print('state_id11')
    state_id = request.GET.get('state_id')
    
    if not state_id:
        return JsonResponse({'units': []})

    units = CustomUser.objects.filter(state_id=state_id).exclude(unit__isnull=True).exclude(unit='').values_list('unit', flat=True).distinct()
    return JsonResponse({'units': list(units)})

@csrf_exempt
def get_users_by_state_unit(request):
    state_id = request.GET.get("state_id")
    unit = request.GET.get("unit")
    app_label = settings.APPNGO

    # Need both filters for location
    if not state_id or not unit:
        return JsonResponse({"users": []})

    # Resolve the AppModule first (case-insensitive, must be active)
    app = AppModule.objects.filter(label__iexact=app_label, is_active=True).first()
    if not app:
        return JsonResponse({"users": []})

    # Query FROM UserAccessControl, then hop to user fields
    uac_qs = (
        UserAccessControl.objects
        .select_related("user")
        .filter(
            user__is_active=True,
            user__state_id=state_id,
            user__unit=unit,
        )
        .filter(
            Q(permissions__app=app, permissions__is_active=True)
            |
            Q(
                roles__is_active=True,
                roles__permissions__app=app,
                roles__permissions__is_active=True
            )
        )
        .distinct()
    )

    # Optional: exclude the requester, if authenticated
    requester_id = getattr(request.user, "id", None)
    if requester_id:
        uac_qs = uac_qs.exclude(user__id=requester_id)

    users_list = []

    for uac in uac_qs:
        u = uac.user

        # ✅ IMPORTANT: if user can create ticket, DO NOT show in list
        try:
            if can_create_ticket(u):
                continue
        except TypeError:
            # fallback if your can_create_ticket signature requires request
            # (rare, but safe)
            pass

        name = (f"{(u.first_name or '').strip()} {(u.last_name or '').strip()}".strip())
        if not name:
            name = (u.email or u.phone_no or "").strip()

        users_list.append({
            "id": u.id,
            "ucs_no": u.ucs_no,
            "rank": u.rank,
            "name": name,
        })

    return JsonResponse({"users": users_list})



@login_required(login_url='login')
def reopen_ticket(request, ticket_id):
    if request.method != "POST":
        return JsonResponse({
            "responseCode": 405,
            "responseMessage": "Method not allowed."
        })
    try:
        # MULTI-SELECT: all selected users from <select multiple name="reopen_for_user">
        reopen_for_user_ids = request.POST.getlist("reopen_for_user")
        request.GET.get("reassign")
        if not reopen_for_user_ids:
            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "Please select at least one user to reopen the ticket for."
            })

        with transaction.atomic():
            # Lock ticket row
            ticket = (
                Ticket.objects
                .select_for_update()
                .select_related("created_by")
                .get(id=ticket_id)
            )
            current_user = request.user

            # Only creator can reopen
            if ticket.created_by_id != current_user.id:
                return JsonResponse({
                    "responseCode": 403,
                    "responseMessage": "You are not allowed to reopen this ticket."
                })

            reason = request.POST.get("reason", "Ticket reopened by creator")
            new_status = "reopened"

            # ---------------------------------------------------------
            # FOR EACH SELECTED USER (e.g. Ravinder, Ankur, or both)
            # ---------------------------------------------------------
            for user_id in reopen_for_user_ids:
                reopen_for_user = get_object_or_404(User, id=user_id)

                # All active answers for this user on this ticket
                old_answers = TicketAnswer.objects.filter(
                    ticket=ticket,
                    answered_by=reopen_for_user,
                    is_active=True,
                )

                for old_ans in old_answers:
                    # 1) Save old value into history
                    TicketQuestionHistory.objects.create(
                        ticket_answer=old_ans,
                        old_answer=old_ans.answer,
                        changed_by=current_user,
                        reason=reason
                    )

                    # 2) Mark old answer inactive
                    old_ans.is_active = False
                    old_ans.save(update_fields=["is_active"])

                    # 3) Compute next version for this (ticket, question, user)
                    last = (
                        TicketAnswer.objects
                        .filter(
                            ticket=ticket,
                            question=old_ans.question,
                            answered_by=reopen_for_user
                        )
                        .order_by("-version")
                        .first()
                    )
                    next_version = (last.version if last else 0) + 1

                    # 4) CREATE NEW FRESH ENTRY WITH BLANK ANSWER
                    TicketAnswer.objects.create(
                        ticket=ticket,
                        question=old_ans.question,
                        answer="",          # ✅ BLANK, no previous answer
                        version=next_version,
                        is_active=True,
                        is_final=False,      # 0 → "fresh" for re-answering
                        answered_by=reopen_for_user,
                    )

            # ---------------------------------------------------------
            # STATUS HISTORY (same as your old code)
            # ---------------------------------------------------------
            old_status = ticket.status

            last_history = (
                TicketStatusHistory.objects
                .filter(ticket=ticket)
                .order_by('-id')
                .first()
            )
            if last_history:
                old_status = last_history.new_status

            if new_status != old_status:
                # TicketStatusHistory.objects.create(
                #     ticket=ticket,
                #     old_status=old_status,
                #     new_status=new_status,
                #     changed_by=current_user
                # )
                add_status_history(ticket, request.user, "reopened")


            ticket.status = new_status
            ticket.is_final = False
            ticket.save(update_fields=["status", "is_final"])

        return JsonResponse({
            "responseCode": 200,
            "responseMessage": "Ticket successfully reopened and fresh blank answers created for selected user(s)."
        })

    except Ticket.DoesNotExist:
        return JsonResponse({
            "responseCode": 404,
            "responseMessage": "Ticket not found."
        })
    except Exception as e:
        return JsonResponse({
            "responseCode": 500,
            "responseMessage": f"Error: {str(e)}"
        })




@login_required(login_url='login') 
def view_ticket_history(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    history = TicketStatusHistory.objects.filter(ticket=ticket).order_by('created_at')
    return render(request, 'tickets/ticket_history.html', {"ticket": ticket, "history": history})

@login_required(login_url='login') 
def dashboard_view(request):
    user = request.user

    # Superuser sees all tickets
    if user.is_superuser:
        tickets = Ticket.objects.all()
    else:
        tickets = Ticket.objects.filter(
            Q(created_by=user) |
            Q(assigned_users=user) |
            Q(reassignments__new_assigned_users=user)
        ).distinct()

    # Count tickets by status
    total_tickets = tickets.count()
    open_tickets = tickets.filter(status='open').count()
    pending_tickets = tickets.filter(status='pending').count()
    closed_tickets = tickets.filter(status='closed').count()
    reassigned_tickets = tickets.filter(reassignments__isnull=False).distinct().count()

    context = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'pending_tickets': pending_tickets,
        'closed_tickets': closed_tickets,
        'reassigned_tickets': reassigned_tickets,
    }

    return render(request, 'tickets/dashboard.html', context)




@login_required(login_url='login') 
@csrf_exempt
def add_registration(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body.decode('utf-8'))
        except Exception as e:
            return JsonResponse({
                "responseCode": 400,
                "responseMessage": f"Invalid JSON payload: {str(e)}"
            })

        ngo_name = data.get('association_name')
        rcn = data.get('rcn')
        map_status = data.get('mapStatus', 'Pending')

        if not ngo_name:
            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "NGO Name is required."
            })  
        now = timezone.now()
        try:
            registration = RegistrationsOther.objects.using('secondary').create(
                association_name=ngo_name,
                rcn=rcn,
                created_at=now,
                updated_at=now
                # mapStatus=map_status
            )

            return JsonResponse({
                "responseCode": 200,
                "responseMessage": "NGO added successfully!",
                "id": registration.id
            })

        except Exception as e:
            return JsonResponse({
                "responseCode": 500,
                "responseMessage": f"Error: {str(e)}"
            })

    return JsonResponse({
        "responseCode": 405,
        "responseMessage": "Invalid request method."
    })



@login_required(login_url="login")
def create_ticket_feedback(request, ticket_id):
    """
    Store feedback/chat user-wise between creator and assignees.
    - creator <-> assignee Ankur
    - creator <-> reassignee Ravinder
    Each pair's messages are separated via TicketLog.user / TicketLog.to_user.
    """
    def is_ajax(req):
        # Works on Django 2+ and older versions
        return (
            req.headers.get("x-requested-with") == "XMLHttpRequest"
            or req.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest"
        )

    ticket = get_object_or_404(Ticket, pk=ticket_id)
    user = request.user

    # Only allow POST
    if request.method != "POST":
        if is_ajax(request):
            return JsonResponse(
                {"success": False, "error": "Invalid HTTP method."},
                status=405,
            )
        return redirect("tickets:ticket_detail", pk=ticket.id)

    message = (request.POST.get("message") or "").strip()
    to_user_id = request.POST.get("to_user_id")

    # Basic validation
    if not message or not to_user_id:
        if is_ajax(request):
            return JsonResponse(
                {"success": False, "error": "Message and target user are required."},
                status=400,
            )
        return redirect("tickets:ticket_detail", pk=ticket.id)

    User = get_user_model()
    to_user = get_object_or_404(User, pk=to_user_id)

    # ========== PERMISSIONS ==========
    # Allowed senders:
    #   - ticket creator
    #   - any assigned user
    #   - any reassigned user for this ticket
    is_creator = (user == ticket.created_by)
    is_assignee = ticket.assigned_users.filter(id=user.id).exists()
    is_reassign_assignee = TicketReassignment.objects.filter(
        ticket=ticket, new_assigned_users=user
    ).exists()

    if not (is_creator or is_assignee or is_reassign_assignee or user.is_superuser):
        if is_ajax(request):
            return JsonResponse(
                {"success": False, "error": "You are not allowed to send feedback for this ticket."},
                status=403,
            )
        raise PermissionDenied("You are not allowed to send feedback for this ticket.")

    # Don't allow sending on closed ticket (optional)
    if ticket.status == "closed":
        if is_ajax(request):
            return JsonResponse(
                {"success": False, "error": "Ticket is closed."},
                status=400,
            )
        return redirect("tickets:ticket_detail", pk=ticket.id)

    # ========== CREATE LOG ==========
    log = TicketLog.objects.create(
        ticket=ticket,
        user=user,        # from_user
        to_user=to_user,  # target user (controls which tab/thread it appears in)
        message=message,
    )

    # ========== ATTACH FILES (optional) ==========
    uploaded_files = []
    for f in request.FILES.getlist("files"):
        fu = FileUpload.objects.create(
            ticket=ticket,
            log=log,
            user=user,
            file=f,
        )
        uploaded_files.append({
            "url": fu.file.url,
            "name": fu.file.name,
        })

    # If this is an AJAX request, return JSON instead of redirect
    if is_ajax(request):
        created_at_display = make_naive(log.created_at).strftime("%d-%m-%Y %H:%M")
        source = (
            "Ticket"
            if log.ticketReassignment_id is None
            else f"Reassign #{log.ticketReassignment_id}"
        )

        return JsonResponse(
            {
                "success": True,
                "message": log.message,
                "author_id": user.id,
                "author_name": f"{user.first_name} {user.last_name}".strip() or user.username,
                "created_at": created_at_display,
                "source": source,
                "files": uploaded_files,
            }
        )

    # Non-AJAX fallback: old behavior
    return redirect("tickets:ticket_detail", pk=ticket.id)



def search_ngo(request):
    q = (request.GET.get("q") or "").strip()

    if len(q) < 2:
        return JsonResponse([], safe=False)

    qs = (
        OtherServicesDocumentsLink.objects.using("secondary")
        .filter(association_name__icontains=q)
        .values("id", "association_name", "rcn","registration_application_id","association_address")
        .order_by("-id")[:20]
    )
    
    result = []
    for row in qs:
        result.append({
            "id": row["id"],
            "association_name": row["association_name"],
            "registration_application_id": row["registration_application_id"],
            "association_address": row["association_address"],
            "rcn": row["rcn"] if row.get("rcn") else "-",  # ✅ replace blank
        })
    return JsonResponse(list(result), safe=False)


def first_form(request):
    # return render(request, "tickets/new_forms/single_page.html")
    return render(request, "tickets/new_forms/read_pdf.html")



# ✅ adjust import to your project
# from accounts.models import CustomUser

STATIC_USER_ID=settings.USERID
THRESHOLD = 0.90
PREVIEW_CACHE_TTL = 60 * 30  # 30 minutes


# ============================================================
#  PAGE
# ============================================================
# def first_form(request):
#     return render(request, "tickets/first_form.html")


# ============================================================
#  NGO SEARCH API
# ============================================================
def search_ngo(request):
    q = (request.GET.get("q") or "").strip()
    if len(q) < 2:
        return JsonResponse([], safe=False)

    qs = (
        OtherServicesDocumentsLink.objects.using("secondary")
        .filter(association_name__icontains=q)
        .values("id", "association_name", "rcn","registration_application_id","association_address")
        .order_by("-id")[:20]
    )

    return JsonResponse([
        {
            "id": row["id"],
            "association_name": row["association_name"],
            "registration_application_id": row["registration_application_id"],
            "association_address": row["association_address"],
            "rcn": row.get("rcn") or "-",
        } for row in qs
    ], safe=False)


# ============================================================
#  SMALL HELPERS
# ============================================================
def _trim(s: str, n: int = 220) -> str:
    s = (s or "").strip()
    return s if len(s) <= n else s[:n] + "..."


def _read_json(request):
    try:
        return json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return {}


# ============================================================
#  QUESTION MATCHING HELPERS
# ============================================================
STOPWORDS = {
    "whether", "the", "is", "are", "any", "of", "to", "a", "an", "and", "or",
    "if", "yes", "no", "give", "details", "detail", "particulars", "please",
    "in", "on", "under", "before", "after", "for", "from", "with", "by", "as",
    "have", "has", "had", "been", "being", "was", "were", "it", "this", "that",
}

def normalize_text(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = s.replace("organisation", "organization")
    s = s.replace("costal", "coastal")

    qm = s.find("?")
    if qm != -1:
        s = s[:qm + 1]

    s = re.sub(r"\?\s*if yes\s*,?\s*(particulars|give details)\s*\.?\s*$", "?", s)
    s = re.sub(r"\b,\s*\d{4}\b", "", s)
    s = re.sub(r"\s+\?", "?", s)

    s = re.sub(r"[^a-z0-9\s\?\/\-]", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def tokenize(s: str):
    s = normalize_text(s).replace("?", "")
    tokens = re.findall(r"[a-z0-9]+", s)
    return [t for t in tokens if t not in STOPWORDS and len(t) > 1]

def token_set_similarity(a: str, b: str) -> float:
    a_set = set(tokenize(a))
    b_set = set(tokenize(b))
    if not a_set or not b_set:
        return 0.0

    inter = sorted(a_set & b_set)
    diff_a = sorted(a_set - b_set)
    diff_b = sorted(b_set - a_set)

    s_inter = " ".join(inter)
    s_a = " ".join(inter + diff_a)
    s_b = " ".join(inter + diff_b)

    return max(
        SequenceMatcher(None, s_inter, s_a).ratio(),
        SequenceMatcher(None, s_inter, s_b).ratio(),
        SequenceMatcher(None, s_a, s_b).ratio(),
    )

def build_question_bank():
    bank = []
    for q in Question.objects.all().only("id", "text"):
        db_norm = normalize_text(q.text)
        if db_norm:
            bank.append((q.id, q.text, db_norm))
    return bank

def match_question_id(excel_question: str, question_bank, threshold: float = THRESHOLD):
    excel_norm = normalize_text(excel_question)

    for qid, db_text, db_norm in question_bank:
        if excel_norm == db_norm:
            return qid, 1.0, db_text

    best_qid = None
    best_score = 0.0
    best_db_text = None

    for qid, db_text, _ in question_bank:
        sc = token_set_similarity(excel_question, db_text)
        if sc > best_score:
            best_score = sc
            best_qid = qid
            best_db_text = db_text

    if best_qid and best_score >= threshold:
        return best_qid, best_score, best_db_text

    return None, best_score, best_db_text


# ============================================================
#  OSDL FETCH
# ============================================================
def find_osdl_by_id(osdl_id: int):
    try:
        obj = OtherServicesDocumentsLink.objects.using("secondary").filter(id=osdl_id).first()
        if obj:
            return obj, "secondary"
    except ProgrammingError:
        pass

    try:
        obj = OtherServicesDocumentsLink.objects.using("default").filter(id=osdl_id).first()
        if obj:
            return obj, "default"
    except ProgrammingError:
        pass

    return None, None


# ============================================================
#  DATE HELPERS (optional)
# ============================================================
def _expand_2digit_year(yy: int, pivot: int = 26) -> int:
    return 2000 + yy if 0 <= yy <= pivot else 1900 + yy

def _parse_ddmmyyyy_or_ddmmyy(date_str: str):
    if not date_str:
        return None

    s = str(date_str).strip().replace("/", "-")
    m = re.search(r"\b(\d{1,2})-(\d{1,2})-(\d{2}|\d{4})\b", s)
    if not m:
        return None

    dd = int(m.group(1))
    mm = int(m.group(2))
    yy_raw = m.group(3)

    yyyy = _expand_2digit_year(int(yy_raw)) if len(yy_raw) == 2 else int(yy_raw)
    dt = datetime(yyyy, mm, dd, 0, 0, 0)
    return timezone.make_aware(dt, timezone.get_current_timezone())

def extract_excel_header_date(ws):
    top_cells = []
    for r in ws.iter_rows(min_row=1, max_row=12, values_only=True):
        for v in r:
            if v is None:
                continue
            top_cells.append(str(v))

    head_text = "\n".join(top_cells)

    m_any = re.search(r"\b(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})\b", head_text)
    if m_any:
        return _parse_ddmmyyyy_or_ddmmyy(m_any.group(1))
    return None


# ============================================================
#  EXCEL Q/A EXTRACT
# ============================================================
def find_qa_pairs_from_excel(ws):
    """
    Excel format:
      col A: 1 or 1. or 1) or 1-
      col B: Question...? Answer...
    """
    pairs = []

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 2:
            continue

        a, b = row[0], row[1]
        if a is None or b is None:
            continue

        a_str = str(a).strip()
        if not re.match(r"^\d{1,3}\s*[\.\)\-]?\s*$", a_str):
            continue

        block = re.sub(r"\s*\n\s*", " ", str(b).strip())
        if not block:
            continue

        qm = block.find("?")
        if qm != -1:
            q_text = block[:qm + 1].strip()
            ans_text = block[qm + 1:].strip()
        else:
            q_text = block.strip()
            ans_text = ""

        q_text = re.sub(r"\s+", " ", q_text).strip()
        ans_text = re.sub(r"\s+", " ", ans_text).strip()

        if len(q_text) >= 6:
            pairs.append((q_text, ans_text))

    # de-dup by normalized question
    seen = set()
    uniq = []
    for q, a in pairs:
        key = normalize_text(q)
        if key and key not in seen:
            seen.add(key)
            uniq.append((q, a))

    return uniq


# ============================================================
#  NGO NAME EXTRACTION (Respect Of Fix)
# ============================================================
def _norm_name(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[^a-z0-9\s]", "", s)
    return s.strip()

_BAD_HEADER_WORDS = {
    "sd", "sd/", "s/d", "additional deputy director", "deputy director",
    "ministry", "government", "regarding", "memo", "dated", "old",
    "seriatim", "reply", "yours", "faithfully"
}

def _guess_ngo_name_from_excel(ws) -> str:
    texts = []
    for r in ws.iter_rows(min_row=1, max_row=40, values_only=True):
        for v in r:
            if v is None:
                continue
            t = str(v).strip()
            if len(t) >= 3:
                texts.append(t)

    blob = "\n".join(texts)

    # ✅ Extract after "respect of" (your requirement)
    m = re.search(
        r"(?is)\b(?:in\s+)?respect\s+of\s+([A-Za-z][A-Za-z .&\-']{2,80})(?=,|\.|\n|$)",
        blob
    )
    if m:
        cand = re.sub(r"\s+", " ", m.group(1).strip())
        if 3 <= len(cand) <= 80:
            return cand

    # Fallback patterns
    m2 = re.search(r"(?is)\bname\s+of\s+(association|organisation|organization)\b\s*[:\-]?\s*([^\n,\.]{3,80})", blob)
    if m2:
        return re.sub(r"\s+", " ", m2.group(2).strip())

    m3 = re.search(r"(?is)\bassociation\s+name\b\s*[:\-]?\s*([^\n,\.]{3,80})", blob)
    if m3:
        return re.sub(r"\s+", " ", m3.group(1).strip())

    candidates = []
    for t in texts:
        nt = _norm_name(t)
        if len(nt) < 6 or len(nt) > 80:
            continue
        if any(w in nt for w in _BAD_HEADER_WORDS):
            continue
        letters = sum(ch.isalpha() for ch in t)
        digits = sum(ch.isdigit() for ch in t)
        if letters < 4 or digits > 6:
            continue
        candidates.append(t.strip())

    if candidates:
        candidates.sort(key=lambda x: len(x), reverse=True)
        return candidates[0]

    return ""

def _find_ngo_in_db(candidate_name: str):
    """
    Return: (id, association_name, db_alias, score)
    """
    cand_norm = _norm_name(candidate_name)
    if not cand_norm:
        return None, None, None, 0.0

    cand_tokens = set(cand_norm.split())
    first_token = candidate_name.split()[0]

    for alias in ("secondary", "default"):
        try:
            qs = (
                OtherServicesDocumentsLink.objects.using(alias)
                .filter(association_name__icontains=first_token)
                .values("id", "association_name")[:300]
            )

            best = None
            best_score = 0.0

            for row in qs:
                db_norm = _norm_name(row["association_name"])
                db_tokens = set(db_norm.split())

                # token containment => perfect match
                if cand_tokens.issubset(db_tokens):
                    return row["id"], row["association_name"], alias, 1.0

                sc = SequenceMatcher(None, cand_norm, db_norm).ratio()
                if sc > best_score:
                    best_score = sc
                    best = row

            if best and best_score >= 0.75:
                return best["id"], best["association_name"], alias, best_score

        except ProgrammingError:
            continue

    return None, None, None, 0.0


# ============================================================
#  SAFE WORKBOOK LOADER (NO "seek of closed file")
# ============================================================
def load_workbook_from_bytes(excel_bytes: bytes):
    return load_workbook(filename=BytesIO(excel_bytes), read_only=True, data_only=True)


# ============================================================
#  ENDPOINT 1: PREVIEW EXCEL
# ============================================================
@require_POST
def preview_from_excel(request):
    excel = request.FILES.get("excel")
    if not excel:
        return JsonResponse({"responseCode": 400, "responseMessage": "Excel file missing."}, status=400)

    if not excel.name.lower().endswith((".xlsx", ".xlsm", ".xltx", ".xltm")):
        return JsonResponse({"responseCode": 400, "responseMessage": "Please upload valid .xlsx file."}, status=400)

    # ✅ Read once
    excel_bytes = excel.read()

    # ✅ Save temp for token
    token = uuid.uuid4().hex
    tmp_path = f"tmp/excel_preview/{token}.xlsx"
    default_storage.save(tmp_path, ContentFile(excel_bytes))
    cache.set(f"EXCEL_PREVIEW:{token}", tmp_path, timeout=PREVIEW_CACHE_TTL)

    # ✅ Load workbook from bytes
    try:
        wb = load_workbook_from_bytes(excel_bytes)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        return JsonResponse({"responseCode": 400, "responseMessage": f"Invalid Excel: {str(e)}"}, status=400)

    # NGO detect
    candidate = _guess_ngo_name_from_excel(ws)
    ngo_id, ngo_name, db_used, score = _find_ngo_in_db(candidate) if candidate else (None, None, None, 0.0)

    # Date (optional)
    header_dt = extract_excel_header_date(ws)
    header_date = header_dt.strftime("%d-%m-%Y") if header_dt else ""

    # Q/A
    qa_pairs = find_qa_pairs_from_excel(ws)
    if not qa_pairs:
        return JsonResponse({"responseCode": 400, "responseMessage": "No questions found in Excel."}, status=400)

    question_bank = build_question_bank()
    if not question_bank:
        return JsonResponse({"responseCode": 400, "responseMessage": "Question bank is empty in DB."}, status=400)

    matched_rows = []
    unmatched = []

    for q_text, ans_text in qa_pairs:
        qid, scq, best_db = match_question_id(q_text, question_bank, threshold=THRESHOLD)
        ans_text = ans_text or ""

        if qid:
            matched_rows.append({
                "question_id": qid,
                "excel_question": q_text,
                "db_question": best_db,
                "match_percent": int(round(scq, 3) * 100),
                "answer": ans_text,
                "answer_missing": not bool(ans_text.strip()),
            })
        else:
            unmatched.append({
                "excel_question": _trim(q_text, 240),
                "best_score": round(scq, 3),
                "best_db_suggestion": _trim(best_db or "", 240),
            })

    # If any unmatched -> return error (your current behavior)
    if unmatched:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "Some Excel questions are not matched in DB Question bank.",
            "total_excel_questions": len(qa_pairs),
            "matched_count": len(matched_rows),
            "unmatched_count": len(unmatched),
            "matched_preview": matched_rows[:10],
            "unmatched_questions": unmatched[:50],
            "note": f"Add missing questions to DB or tune THRESHOLD (currently {THRESHOLD})."
        }, status=400)

    return JsonResponse({
        "responseCode": 200,
        "responseMessage": "Preview ready",
        "token": token,
        "header_date": header_date,
        "ngo": {
            "candidate_name": candidate or "",
            "matched": bool(ngo_id),
            "id": ngo_id,
            "association_name": ngo_name or "",
            "db_used": db_used or "",
            "match_percent": int(score * 100) if ngo_id else 0,
        },
        "rows": matched_rows,
    }, status=200)


# ============================================================
#  ENDPOINT 2: CREATE TICKET FROM PREVIEW (NO EXCEL READ HERE)
# ============================================================

def _compact_unmatched(unmatched, max_items=50):
    out = []
    for item in unmatched[:max_items]:
        out.append({
            "excel_question": _trim(item.get("excel_question", ""), 240),
            "best_score": item.get("best_score"),  # 0.0 - 1.0
            "best_db_suggestion": _trim(item.get("best_db_suggestion") or "", 240),
        })
    return out

@require_POST
def create_ticket_from_preview(request):
    payload = _read_json(request)
    if not payload:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "Invalid JSON payload."
        }, status=400)

    token = payload.get("token")
    ngo_id = payload.get("ngo_id")
    questions_data = payload.get("rows")  # ✅ ALWAYS defined

    # ---------- VALIDATIONS ----------
    if not token:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "Missing preview token."
        }, status=400)

    if not ngo_id:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "NGO not selected."
        }, status=400)

    if not isinstance(questions_data, list) or not questions_data:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "No questions data received."
        }, status=400)

    # ---------- CHECK PREVIEW EXISTS ----------
    tmp_path = cache.get(f"EXCEL_PREVIEW:{token}")
    if not tmp_path:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "Preview expired. Please upload Excel again."
        }, status=400)

    # ---------- VALIDATE NGO ----------
    osdl_obj, db_used = find_osdl_by_id(int(ngo_id))
    if not osdl_obj:
        return JsonResponse({
            "responseCode": 404,
            "responseMessage": "NGO not found."
        }, status=404)

    ngo_name = osdl_obj.association_name.strip()

    # ---------- VALIDATE ANSWERS ----------
    for row in questions_data:
        if not (row.get("answer") or "").strip():
            return JsonResponse({
                "responseCode": 400,
                "responseMessage": "All answers are required before creating ticket."
            }, status=400)

    # ---------- USERS ----------
    assigned_user = CustomUser.objects.filter(id=STATIC_USER_ID).first()
    if not assigned_user:
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": "Assigned user not found."
        }, status=400)

    created_by = request.user if request.user.is_authenticated else assigned_user

    # ---------- CREATE TICKET ----------
    with transaction.atomic():
        ticket = Ticket.objects.create(
            title=f"Auto Ticket - {ngo_name}",
            description="Created from Excel preview",
            created_by=created_by,
            models_name="NGO",
            models_id=osdl_obj.id,
            models_object=ngo_name,
            request_type="single",
            status="open",
        )

        ticket.assigned_users.add(assigned_user)
        TicketLog.objects.create(ticket=ticket, user=assigned_user)

        question_ids = [int(r["question_id"]) for r in questions_data]
        questions = Question.objects.filter(id__in=question_ids)
        qmap = {q.id: q for q in questions}

        for row in questions_data:
            qobj = qmap.get(int(row["question_id"]))
            if not qobj:
                continue

            TicketAnswer.objects.create(
                ticket=ticket,
                question=qobj,
                answered_by=assigned_user,
                answer=row["answer"],
                is_final=True,
                is_active=True,
            )

    # ---------- CLEANUP ----------
    cache.delete(f"EXCEL_PREVIEW:{token}")
    try:
        if default_storage.exists(tmp_path):
            default_storage.delete(tmp_path)
    except Exception:
        pass

    return JsonResponse({
        "responseCode": 200,
        "responseMessage": "Ticket created successfully.",
        "ticket_id": ticket.id
    })





def _clean(v, default="-"):
    v = "" if v is None else str(v).strip()
    return v if v else default

def user_renk(u):
    if not u:
        return "-"
    rank = (getattr(u, "rank", "") or "").strip()
    return rank if rank else "-"

def user_display_name(u):
    if not u:
        return "-"
    fn = getattr(u, "get_full_name", None)
    if callable(fn):
        name = (fn() or "").strip()
        if name:
            return name
    first = (getattr(u, "first_name", "") or "").strip()
    last = (getattr(u, "last_name", "") or "").strip()
    name = (first + " " + last).strip()
    if name:
        return name
    return _clean(getattr(u, "username", None), "-")



@login_required
@require_GET
def get_office_bearers(request):
    """
    Priority:
      1) if rcn exists -> CommitteeMembersData by fcra_registration_no
      2) else if application_id exists -> CommitteeDetails by application_id
      3) else -> 400

    Duplicate handling:
      - De-duplicate by member_name
      - Keep only one row per member_name
      - If duplicates exist, keep the row with the most filled fields
      - If tie, keep the first row found
    """

    def clean_or_dash(v):
        v = "" if v is None else str(v).strip()
        if not v or v.lower() in ("nan", "null", "none", "-"):
            return "-"
        return v

    def filled_score(item):
        return sum(1 for v in item.values() if v != "-")

    rcn = clean_or_dash(request.GET.get("ngoHiddenRcn"))
    application_id = clean_or_dash(request.GET.get("application_id"))

    if rcn == "-" and application_id == "-":
        return JsonResponse(
            {"ok": False, "message": "rcn or application_id is required"},
            status=400,
        )

    if rcn != "-":
        qs = (
            CommitteeMembersData.objects.using("secondary")
            .filter(fcra_registration_no=str(rcn))
            .values(
                "id",
                "member_name",
                "father_husband_name",
                "nationality",
                "occupation",
                "post_in_association",
                "relationship_with_other_member",
                "address",
                "application_id",
                "pan_no",
            )
        )
    else:
        qs = None

    if (qs is None) or (not qs.exists()):
        if application_id == "-":
            return JsonResponse(
                {"ok": False, "message": "No data found by RCN, and application_id not provided"},
                status=400,
            )

        qs = (
            CommitteeDetails.objects.using("secondary")
            .filter(application_id=str(application_id))
            .values(
                "id",
                "member_name",
                "father_husband_name",
                "nationality",
                "occupation",
                "post_in_association",
                "relationship_with_other_member",
                "address",
                "application_id",
                "pan_no",
            )
        )

    members_map = {}

    for row in qs:
        item = {
            "id": clean_or_dash(row.get("id")),
            "member_name": clean_or_dash(row.get("member_name")),
            "father_husband_name": clean_or_dash(row.get("father_husband_name")),
            "nationality": clean_or_dash(row.get("nationality")),
            "occupation": clean_or_dash(row.get("occupation")),
            "post_in_association": clean_or_dash(row.get("post_in_association")),
            "relationship_with_other_member": clean_or_dash(row.get("relationship_with_other_member")),
            "address": clean_or_dash(row.get("address")),
            "application_id": clean_or_dash(row.get("application_id")),
            "pan_no": clean_or_dash(row.get("pan_no")),
        }

        if all(v == "-" for v in item.values()):
            continue

        name_key = item["member_name"].strip().lower()

        if name_key == "-":
            name_key = f"blank_name_{item['id']}"

        if name_key not in members_map:
            members_map[name_key] = item
        else:
            if filled_score(item) > filled_score(members_map[name_key]):
                members_map[name_key] = item

    members = list(members_map.values())

    members.sort(key=lambda x: (
        (x["post_in_association"] or "-").lower(),
        (x["member_name"] or "-").lower()
    ))

    return JsonResponse({"ok": True, "members": members})
    
    
def _pick_office_bearers(app_id: str, osdl):
    """
    Return ALL committee members for this NGO (all rows),
    with these columns:
      member_name, father_husband_name, nationality, occupation,
      post_in_association, relationship_with_other_member,
      address, application_id, pan_no
    """

    def clean_or_dash(v):
        if v is None:
            return "-"
        s = str(v).strip()
        if not s or s.lower() in ("nan", "null", "none", "-"):
            return "-"
        return s

    application_id = clean_or_dash(getattr(osdl, "registration_application_id", None))
    rcn = clean_or_dash(getattr(osdl, "rcn", None))
    if rcn == "-":
        rcn = None

    fields = [
        "member_name",
        "father_husband_name",
        "nationality",
        "occupation",
        "post_in_association",
        "relationship_with_other_member",
        "address",
        "application_id",
        "pan_no",
        "mobile",
    ]

    # ✅ Primary: if rcn exists, pull from CommitteeMembersData (only if it really has these fields)
    qs = None
    if rcn:
        qs = (
            CommitteeMembersData.objects.using("secondary")
            .filter(fcra_registration_no=str(rcn))
        )
        # if this table doesn't have pan_no etc, fallback automatically
        try:
            rows = list(qs.values(*fields))
        except Exception:
            rows = []
    else:
        rows = []

    # ✅ Fallback: CommitteeDetails by application_id
    if not rows:
        rows = list(
            CommitteeDetails.objects.using("secondary")
            .filter(application_id=str(application_id))
            .values(*fields)
        )

    members = []
    for row in rows:
        item = {k: clean_or_dash(row.get(k)) for k in fields}
        # skip fully empty rows
        if all(v == "-" for v in item.values()):
            continue
        members.append(item)

    # optional ordering
    members.sort(key=lambda x: ((x["post_in_association"] or "-").lower(), (x["member_name"] or "-").lower()))
    return members



def _pick_office_bearers_for_pdf(app_id: str):
    """
    NOW: returns ALL committee members for this application_id (not only 3 posts)

    Returns: list of dicts like:
    [
      {
        "member_name": "...",
        "father_husband_name": "...",
        "nationality": "...",
        "occupation": "...",
        "post_in_association": "...",
        "relationship_with_other_member": "...",
        "address": "...",
        "application_id": "...",
        "pan_no": "..."
      },
      ...
    ]
    """

    def clean_or_dash(v):
        if v is None:
            return "-"
        s = str(v).strip()
        if not s or s.lower() in ("nan", "null", "none", "-"):
            return "-"
        return s

    qs = (
        CommitteeMembersData.objects.using("secondary")
        .filter(application_id=str(app_id))
        .values(
            "member_name",
            "father_husband_name",
            "nationality",
            "occupation",
            "post_in_association",
            "relationship_with_other_member",
            "address",
            "application_id",
            "pan_no",
        )
        .order_by("id")
    )

    members = []
    for row in qs:
        members.append({
            "member_name": clean_or_dash(row.get("member_name")),
            "father_husband_name": clean_or_dash(row.get("father_husband_name")),
            "nationality": clean_or_dash(row.get("nationality")),
            "occupation": clean_or_dash(row.get("occupation")),
            "post_in_association": clean_or_dash(row.get("post_in_association")),
            "relationship_with_other_member": clean_or_dash(row.get("relationship_with_other_member")),
            "address": clean_or_dash(row.get("address")),
            "application_id": clean_or_dash(row.get("application_id")),
            "pan_no": clean_or_dash(row.get("pan_no")),
        })

    # optional: sort by post then name
    members.sort(key=lambda x: ((x["post_in_association"] or "-").lower(), (x["member_name"] or "-").lower()))
    return members

def _pick_bank_details(app_id: str):
    """
    Bank details from utilization_bnk_details using application_id (=osdl.id)
    If multiple rows exist, take first non-empty for each field.
    """
    qs = UtilizationBnkDetails.objects.using("secondary").filter(application_id=str(app_id))

    def first_non_empty(field, default="-"):
        for obj in qs:
            val = _clean(getattr(obj, field, None), "")
            if val:
                return val
        return default

    bank_name = first_non_empty("bank_name", "-")
    bank_address = first_non_empty("bank_address", "-")
    account_no = first_non_empty("account_no", "-")
    return bank_name, bank_address, account_no


# ----------------------------
# main view
# ----------------------------

@login_required(login_url="login")
def ticket_assignee_questions_pdf(request, pk):
    ticket = (
        Ticket.objects
        .select_related("created_by")
        .prefetch_related("assigned_users")
        .get(pk=pk)
    )

    user = request.user
    is_creator = (user.id == ticket.created_by_id)
    is_assignee = ticket.assigned_users.filter(id=user.id).exists()
    is_reassign_assignee = TicketReassignment.objects.filter(ticket=ticket, new_assigned_users=user).exists()

    if not (is_creator or is_assignee or is_reassign_assignee or user.is_superuser):
        raise PermissionDenied()

    # -------------------------------------------------------
    # target user (creator can download for assignee using ?user_id=)
    # -------------------------------------------------------
    target_user = user
    if is_creator or user.is_superuser:
        user_id = request.GET.get("user_id")
        if user_id:
            User = get_user_model()
            target_user = User.objects.filter(id=user_id).first() or user
        else:
            first_assignee = ticket.assigned_users.first()
            if first_assignee:
                target_user = first_assignee

    # -------------------------------------------------------
    # Fetch OSDL (secondary DB) using ticket.models_id
    # -------------------------------------------------------
    osdl = None
    if getattr(ticket, "models_id", None):
        osdl = (
            OtherServicesDocumentsLink.objects.using("secondary")
            .filter(id=ticket.models_id)
            .first()
        )

    # -------------------------------------------------------
    # Questions list for this assignee
    # -------------------------------------------------------
    q_ids = list(
        TicketAnswer.objects
        .filter(ticket=ticket, answered_by=target_user)
        .values_list("question_id", flat=True)
        .distinct()
    )

    if q_ids:
        questions = list(Question.objects.filter(id__in=q_ids, status="active").order_by("id"))
    else:
        questions = list(Question.objects.filter(status="active").order_by("id")[:12])

    # -------------------------------------------------------
    # Organization info (from OSDL) - keep as you had
    # -------------------------------------------------------
    org_name = _clean(getattr(osdl, "association_name", None), "-")
    org_address = _clean(getattr(osdl, "association_address", None), "-")  # <-- you used osdl.address earlier, but your OSDL mapping shows association_address
    file_no = _clean(getattr(ticket, "file_no", None), "-")
    # dated = _clean(getattr(ticket, "dated", None), "-")  # if ticket doesn't have, keep "-"
    created_date = "-"
    if getattr(ticket, "created_at", None):
        try:
            created_date = ticket.created_at.strftime("%d-%m-%Y")
        except Exception:
            created_date = _clean(getattr(ticket, "created_at", None), "-")
        
    nature_of_association = _clean(getattr(osdl, "nature_of_association", None), "-")

    # -------------------------------------------------------
    # ✅ NEW: Office bearers + Bank details from new tables
    # -------------------------------------------------------
    app_id = str(ticket.models_id) if getattr(ticket, "models_id", None) else None

    if app_id:
        office_bearers = _pick_office_bearers_for_pdf(app_id)
        bank_name, bank_address, account_no = _pick_bank_details(app_id)
    else:
        office_bearers = []
        bank_name = bank_address = account_no = "-"


    # -------------------------------------------------------
    # From / To (rank)
    # -------------------------------------------------------
    creator_user = ticket.created_by
    assignee_user = target_user

    from_line = user_renk(creator_user)
    to_line = user_renk(assignee_user)

    # -------------------------------------------------------
    # Build PDF
    # -------------------------------------------------------
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"Ticket_{ticket.id}_Message",
    )

    styles = getSampleStyleSheet()
    P = ParagraphStyle("P", parent=styles["Normal"], fontName="Helvetica", fontSize=10, leading=14)
    PB = ParagraphStyle("PB", parent=P, fontName="Helvetica-Bold")
    PC = ParagraphStyle("PC", parent=PB, alignment=1)  # center
    PR = ParagraphStyle("PR", parent=PB, alignment=2)  # right

    story = []

    # Top header
    top = Table(
        [[Paragraph("Message", PC), Paragraph("SECRET/TIME BOUND(30 DAYS)", PR)]],
        colWidths=[110 * mm, 60 * mm],
    )
    top.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (0, 0), 0.25, colors.black),
        ("LINEBELOW", (1, 0), (1, 0), 0.25, colors.black),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(top)
    story.append(Spacer(1, 6))

    # Meta
    meta = Table(
        [
            [Paragraph(f"<b>From:</b> {_clean(from_line)}", P), Paragraph(f"<b>Dated :</b> {_clean(created_date)}", P)],
            [Paragraph(f"<b>To :</b> {_clean(to_line)}", P), ""],
            [Paragraph(f"<b>File No.</b> {_clean(file_no)}", P), ""],
        ],
        colWidths=[120 * mm, 50 * mm],
    )
    meta.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(meta)
    story.append(Spacer(1, 8))

    # NAME & ADDRESS
    story.append(Paragraph("<b>NAME &amp; ADDRESS OF THE ORGANIZATION</b>", PB))
    story.append(Paragraph(f"{org_name} &nbsp;&nbsp; {org_address}", P))
    story.append(Spacer(1, 6))

    # REGISTRATION UNDER SOCIETIES ACT
    story.append(Paragraph("<b>REGISTRATION UNDER SOCIETIES ACT:</b>", PB))
    story.append(Spacer(1, 4))

    # Office bearers block (from committee_members_data)
    story.append(Paragraph("<b>COMMITTEE MEMBERS</b>", PB))
    story.append(Spacer(1, 4))

    ob_rows = [[
        Paragraph("<b>S.No</b>", PB),
        Paragraph("<b>Post</b>", PB),
        Paragraph("<b>Name</b>", PB),
        Paragraph("<b>Father/Husband</b>", PB),
        Paragraph("<b>Nationality</b>", PB),
        Paragraph("<b>Occupation</b>", PB),
        Paragraph("<b>Relation</b>", PB),
        Paragraph("<b>PAN</b>", PB),
    ]]

    if office_bearers:
        for i, m in enumerate(office_bearers, start=1):
            ob_rows.append([
                Paragraph(str(i), P),
                Paragraph(_clean(m.get("post_in_association")), P),
                Paragraph(_clean(m.get("member_name")), P),
                Paragraph(_clean(m.get("father_husband_name")), P),
                Paragraph(_clean(m.get("nationality")), P),
                Paragraph(_clean(m.get("occupation")), P),
                Paragraph(_clean(m.get("relationship_with_other_member")), P),
                Paragraph(_clean(m.get("pan_no")), P),
            ])
    else:
        ob_rows.append([
            Paragraph("1", P),
            Paragraph("-", P),
            Paragraph("No committee members found.", P),
            Paragraph("-", P),
            Paragraph("-", P),
            Paragraph("-", P),
            Paragraph("-", P),
            Paragraph("-", P),
        ])

    office_table = Table(
        ob_rows,
        colWidths=[10*mm, 22*mm, 32*mm, 30*mm, 20*mm, 22*mm, 22*mm, 22*mm],
    )

    office_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(office_table)
    story.append(Spacer(1, 8))


    # BANK DETAILS (from utilization_bnk_details)
    story.append(Paragraph("<b>BANK DETAILS:</b>", PB))
    story.append(Spacer(1, 4))

    bank = Table(
        [
            [Paragraph("Bank Name:", P), Paragraph(_clean(bank_name), P)],
            [Paragraph("Address:", P), Paragraph(_clean(bank_address), P)],
            [Paragraph("Account No.:", P), Paragraph(_clean(account_no), P)],
            [Paragraph("Nature of Association", P), Paragraph(_clean(nature_of_association), P)],
        ],
        colWidths=[45 * mm, 125 * mm],
    )
    bank.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(bank)
    story.append(Spacer(1, 10))

    # Request paragraph
    story.append(Paragraph(
        "REQUEST TO SEND SERIATIM REPLY ON THE FOLLOWING POINTS ABOUT THIS ORGANISATION WHICH "
        "HAS APPLIED FOR FCRA REGISTRATION / PRIOR PERMISSION WITHIN 15 DAYS. "
        "A COPY OF THE APPLICATION RECEIVED FROM THE ASSOCIATION IS ENCLOSED HEREWITH:",
        P
    ))
    story.append(Spacer(1, 10))

    # Questions table
    q_rows = []
    for idx, q in enumerate(questions, start=1):
        q_rows.append([Paragraph(str(idx), P), Paragraph(_clean(getattr(q, "text", "")), P)])

    if not q_rows:
        q_rows = [[Paragraph("1", P), Paragraph("No questions assigned.", P)]]

    q_table = Table(q_rows, colWidths=[10 * mm, 160 * mm])
    q_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(q_table)
    story.append(Spacer(1, 20))

    # Signature
    sig = Table(
        [[Paragraph("Sd/-", PB)],
         [Paragraph("ADDITIONAL DEPUTY DIRECTOR", PB)]],
        colWidths=[170 * mm],
        hAlign="CENTER"
    )
    sig.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(sig)

    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"ticket_{ticket.id}_message.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp




def _clean(value, default="-"):
    if value is None:
        return default
    value = str(value).strip()
    return value if value else default


def safe_reportlab_text(value, default="-"):
    """
    Final safe text for ReportLab Paragraph:
    - plain text only
    - XML escaped
    - line breaks preserved
    """
    value = _clean(value, default)
    value = escape(str(value))
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = value.replace("\n", "<br/>")
    return value


def html_to_plain_text(value, default="-"):
    """
    Convert arbitrary user HTML / pasted content to plain readable text.
    No HTML is preserved.
    """
    if value is None:
        return default

    text = str(value).strip()
    if not text:
        return default

    replacements = [
        (r'(?is)<br\s*/?>', '\n'),
        (r'(?is)</p\s*>', '\n'),
        (r'(?is)</div\s*>', '\n'),
        (r'(?is)</li\s*>', '\n'),
        (r'(?is)</ul\s*>', '\n'),
        (r'(?is)</ol\s*>', '\n'),
        (r'(?is)</tr\s*>', '\n'),
        (r'(?is)</table\s*>', '\n'),
        (r'(?is)</h[1-6]\s*>', '\n'),
        (r'(?is)<li[^>]*>', '• '),
        (r'(?is)</td\s*>', ' | '),
        (r'(?is)</th\s*>', ' | '),
    ]
    for pattern, replacement in replacements:
        text = re.sub(pattern, replacement, text)

    text = re.sub(r'(?is)<script.*?>.*?</script>', ' ', text)
    text = re.sub(r'(?is)<style.*?>.*?</style>', ' ', text)
    text = re.sub(r'(?is)<[^>]+>', ' ', text)

    text = unescape(text).replace("\xa0", " ")

    text = re.sub(r'[ \t]+\n', '\n', text)
    text = re.sub(r'\n[ \t]+', '\n', text)
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'(\s*\|\s*){2,}', ' | ', text)
    text = re.sub(r'\n{3,}', '\n\n', text)

    lines = []
    for line in text.splitlines():
        line = re.sub(r'\s+', ' ', line).strip(" | \t")
        if line:
            lines.append(line)

    text = "\n".join(lines).strip()
    return text if text else default


def extract_html_table_data(value):
    """
    Parse simple pasted HTML tables using regex only.
    Returns:
        {
            "rows": [...],
            "has_header": True/False
        }
    or None
    """
    if value is None:
        return None

    html = str(value).strip()
    if not html:
        return None

    if not re.search(r'(?is)<table\b', html):
        return None

    table_match = re.search(r'(?is)<table\b.*?>.*?</table>', html)
    if not table_match:
        return None

    table_html = table_match.group(0)

    tr_matches = re.findall(r'(?is)<tr\b.*?>(.*?)</tr>', table_html)
    if not tr_matches:
        return None

    rows = []
    has_header = False

    for tr_html in tr_matches:
        row = []

        th_matches = re.findall(r'(?is)<th\b.*?>(.*?)</th>', tr_html)
        td_matches = re.findall(r'(?is)<td\b.*?>(.*?)</td>', tr_html)

        if th_matches:
            has_header = True
            cell_matches = th_matches
        else:
            cell_matches = td_matches

        for cell_html in cell_matches:
            cell_text = re.sub(r'(?is)<br\s*/?>', '\n', cell_html)
            cell_text = re.sub(r'(?is)</p\s*>', '\n', cell_text)
            cell_text = re.sub(r'(?is)</div\s*>', '\n', cell_text)
            cell_text = re.sub(r'(?is)<li[^>]*>', '• ', cell_text)
            cell_text = re.sub(r'(?is)</li\s*>', '\n', cell_text)
            cell_text = re.sub(r'(?is)<[^>]+>', ' ', cell_text)
            cell_text = unescape(cell_text).replace("\xa0", " ")
            cell_text = re.sub(r'[ \t]+', ' ', cell_text)
            cell_text = re.sub(r'\n{3,}', '\n\n', cell_text)
            cell_text = cell_text.strip()

            row.append(cell_text if cell_text else "-")

        if row:
            rows.append(row)

    if not rows:
        return None

    max_cols = max(len(r) for r in rows)
    normalized_rows = []
    for row in rows:
        normalized_rows.append(row + [""] * (max_cols - len(row)))

    return {
        "rows": normalized_rows,
        "has_header": has_header,
    }


def split_html_into_blocks(value):
    """
    Split answer HTML into ordered blocks:
    - {"type": "text", "content": "..."}
    - {"type": "table", "content": "<table>...</table>"}
    """
    if value is None:
        return [{"type": "text", "content": "-"}]

    html = str(value).strip()
    if not html:
        return [{"type": "text", "content": "-"}]

    blocks = []
    pattern = r'(?is)<table\b.*?>.*?</table>'

    last_end = 0
    for match in re.finditer(pattern, html):
        start, end = match.span()

        before_html = html[last_end:start]
        if before_html and before_html.strip():
            blocks.append({
                "type": "text",
                "content": before_html,
            })

        blocks.append({
            "type": "table",
            "content": match.group(0),
        })

        last_end = end

    after_html = html[last_end:]
    if after_html and after_html.strip():
        blocks.append({
            "type": "text",
            "content": after_html,
        })

    if not blocks:
        blocks.append({
            "type": "text",
            "content": html,
        })

    return blocks


@login_required(login_url="login")
def ticket_assignee_questions_pdf_with_answer(request, pk):
    ticket = (
        Ticket.objects
        .select_related("created_by", "signater")
        .prefetch_related("assigned_users")
        .get(pk=pk)
    )

    user = request.user
    is_creator = (user.id == ticket.created_by_id)
    is_assignee = ticket.assigned_users.filter(id=user.id).exists()
    is_reassign_assignee = TicketReassignment.objects.filter(
        ticket=ticket,
        new_assigned_users=user
    ).exists()

    if not (is_creator or is_assignee or is_reassign_assignee or user.is_superuser):
        raise PermissionDenied()

    # ---------------------------------------------------
    # Resolve target user
    # ---------------------------------------------------
    target_user = user

    if is_creator or user.is_superuser:
        user_id = request.GET.get("user_id")

        valid_user_ids = set(ticket.assigned_users.values_list("id", flat=True))
        valid_user_ids.update(
            TicketReassignment.objects.filter(ticket=ticket)
            .values_list("new_assigned_users__id", flat=True)
        )

        if user_id:
            try:
                user_id = int(user_id)
            except (TypeError, ValueError):
                user_id = None

            if user_id and user_id in valid_user_ids:
                User = get_user_model()
                picked = User.objects.filter(id=user_id).first()
                if picked:
                    target_user = picked
            else:
                first_assignee = ticket.assigned_users.first()
                if first_assignee:
                    target_user = first_assignee
        else:
            first_assignee = ticket.assigned_users.first()
            if first_assignee:
                target_user = first_assignee

    # ---------------------------------------------------
    # NGO details
    # ---------------------------------------------------
    osdl = None
    if getattr(ticket, "models_id", None):
        osdl = (
            OtherServicesDocumentsLink.objects.using("secondary")
            .filter(id=ticket.models_id)
            .first()
        )

    org_name = _clean(getattr(osdl, "association_name", None), "-")
    org_state = _clean(getattr(osdl, "association_state", None), "-")
    user_unit = _clean(getattr(target_user, "unit", None), "-")
    d1_file_no = _clean(getattr(ticket, "file_no", None), "-")
    mha_file_no = _clean(getattr(ticket, "mha_file_no", None), "-")
    
    created_date = "-"
    if getattr(ticket, "created_at", None):
        try:
            created_date = ticket.created_at.strftime("%d-%m-%Y")
        except Exception:
            created_date = _clean(getattr(ticket, "created_at", None), "-")


    updated_date = "-"
    if getattr(ticket, "updated_at", None):
        try:
            updated_date = ticket.updated_at.strftime("%d-%m-%Y")
        except Exception:
            updated_date = _clean(getattr(ticket, "updated_at", None), "-")
    
    # ---------------------------------------------------
    # Find relevant reassignment for this target user
    # ---------------------------------------------------
    reassignment_used = None
    reassign_id = request.GET.get("reassign_id")

    if reassign_id:
        reassignment_used = (
            TicketReassignment.objects
            .filter(ticket=ticket, id=reassign_id, new_assigned_users=target_user)
            .order_by("-created_at")
            .first()
        )
    else:
        reassignment_used = (
            TicketReassignment.objects
            .filter(ticket=ticket, new_assigned_users=target_user)
            .order_by("-created_at")
            .first()
        )
    final_updated_date = "-"
    if reassignment_used:
        final_answer = (
            ReassignmentAnswer.objects
            .filter(
                reassignment=reassignment_used,
                answered_by=target_user,
                is_active=True,
                is_final=True
            )
            .order_by("-created_at", "-id")
            .first()
        )

        final_date_obj = getattr(final_answer, "created_at", None) if final_answer else None

    else:
        final_answer = (
            TicketAnswer.objects
            .filter(
                ticket=ticket,
                answered_by=target_user,
                is_active=True,
                is_final=True
            )
            .order_by("-updated_at", "-id")
            .first()
        )

        final_date_obj = None
        if final_answer:
            final_date_obj = getattr(final_answer, "updated_at", None) or getattr(final_answer, "created_at", None)

    if final_date_obj:
        try:
            final_updated_date = final_date_obj.strftime("%d-%m-%Y")
        except Exception:
            final_updated_date = _clean(final_date_obj, "-")
    # ---------------------------------------------------
    # Find relevant reassignment for this target user
    # ---------------------------------------------------
    reassignment_used = None
    reassign_id = request.GET.get("reassign_id")

    if reassign_id:
        reassignment_used = (
            TicketReassignment.objects
            .filter(ticket=ticket, id=reassign_id, new_assigned_users=target_user)
            .order_by("-created_at")
            .first()
        )
    else:
        reassignment_used = (
            TicketReassignment.objects
            .filter(ticket=ticket, new_assigned_users=target_user)
            .order_by("-created_at")
            .first()
        )

    # ---------------------------------------------------
    # Build Q/A rows
    # ---------------------------------------------------
    qa_rows_data = []

    if reassignment_used:
        base_qs = (
            ReassignmentAnswer.objects
            .filter(
                reassignment=reassignment_used,
                answered_by=target_user,
            )
            .select_related("question")
            .order_by("question__id", "id")
        )

        grouped = defaultdict(list)
        for ra in base_qs:
            grouped[ra.question_id].append(ra)

        for qid in sorted(grouped.keys()):
            latest_ra = grouped[qid][-1]
            qa_rows_data.append({
                "q": html_to_plain_text(getattr(latest_ra.question, "text", None), "-"),
                "a": getattr(latest_ra, "answer", None),
            })

    else:
        base_qs = (
            TicketAnswer.objects
            .filter(ticket=ticket, answered_by=target_user)
            .select_related("question")
            .order_by("question__id", "version", "id")
        )

        grouped = defaultdict(list)
        for ta in base_qs:
            grouped[ta.question_id].append(ta)

        for qid in sorted(grouped.keys()):
            latest_ta = grouped[qid][-1]
            qa_rows_data.append({
                "q": html_to_plain_text(getattr(latest_ta.question, "text", None), "-"),
                "a": getattr(latest_ta, "answer", None),
            })

    # ---------------------------------------------------
    # Office file no
    # ---------------------------------------------------
    latest_action_trace = (
        TicketActionTrace.objects
        .filter(ticket=ticket)
        .order_by("-performed_at", "-id")
        .first()
    )
    office_file_no = _clean(getattr(latest_action_trace, "office_file_no", None), "-")

    # ---------------------------------------------------
    # Signatures
    # ---------------------------------------------------
    left_signater_rank = "-"
    if ticket.signater_id and ticket.signater:
        left_signater_rank = _clean(ticket.signater.ranks, "-")

    signater_from_row = (
        TicketSignater.objects
        .select_related("signater")
        .filter(ticket=ticket, user=target_user, action_type__iexact="from")
        .order_by("-id")
        .first()
    )

    right_signater_rank = "-"
    if signater_from_row and signater_from_row.signater:
        right_signater_rank = _clean(signater_from_row.signater.ranks, "-")

    # ---------------------------------------------------
    # PDF constants
    # ---------------------------------------------------
    FIXED_HEADER_LINE_1 = "Subsidiary Intelligence Bureau"
    FIXED_HEADER_LINE_2 = "(Ministry of Home Affairs)"
    FIXED_HEADER_LINE_3 = "Government of India"

    FIXED_RIGHT_SIGN_LABEL = "SIB"
    FIXED_LEFT_SIGN_LABEL = "IB Hqrc, New Delhi"

    # ---------------------------------------------------
    # PDF styles
    # ---------------------------------------------------
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=18 * mm,
        bottomMargin=14 * mm,
        title=f"Ticket_{ticket.id}_questions_answers",
    )

    styles = getSampleStyleSheet()

    normal = ParagraphStyle(
        "normal",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        alignment=TA_LEFT,
        spaceAfter=0,
        spaceBefore=0,
    )

    center_normal = ParagraphStyle(
        "center_normal",
        parent=normal,
        alignment=TA_CENTER,
        fontSize=10,
        leading=13,
    )

    center_bold = ParagraphStyle(
        "center_bold",
        parent=normal,
        fontName="Helvetica-Bold",
        alignment=TA_CENTER,
        fontSize=11,
        leading=14,
    )

    right_normal = ParagraphStyle(
        "right_normal",
        parent=normal,
        alignment=TA_RIGHT,
        fontSize=10,
        leading=13,
    )

    justify_style = ParagraphStyle(
        "justify_style",
        parent=normal,
        alignment=TA_JUSTIFY,
        fontName="Helvetica",
        fontSize=10,
        leading=14,
    )

    q_style = ParagraphStyle(
        "q_style",
        parent=normal,
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        alignment=TA_LEFT,
        leftIndent=0,
        firstLineIndent=0,
        spaceAfter=0,
    )

    a_style = ParagraphStyle(
        "a_style",
        parent=normal,
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        alignment=TA_LEFT,
        leftIndent=0,
        firstLineIndent=0,
        spaceAfter=4,
    )

    table_cell_style = ParagraphStyle(
        "table_cell_style",
        parent=normal,
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        alignment=TA_LEFT,
        leftIndent=0,
        firstLineIndent=0,
        spaceAfter=0,
    )

    left_signature_style = ParagraphStyle(
        "left_signature_style",
        parent=normal,
        alignment=TA_LEFT,
        fontName="Helvetica",
        fontSize=10,
        leading=13,
    )

    right_signature_style = ParagraphStyle(
        "right_signature_style",
        parent=normal,
        alignment=TA_RIGHT,
        fontName="Helvetica",
        fontSize=10,
        leading=13,
    )

    story = []

    # ---------------------------------------------------
    # Header
    # ---------------------------------------------------
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"<b>{safe_reportlab_text(office_file_no)}</b>", center_normal))
    story.append(Paragraph(f"<b>{safe_reportlab_text(FIXED_HEADER_LINE_1)}</b>", center_normal))
    story.append(Paragraph(f"<b>{safe_reportlab_text(FIXED_HEADER_LINE_2)}</b>", center_normal))
    story.append(Paragraph(f"<b>{safe_reportlab_text(FIXED_HEADER_LINE_3)}</b>", center_normal))
    story.append(Spacer(1, 18))

    story.append(Paragraph(f"{safe_reportlab_text(user_unit)}, {safe_reportlab_text(final_updated_date)}", right_normal))
    story.append(Spacer(1, 14))

    story.append(Paragraph("<b>Memorandum</b>", center_bold))
    story.append(Spacer(1, 14))

    body_width = 165 * mm

    # ---------------------------------------------------
    # Memo paragraph
    # ---------------------------------------------------
    memo_text = (
        f"Please refer to your Memo issued vide File No {d1_file_no} dated {created_date} "
        f"in reference to MHA File No {mha_file_no},"
        f"regarding FCRA registration in respect of "
        f"{org_name}. "
        f"Our seriatim reply is as under."
    )

    memo_block = [
        [Paragraph(safe_reportlab_text(memo_text), justify_style)],
    ]
    memo_table = Table(
        memo_block,
        colWidths=[body_width],
        hAlign="CENTER",
    )
    memo_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 1),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ]))
    story.append(memo_table)
    story.append(Spacer(1, 12))

    # ---------------------------------------------------
    # Q/A
    # ---------------------------------------------------
    if qa_rows_data:
        for idx, item in enumerate(qa_rows_data, start=1):
            story.append(Paragraph(f"{idx}. {safe_reportlab_text(item['q'])}", q_style))
            story.append(Spacer(1, 2))

            answer_blocks = split_html_into_blocks(item["a"])

            for block in answer_blocks:
                if block["type"] == "text":
                    plain_text = html_to_plain_text(block["content"], "").strip()
                    if plain_text:
                        story.append(Paragraph(
                            safe_reportlab_text(plain_text, "-"),
                            a_style
                        ))
                        story.append(Spacer(1, 4))

                elif block["type"] == "table":
                    table_data = extract_html_table_data(block["content"])

                    if table_data:
                        table_rows = []
                        for row in table_data["rows"]:
                            table_rows.append([
                                Paragraph(safe_reportlab_text(cell or "-"), table_cell_style)
                                for cell in row
                            ])

                        col_count = len(table_rows[0]) if table_rows else 1
                        col_width = body_width / max(col_count, 1)
                        col_widths = [col_width] * col_count

                        answer_table = Table(
                            table_rows,
                            colWidths=col_widths,
                            hAlign="LEFT",
                        )

                        style_cmds = [
                            ("GRID", (0, 0), (-1, -1), 0.8, colors.HexColor("#42566e")),
                            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("LEFTPADDING", (0, 0), (-1, -1), 8),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                            ("TOPPADDING", (0, 0), (-1, -1), 10),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
                        ]

                        if table_data["has_header"]:
                            style_cmds.extend([
                                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                                ("LINEBELOW", (0, 0), (-1, 0), 1.4, colors.HexColor("#42566e")),
                            ])

                        answer_table.setStyle(TableStyle(style_cmds))
                        story.append(answer_table)
                        story.append(Spacer(1, 4))

            story.append(Spacer(1, 6))
    else:
        story.append(Paragraph("1. No questions assigned.", q_style))
        story.append(Spacer(1, 2))
        story.append(Paragraph("-", a_style))

    # ---------------------------------------------------
    # Right signature
    # ---------------------------------------------------
    story.append(Spacer(1, 26))
    right_signature_table = Table(
        [[
            Paragraph("", normal),
            Paragraph(
                f"{safe_reportlab_text(right_signater_rank)}<br/>{safe_reportlab_text(FIXED_RIGHT_SIGN_LABEL)}",
                right_signature_style
            ),
        ]],
        colWidths=[90 * mm, 75 * mm],
        hAlign="CENTER",
    )
    right_signature_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
    ]))
    story.append(right_signature_table)

    # ---------------------------------------------------
    # Left signature
    # ---------------------------------------------------
    story.append(Spacer(1, 18))
    left_signature_table = Table(
        [[
            Paragraph(
                f"{safe_reportlab_text(left_signater_rank)}<br/>{safe_reportlab_text(FIXED_LEFT_SIGN_LABEL)}",
                left_signature_style
            ),
            Paragraph("", normal),
        ]],
        colWidths=[75 * mm, 90 * mm],
        hAlign="CENTER",
    )
    left_signature_table.setStyle(TableStyle([
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (0, 0), "LEFT"),
    ]))
    story.append(left_signature_table)

    def draw_secret(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica-Bold", 11)
        x = A4[0] - doc.rightMargin
        y = A4[1] - 10 * mm
        canvas.drawRightString(x, y, "SECRET")
        canvas.restoreState()

    doc.build(
        story,
        onFirstPage=draw_secret,
        onLaterPages=draw_secret,
    )

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"ticket_{ticket.id}_questions_answers.pdf"
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response








PARSE_CHUNK_SIZE = 5000
SQL_BATCH_SIZE = 1000

IMPORT_CONFIG = {
    "reg": {
        "model": RegistrationApplications,
        "unique_field": "registration_application_id",
        "fields": [
            "registration_application_id",
            "section_file_number",
            "form_submission_date",
            "darpan_id",
            "association_name",
            "association_address",
            "association_state",
            "association_district",
            "association_official_telephone",
            "email_id",
            "association_official_website",
            "pan_number",
            "nature_of_association",
            "religion",
            "registration_date",
            "act_registration_name",
            "act_registration_number",
            "place_of_act_registration",
            "association_chief_functionary_phone_number",
            "association_chief_functionary_mobile_number",
            "association_status",
            "bank_name",
            "bank_address",
            "bank_email_id",
            "ifsc_code",
            "account_number",
            "registration_applications_documents_location",
            "created_at",
        ],
        "log_table": "application_details",
        "log_type": "app",
        "timestamp_fields": ["created_at"],
        "use_hash_validation": False,
    },

    "com": {
        "model": CommitteeDetails,
        "unique_field": "uniq_hash",
        "fields": [
            "application_id",
            "member_name",
            "father_husband_name",
            "nationality",
            "occupation",
            "post_in_association",
            "relationship_with_other_member",
            "address",
            "address_foreign_country",
            "pan_no",
            "passport_number",
            "whether_indian_origin",
            "date_from_residing_india",
            "email",
            "landline",
            "mobile",
            "date_place_of_birth",
            "created_at",
            "updated_at",
        ],
        "hash_fields": [
            "application_id",
            "member_name",
            "father_husband_name",
            "nationality",
            "occupation",
            "post_in_association",
            "relationship_with_other_member",
            "address",
            "address_foreign_country",
            "pan_no",
            "passport_number",
            "whether_indian_origin",
            "date_from_residing_india",
            "email",
            "landline",
            "mobile",
            "date_place_of_birth",
        ],
        "log_table": "committee_details",
        "log_type": "app",
        "timestamp_fields": ["created_at", "updated_at"],
        "use_hash_validation": True,
    },

    "mem": {
        "model": CommitteeMembersData,
        "unique_field": "uniq_hash",
        "fields": [
            "fcra_registration_no",
            "association_name",
            "state",
            "district",
            "registration_date",
            "member_name",
            "father_husband_name",
            "nationality",
            "occupation",
            "post_in_association",
            "relationship_with_other_member",
            "address",
            "mobile",
            "application_id",
            "pan_no",
            "aadhaar_number",
            "created_at",
            "updated_at",
        ],
        "hash_fields": [
            "fcra_registration_no",
            "association_name",
            "state",
            "district",
            "registration_date",
            "member_name",
            "father_husband_name",
            "nationality",
            "occupation",
            "post_in_association",
            "relationship_with_other_member",
            "address",
            "mobile",
            "application_id",
            "pan_no",
            "aadhaar_number",
        ],
        "log_table": "committee_members_data",
        "log_type": "renewal",
        "timestamp_fields": ["created_at", "updated_at"],
        "use_hash_validation": True,
    },

    "logs": {
        "model": OtherServicesDocumentsLink_logs,
        "unique_field": "rcn",
        "fields": [
            "rcn",
            "registration_application_id",
            "section_file_number",
            "form_submission_date",
            "darpan_id",
            "association_name",
            "registration_date",
            "association_address",
            "association_state",
            "association_district",
            "association_official_telephone",
            "email_id",
            "association_official_website",
            "association_chief_functionary_phone_number",
            "association_chief_functionary_mobile_number",
            "act_registration_name",
            "act_registration_number",
            "date_of_act_registration",
            "place_of_act_registration",
            "pan_number",
            "nature_of_association",
            "religion",
            "bank_name",
            "bank_address",
            "bank_email_id",
            "ifsc_code",
            "account_number",
            "association_status",
            "cancelled_suspended_date",
            "cancelled_suspended_remarks",
            "cancellation_reason",
            "registration_certificate_location",
            "registration_applications_documents_location",
            "created_at",
        ],
        "log_table": "other_services_documents_link_logs",
        "log_type": "renewal",
        "timestamp_fields": ["created_at"],
        "use_hash_validation": False,
    },
}


def import_page(request):
    tables = [
        "application_details",
        "committee_details",
        "committee_members_data",
        "other_services_documents_link_logs",
    ]

    logs_by_table = {}
    for t in tables:
        logs_by_table[t] = (
            ImportSyncLogs.objects.using("secondary")
            .filter(table_name=t)
            .order_by("-updated_at", "-id")
            .first()
        )

    return render(
        request,
        "tickets/upload_reg_app.html",
        {"logs_by_table": logs_by_table},
    )


def save_import_log(kind, filename, inserted, skipped, updated=0):
    cfg = IMPORT_CONFIG[kind]
    now = timezone.now()

    obj = (
        ImportSyncLogs.objects.using("secondary")
        .filter(type=cfg["log_type"], table_name=cfg["log_table"])
        .first()
    )

    if obj:
        obj.file_name = filename
        obj.inserted = int(inserted or 0)
        obj.updated = int(updated or 0)
        obj.skipped = int(skipped or 0)
        obj.last_sync_at = now
        obj.updated_at = now
        obj.save(using="secondary")
        return obj

    return ImportSyncLogs.objects.using("secondary").create(
        type=cfg["log_type"],
        table_name=cfg["log_table"],
        file_name=filename,
        inserted=int(inserted or 0),
        updated=int(updated or 0),
        skipped=int(skipped or 0),
        last_sync_at=now,
        created_at=now,
        updated_at=now,
    )


def _chunks(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def _normalize_hash_value(value):
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    return str(value).strip()

def _normalize_registration_date(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")

    raw = str(value).strip()
    if not raw:
        return None

    raw = raw.replace("/", "-").upper()
    raw = raw.split(" ")[0]

    formats = [
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d-%b-%Y",
        "%d-%B-%Y",
        "%d-%m-%y",
        "%d-%b-%y",
        "%d-%B-%y",
    ]

    for fmt in formats:
        try:
            parsed = datetime.strptime(raw, fmt)
            return parsed.strftime("%Y-%m-%d")
        except ValueError:
            continue

    return None


def _build_row_uniq_hash(row, hash_fields):
    parts = []
    all_empty = True

    for field_name in hash_fields:
        val = _normalize_hash_value(row.get(field_name))
        if val != "":
            all_empty = False
        parts.append(val)

    if all_empty:
        return ""

    raw = "|".join(parts)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def _resolve_model_column(model_cls, field_name):
    try:
        return model_cls._meta.get_field(field_name).column
    except Exception:
        return field_name


def _get_model_sql_meta(model_cls, field_names, unique_field_name):
    table_name = model_cls._meta.db_table

    db_columns = []
    for fname in field_names:
        db_columns.append(_resolve_model_column(model_cls, fname))

    unique_column = _resolve_model_column(model_cls, unique_field_name)

    return table_name, db_columns, unique_column


def _get_existing_keys(cursor, table_name, unique_column, keys):
    existing = set()
    if not keys:
        return existing

    for part in _chunks(keys, SQL_BATCH_SIZE):
        placeholders = ",".join(["%s"] * len(part))
        sql = f"SELECT {unique_column} FROM {table_name} WHERE {unique_column} IN ({placeholders})"
        cursor.execute(sql, part)
        existing.update("" if row[0] is None else str(row[0]).strip() for row in cursor.fetchall())

    return existing


def _insert_rows_raw(cursor, table_name, db_columns, rows, ignore_duplicates=False):
    if not rows:
        return 0

    placeholders = ", ".join(["%s"] * len(db_columns))
    insert_keyword = "INSERT IGNORE" if ignore_duplicates else "INSERT"

    sql = f"""
        {insert_keyword} INTO {table_name} ({", ".join(db_columns)})
        VALUES ({placeholders})
    """

    values = []
    for row in rows:
        values.append(tuple(row.get(col) for col in db_columns))

    cursor.executemany(sql, values)

    if ignore_duplicates:
        return max(cursor.rowcount, 0)

    return len(rows)


@require_http_methods(["POST"])
@csrf_protect
def import_single_file(request):
    kind = (request.POST.get("kind") or "").strip()
    uploaded_file = request.FILES.get("excel")

    if kind not in IMPORT_CONFIG:
        return JsonResponse({"responseCode": 400, "responseMessage": "Invalid section"})

    if not uploaded_file:
        return JsonResponse({"responseCode": 400, "responseMessage": "File is required"})

    cfg = IMPORT_CONFIG[kind]
    model_cls = cfg["model"]

    table_name, db_columns, unique_column = _get_model_sql_meta(
        model_cls=model_cls,
        field_names=cfg["fields"],
        unique_field_name=cfg["unique_field"],
    )

    field_to_column = {
        fname: _resolve_model_column(model_cls, fname)
        for fname in cfg["fields"]
    }

    total_inserted = 0
    total_skipped = 0
    first_chunk_checked = False

    try:
        with transaction.atomic(using="secondary"):
            conn = connections["secondary"]

            with conn.cursor() as cursor:
                for chunk in iter_upload_chunks(
                    uploaded_file=uploaded_file,
                    kind=kind,
                    chunk_size=PARSE_CHUNK_SIZE,
                ):
                    if not chunk:
                        continue

                    if not first_chunk_checked:
                        rcn_err = validate_rcn_rule(kind, chunk[:1000])
                        if rcn_err:
                            return JsonResponse({
                                "responseCode": 400,
                                "responseMessage": rcn_err
                            })
                        first_chunk_checked = True

                    now = timezone.now()
                    prepared_rows = []
                    incoming_keys = []

                    for row in chunk:
                        row_data = dict(row)

                        for tf in cfg["timestamp_fields"]:
                            row_data[tf] = now

                        # Only for application_details: normalize registration_date
                        # Only for application_details: normalize date fields
                        if kind == "reg":
                            normalized_reg_date = _normalize_registration_date(row_data.get("registration_date"))
                            if row_data.get("registration_date") and not normalized_reg_date:
                                total_skipped += 1
                                continue
                            row_data["registration_date"] = normalized_reg_date

                            normalized_form_submission_date = _normalize_registration_date(row_data.get("form_submission_date"))
                            if row_data.get("form_submission_date") and not normalized_form_submission_date:
                                total_skipped += 1
                                continue
                            row_data["form_submission_date"] = normalized_form_submission_date

                        if cfg.get("use_hash_validation"):
                            unique_val = _build_row_uniq_hash(row_data, cfg["hash_fields"])
                            if not unique_val:
                                total_skipped += 1
                                continue
                            row_data["uniq_hash"] = unique_val
                        else:
                            unique_val = _normalize_hash_value(row_data.get(cfg["unique_field"]))
                            if not unique_val:
                                total_skipped += 1
                                continue

                        prepared_rows.append(row_data)
                        incoming_keys.append(unique_val)

                    existing_keys = _get_existing_keys(
                        cursor=cursor,
                        table_name=table_name,
                        unique_column=unique_column,
                        keys=incoming_keys,
                    )

                    final_rows = []
                    seen_in_same_file_chunk = set()

                    for row in prepared_rows:
                        unique_val = _normalize_hash_value(row.get(cfg["unique_field"]))

                        if unique_val in existing_keys or unique_val in seen_in_same_file_chunk:
                            total_skipped += 1
                            continue

                        seen_in_same_file_chunk.add(unique_val)

                        sql_row = {}
                        for fname in cfg["fields"]:
                            col = field_to_column[fname]
                            sql_row[col] = row.get(fname)

                        final_rows.append(sql_row)

                    for part in _chunks(final_rows, SQL_BATCH_SIZE):
                        inserted_count = _insert_rows_raw(
                            cursor=cursor,
                            table_name=table_name,
                            db_columns=db_columns,
                            rows=part,
                            ignore_duplicates=cfg.get("use_hash_validation", False),
                        )

                        total_inserted += inserted_count
                        total_skipped += max(len(part) - inserted_count, 0)

        save_import_log(kind, uploaded_file.name, total_inserted, total_skipped, 0)

        return JsonResponse({
            "responseCode": 200,
            "responseMessage": "Imported successfully",
            "kind": kind,
            "inserted": total_inserted,
            "skipped": total_skipped,
        })

    except Exception as e:
        save_import_log(kind, uploaded_file.name, 0, 0, 0)
        return JsonResponse({
            "responseCode": 500,
            "responseMessage": f"Import failed: {str(e)}"
        })


#****************************FOR PDF COMMAN CODE*******************



# ✅ adjust these imports to your actual app/models path
# If these models are in another app, import from there.
# from FIU.models import CommitteeMembersData, CommitteeDetails, UtilizationBnkDetails


# -------------------------------------------------------
# Small helpers
# -------------------------------------------------------

def _clean(v, default="-"):
    v = "" if v is None else str(v).strip()
    return v if v else default


def user_rank_unit(u):
    """
    From: <unit>(<rank>)
    """
    if not u:
        return "-"
    unit = _clean(getattr(u, "unit", None), "").strip()
    rank = _clean(getattr(u, "rank", None), "").strip()
    if unit and rank:
        return f"{unit}({rank})"
    return unit or rank or _clean(getattr(u, "username", None), "-")


def _check_ticket_permission(ticket: Ticket, user):
    is_creator = (user.id == ticket.created_by_id)
    is_assignee = ticket.assigned_users.filter(id=user.id).exists()
    is_reassign_assignee = TicketReassignment.objects.filter(ticket=ticket, new_assigned_users=user).exists()

    if not (is_creator or is_assignee or is_reassign_assignee or user.is_superuser):
        raise PermissionDenied("You do not have permission to view this ticket.")
    return is_creator


def _resolve_target_user(ticket: Ticket, request, user, is_creator: bool):
    """
    Creator/superuser:
      - can pass ?user_id=123 to generate for that user
      - else generates for first assignee
    Others:
      - generates for current user
    """
    target_user = user

    if is_creator or user.is_superuser:
        user_id = request.GET.get("user_id")
        if user_id:
            User = get_user_model()
            target_user = User.objects.filter(id=user_id).first() or user
        else:
            first_assignee = ticket.assigned_users.first()
            if first_assignee:
                target_user = first_assignee

    return target_user


def _get_osdl(ticket: Ticket):
    """
    NGO row from secondary DB.
    """
    if not getattr(ticket, "models_id", None):
        return None
    return (
        OtherServicesDocumentsLink.objects.using("secondary")
        .filter(id=ticket.models_id)
        .first()
    )




from typing import List, Dict

def _pick_office_bearers11111(osdl) -> List[Dict[str, str]]:
    """
    UPDATED (same name): returns ALL committee members for this NGO.

    Priority:
      1) if osdl.rcn exists -> CommitteeMembersData by fcra_registration_no
      2) fallback -> CommitteeDetails by application_id = osdl.registration_application_id (or osdl.id)

    Returns:
      [
        {
          "member_name": "...",
          "father_husband_name": "...",
          "nationality": "...",
          "occupation": "...",
          "post_in_association": "...",
          "relationship_with_other_member": "...",
          "address": "...",
          "application_id": "...",
          "pan_no": "..."
        },
        ...
      ]
    """

    def clean_or_dash(v):
        v = "" if v is None else str(v).strip()
        if not v or v.lower() in ("nan", "null", "none", "-"):
            return "-"
        return v

    rcn = clean_or_dash(getattr(osdl, "rcn", None))
    application_id = clean_or_dash(getattr(osdl, "registration_application_id", None))

    qs = None

    # 1) primary: renewal data by RCN
    if rcn != "-":
        qs = (
            CommitteeMembersData.objects.using("secondary")
            .filter(fcra_registration_no=str(rcn))
            .values(
                "member_name",
                "father_husband_name",
                "nationality",
                "occupation",
                "post_in_association",
                "relationship_with_other_member",
                "address",
                "application_id",
                "pan_no",
            )
        )

    # 2) fallback: application committee details by application_id (or osdl.id)
    if (qs is None) or (not qs.exists()):
        fallback_app_id = application_id if application_id != "-" else clean_or_dash(getattr(osdl, "id", None))
        qs = (
            CommitteeDetails.objects.using("secondary")
            .filter(application_id=str(fallback_app_id))
            .values(
                "member_name",
                "father_husband_name",
                "nationality",
                "occupation",
                "post_in_association",
                "relationship_with_other_member",
                "address",
                "application_id",
                "pan_no",
            )
        )

    members = []
    for row in qs:
        item = {
            "member_name": clean_or_dash(row.get("member_name")),
            "father_husband_name": clean_or_dash(row.get("father_husband_name")),
            "nationality": clean_or_dash(row.get("nationality")),
            "occupation": clean_or_dash(row.get("occupation")),
            "post_in_association": clean_or_dash(row.get("post_in_association")),
            "relationship_with_other_member": clean_or_dash(row.get("relationship_with_other_member")),
            "address": clean_or_dash(row.get("address")),
            "application_id": clean_or_dash(row.get("application_id")),
            "pan_no": clean_or_dash(row.get("pan_no")),
        }

        # skip fully blank rows
        if all(v == "-" for v in item.values()):
            continue

        members.append(item)

    # optional: order by role then name
    members.sort(key=lambda x: (
        (x["post_in_association"] or "-").lower(),
        (x["member_name"] or "-").lower()
    ))

    return members


def _pick_bank_details(osdl) -> Tuple[str, str, str]:
    """
    Bank from UtilizationBnkDetails by application_id = osdl.id (as you designed in import).
    """
    app_id = str(getattr(osdl, "id", "") or "")
    if not app_id:
        return "-", "-", "-"

    qs = UtilizationBnkDetails.objects.using("secondary").filter(application_id=app_id)

    def first_non_empty(field, default="-"):
        for obj in qs:
            val = _clean(getattr(obj, field, None), "")
            if val:
                return val
        return default

    bank_name = first_non_empty("bank_name", "-")
    bank_address = first_non_empty("bank_address", "-")
    account_no = first_non_empty("account_no", "-")

    return bank_name, bank_address, account_no


# -------------------------------------------------------
#  PDF landing page
# -------------------------------------------------------

@login_required(login_url="login")
def ticket_pdfs_page(request, pk):
    ticket = get_object_or_404(
        Ticket.objects.select_related("created_by").prefetch_related("assigned_users"),
        pk=pk
    )
    is_creator = _check_ticket_permission(ticket, request.user)

    target_user = _resolve_target_user(ticket, request, request.user, is_creator)
    osdl = _get_osdl(ticket)

    org_name = _clean(getattr(osdl, "association_name", None), "-")
    org_address = _clean(getattr(osdl, "association_address", None), "-")

    office_bearers = []
    bank_name = bank_address = account_no = "-"

    if osdl:
        office_bearers = _pick_office_bearers11111(osdl)   # <-- ALL members
        bank_name, bank_address, account_no = _pick_bank_details(osdl)

    return render(request, "tickets/ticket_pdfs.html", {
        "ticket": ticket,
        "target_user": target_user,
        "org_name": org_name,
        "org_address": org_address,
        "office_bearers": office_bearers,
        # "vice_president": vice_president,
        # "secretary": secretary,
        "bank_name": bank_name,
        "bank_address": bank_address,
        "account_no": account_no,
        "is_creator": is_creator,
    })


# -------------------------------------------------------
# PDF 1: LETTER TO SIB  (Questions list)
# -------------------------------------------------------



# -------------------------------------------------
# 1️⃣ Safe Display Name
# -------------------------------------------------
def _display_name(user):
    if not user:
        return ""

    # call only if it's actually a function/method
    get_full_name = getattr(user, "get_full_name", None)
    if callable(get_full_name):
        full = (get_full_name() or "").strip()
        if full:
            return full

    # fallback: field-based
    first = (getattr(user, "first_name", "") or "").strip()
    last = (getattr(user, "last_name", "") or "").strip()
    if first or last:
        return f"{first} {last}".strip()

    username = (getattr(user, "username", "") or "").strip()
    if username:
        return username

    email = (getattr(user, "email", "") or "").strip()
    if email:
        return email

    # final fallback: your custom __str__ like "RAVINDER DIWAKAR (976...)"
    return str(user)

# -------------------------------------------------
# 2️⃣ Get Current Assignee
# -------------------------------------------------
def _get_current_assignee(ticket):
    """
    Since assigned_users is ManyToMany,
    we define ONE assignee as primary.
    """
    return ticket.assigned_users.order_by("id").first()


# -------------------------------------------------
# 3️⃣ Core Logic: From / To Flip
# -------------------------------------------------
def _get_from_to_users(ticket, downloader):
    """
    Case 1:
        Assignee downloads
        From = Creator
        To   = Assignee

    Case 2:
        Creator downloads
        From = Assignee
        To   = Creator
    """

    creator = ticket.created_by

    # Is downloader one of the assigned users?
    is_assignee = ticket.assigned_users.filter(id=downloader.id).exists()

    if is_assignee:
        assignee = downloader
    else:
        assignee = _get_current_assignee(ticket)

    to_user = downloader

    if downloader.id == creator.id:
        # Creator downloading
        from_user = assignee
    else:
        # Assignee downloading
        from_user = creator

    # safety fallback
    if from_user is None:
        from_user = creator

    return from_user, to_user


# -------------------------------------------------
# 4️⃣ View: Letter to SIB
# -------------------------------------------------
@login_required
def pdf_letter_to_sib(request, pk):
    """
    IMPORTANT:
    Your URL is passing <int:pk>
    So function MUST accept pk (not ticket_id).
    """

    ticket = get_object_or_404(Ticket, pk=pk)

    from_user, to_user = _get_from_to_users(ticket, request.user)

    context = {
        "ticket": ticket,
        "from_name": _display_name(from_user),
        "to_name": _display_name(to_user),
    }

    # If using HTML template:
    return render(request, "tickets/pdf_letter_to_sib.html", context)

    # If generating actual PDF,
    # keep your existing PDF generation code here


def _pick_one_assignee(ticket):
    """
    Ticket.assigned_users is ManyToMany.
    Choose one assignee. If multiple assigned, first() is used.
    """
    return ticket.assigned_users.order_by("id").first()


def _from_to_lines_for_letter(ticket, downloader):
    """
    Your rule:
    - If assignee downloads -> From=creator, To=assignee(downloader)
    - If creator downloads  -> From=assignee, To=creator(downloader)
    """
    creator = ticket.created_by
    to_user = downloader

    if ticket.assigned_users.filter(id=downloader.id).exists():
        assignee = downloader
    else:
        assignee = _pick_one_assignee(ticket)

    if creator and downloader and downloader.id == creator.id:
        from_user = assignee or creator
    else:
        from_user = creator

    from_line = user_rank_unit(from_user)
    to_line = user_rank_unit(to_user)

    return from_line, to_line, to_user


def _get_dynamic_questions_for_letter(ticket, answered_by_user):
    """
    Returns only dynamic questions actually linked with this ticket + user.
    No static fallback.
    """
    ans_qs = (
        TicketAnswer.objects
        .filter(ticket=ticket, answered_by=answered_by_user)
        .select_related("question")
        .order_by("question_id", "version", "id")
    )

    if not ans_qs.exists():
        return []

    seen = set()
    questions = []

    for a in ans_qs:
        q = getattr(a, "question", None)
        if not q:
            continue
        if q.id in seen:
            continue
        seen.add(q.id)
        questions.append(q)

    return questions


@login_required(login_url="login")
def pdf_letter_to_sib(request, pk):
    ticket = (
        Ticket.objects.select_related("created_by", "signater")
        .prefetch_related("assigned_users")
        .get(pk=pk)
    )

    is_creator = _check_ticket_permission(ticket, request.user)
    target_user = _resolve_target_user(ticket, request, request.user, is_creator)

    osdl = _get_osdl(ticket)

    org_name = _clean(getattr(osdl, "association_name", None), "-")
    org_address = _clean(getattr(osdl, "association_address", None), "-")

    file_no = _clean(getattr(ticket, "file_no", None), "-")

    # dated_val = getattr(ticket, "dated", None)
    # if dated_val:
    #     dated = _clean(dated_val, "-")
    # else:
    #     dated = timezone.now().strftime("%d-%m-%Y")
    
    dated = "-"
    if getattr(ticket, "created_at", None):
        try:
            dated = ticket.created_at.strftime("%d-%m-%Y")
        except Exception:
            dated = _clean(getattr(ticket, "created_at", None), "-")


    # From / To based on downloader
    to_line,from_line,questions_user = _from_to_lines_for_letter(ticket, request.user)

    # Dynamic questions only, same idea as MHA reply
    questions = _get_dynamic_questions_for_letter(ticket, questions_user)

    # If downloader has no answers but target_user has answers, use target_user as backup
    if not questions and target_user:
        questions = _get_dynamic_questions_for_letter(ticket, target_user)

    signater_text = "-"
    if ticket.signater_id and ticket.signater:
        signater_text = _clean(getattr(ticket.signater, "ranks", None), "-")

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
        title=f"LETTER_TO_SIB_Ticket_{ticket.id}",
    )

    styles = getSampleStyleSheet()
    P = ParagraphStyle(
        "P",
        parent=styles["Normal"],
        fontName="Helvetica",
        fontSize=10,
        leading=14,
        alignment=TA_JUSTIFY,
    )
    PB = ParagraphStyle("PB", parent=P, fontName="Helvetica-Bold")
    PC = ParagraphStyle("PC", parent=PB, alignment=1)
    PR = ParagraphStyle("PR", parent=PB, alignment=2)

    story = []

    top = Table(
        [[Paragraph("Message", PC), Paragraph("SECRET/TIME BOUND(30 DAYS)", PR)]],
        colWidths=[110 * mm, 60 * mm]
    )
    top.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (0, 0), 0.25, colors.black),
        ("LINEBELOW", (1, 0), (1, 0), 0.25, colors.black),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(top)
    story.append(Spacer(1, 6))

    meta = Table(
        [
            [Paragraph(f"<b>From:</b> {_clean(from_line)}", P), Paragraph(f"<b>Dated:</b> {_clean(dated)}", P)],
            [Paragraph(f"<b>To :</b> {_clean(to_line)}", P), ""],
            [Paragraph(f"<b>File No.</b> {_clean(file_no)}", P), ""],
        ],
        colWidths=[120 * mm, 50 * mm],
    )
    meta.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(meta)
    story.append(Spacer(1, 8))

    story.append(Paragraph("<b>NAME &amp; ADDRESS OF THE ORGANIZATION</b>", PB))
    story.append(Paragraph(f"{safe_reportlab_text(org_name)} {safe_reportlab_text(org_address)}", P))
    story.append(Spacer(1, 6))

    story.append(Paragraph("<b>REGISTRATION UNDER SOCIETIES ACT:</b>", PB))
    story.append(Spacer(1, 4))

    story.append(Paragraph(
        "REQUEST TO SEND SERIATIM REPLY ON THE FOLLOWING POINTS ABOUT THIS ORGANISATION WHICH "
        "HAS APPLIED FOR FCRA REGISTRATION / PRIOR PERMISSION WITHIN 15 DAYS. "
        "A COPY OF THE APPLICATION RECEIVED FROM THE ASSOCIATION IS ENCLOSED HEREWITH:",
        P
    ))
    story.append(Spacer(1, 10))

    q_rows = []
    for idx, q in enumerate(questions, start=1):
        q_rows.append([
            Paragraph(f"{idx}.", P),
            Paragraph(safe_reportlab_text(_clean(getattr(q, "text", ""))), P)
        ])

    if not q_rows:
        q_rows = [[Paragraph("1.", P), Paragraph("No questions assigned.", P)]]

    q_table = Table(q_rows, colWidths=[10 * mm, 160 * mm])
    q_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(q_table)
    story.append(Spacer(1, 20))

    sig = Table(
        [[Paragraph("Sd/-", PB)],
         [Paragraph(safe_reportlab_text(signater_text), PB)]],
        colWidths=[170 * mm],
        hAlign="CENTER"
    )
    sig.setStyle(TableStyle([
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(sig)

    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"LETTER_TO_SIB_ticket_{ticket.id}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp






# -------------------------------------------------------
# ✅ PDF: SIB REPLY (Questions + Answers)
# -------------------------------------------------------

@login_required(login_url="login")
def pdf_sib_reply(request, pk):
    ticket = (
        Ticket.objects.select_related("created_by")
        .prefetch_related("assigned_users")
        .get(pk=pk)
    )
    is_creator = _check_ticket_permission(ticket, request.user)
    target_user = _resolve_target_user(ticket, request, request.user, is_creator)

    osdl = _get_osdl(ticket)
    org_name = _clean(getattr(osdl, "association_name", None), "-")
    file_no = _clean(getattr(ticket, "file_no", None), "-")

    # Pull latest answers for that user (use your own is_final/is_active rules)
    ans_qs = (
        TicketAnswer.objects
        .filter(ticket=ticket, answered_by=target_user)
        .select_related("question")
        .order_by("question_id", "-version", "-id")
    )

    # de-dup per question_id -> keep latest
    latest_by_q = {}
    for a in ans_qs:
        if a.question_id not in latest_by_q:
            latest_by_q[a.question_id] = a

    rows = list(latest_by_q.values())
    rows.sort(key=lambda x: x.question_id)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    P = ParagraphStyle("P", parent=styles["Normal"], fontName="Helvetica", fontSize=10, leading=14)
    PB = ParagraphStyle("PB", parent=P, fontName="Helvetica-Bold")
    PC = ParagraphStyle("PC", parent=PB, alignment=1)

    story = []
    story.append(Paragraph("REPLY FROM SIB", PC))
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"<b>NGO:</b> {org_name}", P))
    story.append(Paragraph(f"<b>File No:</b> {file_no}", P))
    story.append(Paragraph(f"<b>Reply By:</b> {_clean(user_rank_unit(target_user))}", P))
    story.append(Spacer(1, 10))

    table_rows = [[Paragraph("<b>Q.No</b>", P), Paragraph("<b>Question</b>", P), Paragraph("<b>Answer</b>", P)]]

    if rows:
        for i, a in enumerate(rows, start=1):
            qtext = _clean(getattr(a.question, "text", ""), "-")
            ans = _clean(getattr(a, "answer", ""), "-")
            table_rows.append([Paragraph(str(i), P), Paragraph(qtext, P), Paragraph(ans, P)])
    else:
        table_rows.append([Paragraph("1", P), Paragraph("No answers found.", P), Paragraph("-", P)])

    t = Table(table_rows, colWidths=[12*mm, 88*mm, 70*mm])
    t.setStyle(TableStyle([
        ("GRID", (0,0), (-1,-1), 0.25, colors.black),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("BACKGROUND", (0,0), (-1,0), colors.lightgrey),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]))
    story.append(t)

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    filename = f"SIB_REPLY_ticket_{ticket.id}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


# -------------------------------------------------------
# PDF: RCI SLIP (matches your layout)
# -------------------------------------------------------

@login_required(login_url="login")
def pdf_rci_slip(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    _check_ticket_permission(ticket, request.user)

    osdl = _get_osdl(ticket)
    org_name = _clean(getattr(osdl, "association_name", None), "-")
    org_address = _clean(getattr(osdl, "association_address", None), "-")
    file_no = _clean(getattr(ticket, "file_no", None), "-")

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    W, H = A4

    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, H - 18 * mm, "RCI SLIP (FOR SUBJECT)")
    c.drawRightString(W - 20 * mm, H - 18 * mm, "SECRET")

    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(W / 2, H - 28 * mm, 'ANNEXURE "A"')

    c.setFont("Helvetica", 11)
    c.drawString(20 * mm, H - 40 * mm, "RCI NO.________")
    c.drawString(20 * mm, H - 48 * mm, "SL.NO._________")
    c.drawString(20 * mm, H - 58 * mm, "IDENTIFICATION NO.__________________________________________________________________")
    c.drawString(20 * mm, H - 68 * mm, "MAJOR CLASSIFICATON_______________________________________________________________")
    c.drawString(20 * mm, H - 78 * mm, "MINOR CLASSIFICATION_______________________________________________________________")

    c.drawString(20 * mm, H - 90 * mm, "KEYWORD.1.____________________________")
    c.drawString(115 * mm, H - 90 * mm, f"FILE NO.1.{file_no}")

    c.drawString(20 * mm, H - 98 * mm, "KEYWORD.2.____________________________")
    c.drawString(115 * mm, H - 98 * mm, "FILE NO.2._________________________")

    c.drawString(20 * mm, H - 106 * mm, "KEYWORD.3.____________________________")
    c.drawString(115 * mm, H - 106 * mm, "FILE NO.3__________________________")

    c.drawString(20 * mm, H - 116 * mm, "YEAR OF PREVIOUS INDEXING_____________")
    c.drawString(115 * mm, H - 116 * mm, "FILE NO.4._________________________")

    c.drawString(20 * mm, H - 124 * mm, "SUBSEQUENT INDEXING YEAR_____________")
    c.drawString(115 * mm, H - 124 * mm, "FILE NO.5._________________________")

    c.drawString(20 * mm, H - 132 * mm, "YEAR OF INDEXING _______________________")
    c.drawString(115 * mm, H - 132 * mm, "FILE NO.6._________________________")

    c.drawString(20 * mm, H - 146 * mm, "Initial of D.A")
    c.drawString(20 * mm, H - 156 * mm, "ACTION IN INDEX BRANCH")

    c.line(20 * mm, H - 160 * mm, W - 20 * mm, H - 160 * mm)
    c.line(20 * mm, H - 190 * mm, W - 20 * mm, H - 190 * mm)

    c.drawString(20 * mm, H - 206 * mm, "ACTION IN INDEX BRANCH")
    c.drawString(20 * mm, H - 216 * mm, "The subject has been indexedSUBJECT MATTER:")
    c.line(20 * mm, H - 220 * mm, W - 20 * mm, H - 220 * mm)

    c.setFont("Helvetica", 10)
    c.drawString(
        20 * mm, H - 238 * mm,
        "FOR REFERENCING: Please refer to the above mentioned subject which is not borne on Branch File Registers for the last"
    )
    c.drawString(20 * mm, H - 246 * mm, "five years.")

    c.setFont("Helvetica-Bold", 11)
    c.drawString(20 * mm, H - 258 * mm, org_name)

    c.setFont("Helvetica", 10)
    addr = org_address.replace("\n", " ").strip()
    addr_words = addr.split()
    lines = []
    cur = []
    for w in addr_words:
        cur.append(w)
        if len(" ".join(cur)) > 70:
            cur.pop()
            lines.append(" ".join(cur))
            cur = [w]
    if cur:
        lines.append(" ".join(cur))

    y = H - 268 * mm
    for line in lines[:3]:
        c.drawString(20 * mm, y, line)
        y -= 6 * mm

    c.drawString(20 * mm, y - 8 * mm, "1.")
    c.drawString(20 * mm, y - 16 * mm, "3.2.")
    c.drawString(20 * mm, y - 24 * mm, "4.Initial of S.O.")

    c.line(20 * mm, 18 * mm, W - 20 * mm, 18 * mm)

    c.showPage()
    c.save()

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"RCI_SLIP_ticket_{ticket.id}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp

def _get_dynamic_qa(ticket, answered_by_user):
    """
    Returns list of tuples: [(question_text, answer_text), ...]
    If no answers exist for that user, it falls back to active questions with blank answers.
    """
    ans_qs = (
        TicketAnswer.objects
        .filter(ticket=ticket, answered_by=answered_by_user)
        .select_related("question")
        .order_by("question_id", "id")
    )

    if ans_qs.exists():
        qa = []
        for a in ans_qs:
            q_text = _clean(getattr(a.question, "text", None), "")
            a_text = _clean(getattr(a, "answer", None), "No.")
            qa.append((q_text, a_text))
        return qa

    # fallback: show active questions with "-" answers
    questions = list(Question.objects.filter(status="active").order_by("id"))
    return [(_clean(q.text, ""), "-") for q in questions]






# -------------------------------------------------------
# PDF: MHA Reply
# -------------------------------------------------------





def extract_html_table_data(value):
    """
    Parse simple pasted HTML tables using regex only.
    Returns:
        {
            "rows": [...],
            "has_header": True/False
        }
    or None
    """
    if value is None:
        return None

    html = str(value).strip()
    if not html:
        return None

    if not re.search(r'(?is)<table\b', html):
        return None

    table_match = re.search(r'(?is)<table\b.*?>.*?</table>', html)
    if not table_match:
        return None

    table_html = table_match.group(0)

    tr_matches = re.findall(r'(?is)<tr\b.*?>(.*?)</tr>', table_html)
    if not tr_matches:
        return None

    rows = []
    has_header = False

    for tr_html in tr_matches:
        row = []

        th_matches = re.findall(r'(?is)<th\b.*?>(.*?)</th>', tr_html)
        td_matches = re.findall(r'(?is)<td\b.*?>(.*?)</td>', tr_html)

        if th_matches:
            has_header = True
            cell_matches = th_matches
        else:
            cell_matches = td_matches

        for cell_html in cell_matches:
            cell_text = re.sub(r'(?is)<br\s*/?>', '\n', cell_html)
            cell_text = re.sub(r'(?is)</p\s*>', '\n', cell_text)
            cell_text = re.sub(r'(?is)</div\s*>', '\n', cell_text)
            cell_text = re.sub(r'(?is)<li[^>]*>', '• ', cell_text)
            cell_text = re.sub(r'(?is)</li\s*>', '\n', cell_text)
            cell_text = re.sub(r'(?is)<[^>]+>', ' ', cell_text)
            cell_text = unescape(cell_text).replace("\xa0", " ")
            cell_text = re.sub(r'[ \t]+', ' ', cell_text)
            cell_text = re.sub(r'\n{3,}', '\n\n', cell_text)
            cell_text = cell_text.strip()

            row.append(cell_text if cell_text else "-")

        if row:
            rows.append(row)

    if not rows:
        return None

    max_cols = max(len(r) for r in rows)
    normalized_rows = []
    for row in rows:
        normalized_rows.append(row + [""] * (max_cols - len(row)))

    return {
        "rows": normalized_rows,
        "has_header": has_header,
    }


def split_html_into_blocks(value):
    """
    Split answer HTML into ordered blocks:
    - {"type": "text", "content": "..."}
    - {"type": "table", "content": "<table>...</table>"}
    """
    if value is None:
        return [{"type": "text", "content": "-"}]

    html = str(value).strip()
    if not html:
        return [{"type": "text", "content": "-"}]

    blocks = []
    pattern = r'(?is)<table\b.*?>.*?</table>'

    last_end = 0
    for match in re.finditer(pattern, html):
        start, end = match.span()

        before_html = html[last_end:start]
        if before_html and before_html.strip():
            blocks.append({
                "type": "text",
                "content": before_html,
            })

        blocks.append({
            "type": "table",
            "content": match.group(0),
        })

        last_end = end

    after_html = html[last_end:]
    if after_html and after_html.strip():
        blocks.append({
            "type": "text",
            "content": after_html,
        })

    if not blocks:
        blocks.append({
            "type": "text",
            "content": html,
        })

    return blocks


@login_required(login_url="login")
def pdf_mha_reply(request, pk):
    ticket = (
        Ticket.objects.select_related("created_by")
        .prefetch_related("assigned_users")
        .get(pk=pk)
    )
    is_creator = _check_ticket_permission(ticket, request.user)
    target_user = _resolve_target_user(ticket, request, request.user, is_creator)

    osdl = _get_osdl(ticket)
    org_name = _clean(getattr(osdl, "association_name", None), "-")
    org_address = _clean(getattr(osdl, "association_address", None), "-")

    subject = _clean(
        getattr(ticket, "subject", None),
        f"Permission for grant of Registration/Prior permission under FCRA to {org_name}, {org_address}"
    )
    ref_no = _clean(getattr(ticket, "ref_no", None), "-")
    ref_date = _clean(getattr(ticket, "ref_date", None), "-")

    office_bearers = []
    bank_name = bank_address = account_no = "-"

    if osdl:
        office_bearers = _pick_office_bearers11111(osdl)
        bank_name, bank_address, account_no = _pick_bank_details(osdl)

    qa = _get_dynamic_qa(ticket, target_user)

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
        title=f"MHA_REPLY_Ticket_{ticket.id}",
    )

    styles = getSampleStyleSheet()
    P = ParagraphStyle("P", parent=styles["Normal"], fontName="Helvetica", fontSize=10.5, leading=14.5)
    PB = ParagraphStyle("PB", parent=P, fontName="Helvetica-Bold")
    PC = ParagraphStyle("PC", parent=PB, alignment=1)
    PR = ParagraphStyle("PR", parent=PB, alignment=2)

    TABLE_CELL = ParagraphStyle(
        "TABLE_CELL",
        parent=P,
        fontName="Helvetica",
        fontSize=10,
        leading=13,
        alignment=TA_LEFT,
    )

    story = []

    story.append(Paragraph("SECRET", PR))
    story.append(Spacer(1, 4))
    story.append(Paragraph("IR 002(IS)", PB))
    story.append(Spacer(1, 8))
    story.append(Paragraph("INTELLIGENCE BUREAU", PC))
    story.append(Paragraph("(MINISTRY OF HOME AFFAIRS)", PC))
    story.append(Paragraph("*****", PC))
    story.append(Spacer(1, 12))

    story.append(Paragraph(f"<b>Sub:</b> {safe_reportlab_text(subject)}", P))
    story.append(Spacer(1, 8))

    if ref_no != "-" or ref_date != "-":
        ref_line = f"MHA letter No. {ref_no} dated {ref_date}"
        story.append(Paragraph(f"<b>Ref:</b> {safe_reportlab_text(ref_line)}", P))
        story.append(Spacer(1, 10))

    story.append(Paragraph("Ad - seriatim reply is given below", PB))
    story.append(Spacer(1, 10))

    qa_flow = []
    qa_width = 160 * mm

    for i, (q, a) in enumerate(qa, start=1):
        q_txt = _clean(q, "")
        a_txt = a if a is not None else "No."

        # Question row
        qa_flow.append([
            Paragraph(f"{i})", PB),
            Paragraph(safe_reportlab_text(q_txt, ""), P),
        ])

        # Answer blocks
        answer_blocks = split_html_into_blocks(a_txt)

        first_answer_row = True

        for block in answer_blocks:
            if block["type"] == "text":
                plain_text = html_to_plain_text(block["content"], "").strip()
                if plain_text:
                    qa_flow.append([
                        "" if not first_answer_row else "",
                        Paragraph(safe_reportlab_text(plain_text, "-"), P),
                    ])
                    first_answer_row = False

            elif block["type"] == "table":
                table_data = extract_html_table_data(block["content"])
                if table_data:
                    table_rows = []
                    for row in table_data["rows"]:
                        table_rows.append([
                            Paragraph(safe_reportlab_text(cell or "-"), TABLE_CELL)
                            for cell in row
                        ])

                    col_count = len(table_rows[0]) if table_rows else 1
                    col_width = qa_width / max(col_count, 1)
                    col_widths = [col_width] * col_count

                    inner_table = Table(
                        table_rows,
                        colWidths=col_widths,
                        hAlign="LEFT",
                    )
                    inner_style = [
                        ("GRID", (0, 0), (-1, -1), 0.8, colors.HexColor("#42566e")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 8),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                    ]

                    if table_data["has_header"]:
                        inner_style.extend([
                            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                            ("LINEBELOW", (0, 0), (-1, 0), 1.2, colors.HexColor("#42566e")),
                        ])

                    inner_table.setStyle(TableStyle(inner_style))

                    qa_flow.append([
                        "",
                        inner_table,
                    ])
                    first_answer_row = False

        qa_flow.append(["", Spacer(1, 6)])

    qa_table = Table(qa_flow, colWidths=[10 * mm, 160 * mm])
    qa_table.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
    ]))
    story.append(qa_table)
    story.append(Spacer(1, 14))

    sig_name = _clean(getattr(target_user, "full_name", None), "__________")
    sig_designation = _clean(getattr(target_user, "designation", None), "Joint Deputy Director")

    story.append(Paragraph(f"({safe_reportlab_text(sig_name)})", PB))
    story.append(Paragraph(safe_reportlab_text(sig_designation), PB))
    story.append(Spacer(1, 10))

    uo_no = _clean(getattr(ticket, "uo_no", None), "-")
    uo_date = _clean(getattr(ticket, "uo_date", None), "-")
    if uo_no != "-" or uo_date != "-":
        footer_line = f"DIB U.O.No. {uo_no}    Dated: {uo_date}"
        story.append(Paragraph(safe_reportlab_text(footer_line), P))

    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()

    filename = f"MHA_REPLY_ticket_{ticket.id}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp
# -------------------------------------------------------
# PDF: Application (serve file if local path OR redirect if URL)
# -------------------------------------------------------

@login_required(login_url="login")
def pdf_application(request, pk):
    ticket = get_object_or_404(Ticket, pk=pk)
    _check_ticket_permission(ticket, request.user)

    osdl = _get_osdl(ticket)
    if not osdl:
        raise Http404("NGO/Application not found")

    path_or_url = getattr(osdl, "registration_applications_documents_location", None)

    if path_or_url:
        s = str(path_or_url).strip()

        # If it's a URL -> redirect
        if s.startswith("http://") or s.startswith("https://"):
            return JsonResponse({"redirect": s}, status=200)

        # If it's a local filesystem path -> serve file
        try:
            return FileResponse(open(s, "rb"), as_attachment=True, filename=os.path.basename(s) or "APPLICATION.pdf")
        except Exception:
            pass

    # fallback placeholder
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=18*mm, rightMargin=18*mm, topMargin=12*mm, bottomMargin=12*mm)
    styles = getSampleStyleSheet()
    P = ParagraphStyle("P", parent=styles["Normal"], fontName="Helvetica", fontSize=11, leading=16)
    PB = ParagraphStyle("PB", parent=P, fontName="Helvetica-Bold")

    story = []
    story.append(Paragraph("APPLICATION PDF NOT AVAILABLE", PB))
    story.append(Spacer(1, 10))
    story.append(Paragraph(
        "The application PDF location is not available/valid in the database for this NGO/ticket. "
        "Please store a valid URL or local path in <b>registration_applications_documents_location</b>.",
        P
    ))
    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"APPLICATION_ticket_{ticket.id}.pdf"
    resp = HttpResponse(pdf, content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def run_ticket_action_taken(ticket, user, office_file_no=None):
    office_file_no = (office_file_no or "").strip()

    already_done = TicketActionTrace.objects.filter(
        ticket=ticket,
        action_key="ACTION_TAKEN",
        performed_by=user,
        is_success=True
    ).exists()

    if already_done:
        if office_file_no:
            TicketActionTrace.objects.filter(
                ticket=ticket,
                action_key="ACTION_TAKEN",
                performed_by=user,
                is_success=True
            ).update(
                office_file_no=office_file_no,
                status_text="Action taken successfully"
            )
            return {
                "executed": True,
                "message": "Office file number updated successfully."
            }

        return {
            "executed": False,
            "message": "Action already completed by you for this ticket."
        }
    final_office_file_no = f"{office_file_no}({ticket.id})" if office_file_no else ""

    try:
        with transaction.atomic():
            
            log_ticket_action(
                ticket=ticket,
                action_key="ACTION_TAKEN",
                user=user,
                success=True,
                status_text="Action taken successfully",
                office_file_no=final_office_file_no,
                meta={"ticket_id": ticket.id}
            )

        return {
            "executed": True,
            "message": "Action taken successfully."
        }

    except Exception as e:
        
        log_ticket_action(
            ticket=ticket,
            action_key="ACTION_TAKEN",
            user=user,
            success=False,
            status_text="Action failed",
            error_text=str(e),
            office_file_no=final_office_file_no,
            meta={"ticket_id": ticket.id}
        )
        return {
            "executed": False,
            "message": str(e)
        }

@login_required(login_url="login")
def ticket_action_taken(request, ticket_id):
    if request.method != "POST":
        return JsonResponse({
            "status": "error",
            "message": "Invalid request method."
        }, status=405)

    ticket = get_object_or_404(Ticket, id=ticket_id)
    office_file_no = request.POST.get("office_file_no", "").strip()

    if not office_file_no:
        return JsonResponse({
            "status": "error",
            "message": "File number is required."
        }, status=400)

    result = run_ticket_action_taken(
        ticket=ticket,
        user=request.user,
        office_file_no=office_file_no
    )

    if result["executed"]:
        return JsonResponse({
            "status": "success",
            "message": result["message"]
        })

    return JsonResponse({
        "status": "info",
        "message": result["message"]
    })



@login_required
@user_passes_test(lambda u: u.is_superuser)
def delete_ticket(request, pk):
    if request.method != "POST":
        return JsonResponse({
            "responseCode": 405,
            "responseMessage": "Invalid request method."
        }, status=405)

    ticket = get_object_or_404(
        Ticket.all_objects.prefetch_related("reassignments"),
        pk=pk,
        is_deleted=False
    )

    if not can_admin_modify_ticket(ticket):
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": (
                "Only fresh tickets can be deleted. "
                "Action taken / answer submitted / final submitted / closed ticket cannot be deleted."
            )
        }, status=400)

    ticket.is_deleted = True
    ticket.deleted_at = timezone.now()
    ticket.deleted_by = request.user
    ticket.save(update_fields=["is_deleted", "deleted_at", "deleted_by"])

    return JsonResponse({
        "responseCode": 200,
        "responseMessage": "Ticket deleted successfully."
    })


@login_required(login_url="login")
def edit_ticket(request, pk):
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponseForbidden("You do not have permission to edit tickets.")

    ticket = get_object_or_404(
        Ticket.all_objects.select_related("created_by", "signater").prefetch_related("reassignments"),
        pk=pk,
        is_deleted=False
    )

    if not can_admin_modify_ticket(ticket):
        messages.error(
            request,
            "Only fresh tickets can be edited. "
            "Action taken / answer submitted / final submitted / closed ticket cannot be edited."
        )
        return redirect("tickets:ticket_list")

    states = list(State.objects.values("id", "name"))
    categories = list(Category.objects.all().order_by("id").values("id", "name", "description"))
    signaters = list(Signater.objects.filter(status="active").order_by("id").values("id", "ranks"))

    osdl = None
    if ticket.models_id and str(ticket.models_id).isdigit():
        osdl = OtherServicesDocumentsLink.objects.using("secondary").filter(id=int(ticket.models_id)).first()

    ngo_name = ticket.models_object or ""
    ngo_id = str(ticket.models_id or "")
    ngo_rcn = ticket.rcn or ""
    application_id = ""
    ngo_address = ""

    if osdl:
        ngo_name = getattr(osdl, "association_name", ngo_name) or ngo_name
        ngo_rcn = getattr(osdl, "rcn", ngo_rcn) or ngo_rcn
        application_id = getattr(osdl, "registration_application_id", "") or getattr(osdl, "application_id", "") or ""
        ngo_address = getattr(osdl, "association_address", "") or ""

    ngo_rows = []
    ngo_assignments = (
        AssignUsersCategory.objects
        .filter(ticket=ticket, section_type="NGO")
        .select_related("user", "category")
        .prefetch_related("questions")
    )

    for row in ngo_assignments:
        ngo_rows.append({
            "user_id": row.user_id,
            "state_id": getattr(row.user, "state_id", "") if row.user else "",
            "group": getattr(row.user, "unit", "") if row.user else "",
            "category_id": row.category_id or "",
            "question_ids": list(row.questions.values_list("id", flat=True)),
        })

    bearer_rows = []
    bearer_assignments = (
        AssignUsersCategory.objects
        .filter(ticket=ticket, section_type="OFFICE_BEARER")
        .select_related("user", "category", "member_row")
        .prefetch_related("questions")
        .order_by("id")
    )

    for row in bearer_assignments:
        member = row.member_row
        bearer_rows.append({
            "member_id": member.member_data_id if member else "",
            "member_name": member.member_name if member else "",
            "father_husband_name": member.father_husband_name if member else "",
            "nationality": member.nationality if member else "",
            "occupation": member.occupation if member else "",
            "post_in_association": member.post_in_association if member else "",
            "relationship_with_other_member": member.relationship_with_other_member if member else "",
            "address": member.address if member else "",
            "mobile": member.mobile if member else "",
            "pan_no": member.pan_no if member else "",
            "user_id": row.user_id,
            "state_id": getattr(row.user, "state_id", "") if row.user else "",
            "group": getattr(row.user, "unit", "") if row.user else "",
            "category_id": row.category_id or "",
            "question_ids": list(row.questions.values_list("id", flat=True)),
        })

    all_docs = list(
        TicketApplicationDoc.objects.using("secondary")
        .filter(ticket_id=ticket.id)
        .order_by("-uploaded_at")
    )

    application_docs = []
    supporting_docs = []

    for doc in all_docs:
        file_name = ""
        if getattr(doc, "file", None):
            file_name = str(doc.file.name or "").split("/")[-1]

        if file_name.startswith("Application_"):
            application_docs.append(doc)
        else:
            supporting_docs.append(doc)

    context = {
        "ticket": ticket,
        "states": json.dumps(states),
        "categories": categories,
        "signaters": signaters,
        "ngo_name": ngo_name,
        "ngo_id": ngo_id,
        "ngo_rcn": ngo_rcn,
        "application_id": application_id,
        "ngo_address": ngo_address,
        "edit_ngo_rows": json.dumps(ngo_rows),
        "edit_bearer_rows": json.dumps(bearer_rows),
        "application_docs": application_docs,
        "supporting_docs": supporting_docs,
    }
    return render(request, "tickets/edit_ticket.html", context)
@login_required(login_url="login")
def update_ticket(request, pk):
    if request.method != "POST":
        return JsonResponse({
            "responseCode": 405,
            "responseMessage": "Invalid request method."
        }, status=405)

    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({
            "responseCode": 403,
            "responseMessage": "You do not have permission to edit ticket."
        }, status=403)

    ticket = get_object_or_404(Ticket.all_objects, pk=pk, is_deleted=False)

    if not can_admin_modify_ticket(ticket):
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": (
                "Only fresh tickets can be edited. "
                "Action taken / answer submitted / final submitted / closed ticket cannot be updated."
            )
        }, status=400)

    User = get_user_model()

    try:
        with transaction.atomic():
            data = _get_payload(request)
            description = data.get("description")
            file_no = (request.POST.get("file_no") or "").strip()
            mhafile_no = (request.POST.get("mhafile_no") or "").strip()
            signater_id = (request.POST.get("signater_id") or "").strip()
            ngoId = data.get("ngoId")
            users = data.get("users", [])

            if not ngoId:
                return JsonResponse({
                    "responseCode": 400,
                    "responseMessage": "Please select a valid NGO."
                }, status=400)

            osdl = OtherServicesDocumentsLink.objects.using("secondary").filter(id=int(ngoId)).first()
            if not osdl:
                return JsonResponse({
                    "responseCode": 400,
                    "responseMessage": "Invalid NGO selected."
                }, status=400)

            models_object = getattr(osdl, "association_name", "") or ""
            rcn = getattr(osdl, "rcn", None)

            old_user_ids = set(ticket.assigned_users.values_list("id", flat=True))

            new_user_ids = set()
            for u in users:
                try:
                    uid = int(u.get("user_id"))
                    new_user_ids.add(uid)
                except (TypeError, ValueError):
                    continue

            removed_user_ids = old_user_ids - new_user_ids

            ticket.description = description
            ticket.models_id = osdl.id
            ticket.models_object = models_object
            ticket.rcn = rcn
            ticket.file_no = file_no
            ticket.mha_file_no = mhafile_no
            ticket.signater_id = signater_id if signater_id else None
            ticket.save(update_fields=[
                "description",
                "models_id",
                "models_object",
                "rcn",
                "file_no",
                "mha_file_no",
                "signater",
                "updated_at",
            ])

            if removed_user_ids:
                TicketAnswer.objects.filter(
                    ticket=ticket,
                    answered_by_id__in=removed_user_ids
                ).delete()

                AssignUsersCategory.objects.filter(
                    ticket=ticket,
                    user_id__in=removed_user_ids
                ).delete()

                ReportSummary.objects.filter(
                    ticket=ticket,
                    user_id__in=removed_user_ids,
                    reassignment__isnull=True
                ).delete()

                TicketSignater.objects.filter(
                    ticket=ticket,
                    user_id__in=removed_user_ids
                ).delete()

                old_logs_qs = TicketLog.objects.filter(
                    ticket=ticket,
                    ticketReassignment__isnull=True
                ).filter(
                    Q(user_id__in=removed_user_ids) |
                    Q(to_user_id__in=removed_user_ids)
                )

                old_log_ids = list(old_logs_qs.values_list("id", flat=True))

                if old_log_ids:
                    FileUpload.objects.filter(log_id__in=old_log_ids).delete()
                    old_logs_qs.delete()

                old_reassignments = TicketReassignment.objects.filter(
                    ticket=ticket,
                    new_assigned_users__id__in=removed_user_ids
                ).distinct()

                old_reassignment_ids = list(old_reassignments.values_list("id", flat=True))

                if old_reassignment_ids:
                    ReassignmentAnswer.objects.filter(
                        reassignment_id__in=old_reassignment_ids
                    ).delete()

                    ReportSummary.objects.filter(
                        ticket=ticket,
                        reassignment_id__in=old_reassignment_ids
                    ).delete()

                    FileUpload.objects.filter(
                        reassignment_id__in=old_reassignment_ids
                    ).delete()

                    TicketLog.objects.filter(
                        ticket=ticket,
                        ticketReassignment_id__in=old_reassignment_ids
                    ).delete()

                    old_reassignments.delete()

            ticket.assigned_users.clear()

            AssignUsersCategory.objects.filter(ticket=ticket).delete()
            TicketAnswer.objects.filter(ticket=ticket).delete()
            TicketMemberRow.objects.filter(ticket=ticket).delete()
            TicketSectionMeta.objects.filter(ticket=ticket).delete()

            def _norm_section(val):
                return (val or "").strip().lower()

            def _safe_int(val):
                try:
                    if val is None:
                        return None
                    s = str(val).strip()
                    return int(s) if s else None
                except Exception:
                    return None

            has_user_section = any(_norm_section(u.get("section")) == "user_section" for u in users)
            has_bearer_section = any(_norm_section(u.get("section")) == "office_bearer" for u in users)

            meta_payload = {
                "ngo_id": str(ngoId),
                "ngo_name": models_object,
                "rcn": rcn,
                "application_id": (request.POST.get("application_id") or ""),
                "module": "NGO",
            }

            if has_user_section:
                TicketSectionMeta.objects.get_or_create(
                    ticket=ticket,
                    section_type="USER_SECTION",
                    defaults={"created_by": request.user, "meta": meta_payload},
                )

            if has_bearer_section:
                TicketSectionMeta.objects.get_or_create(
                    ticket=ticket,
                    section_type="OFFICE_BEARER",
                    defaults={"created_by": request.user, "meta": meta_payload},
                )

            bearer_member_cache = {}

            for u in users:
                user_id = u.get("user_id")
                questions = u.get("questions", [])
                section = _norm_section(u.get("section"))
                db_section = "NGO" if section == "user_section" else "OFFICE_BEARER"

                category_obj = None
                cat_id = u.get("category_id")
                if cat_id:
                    try:
                        category_obj = Category.objects.filter(id=int(cat_id)).first()
                    except Exception:
                        category_obj = None

                member_row_obj = None
                if db_section == "OFFICE_BEARER":
                    member_data_id = _safe_int(u.get("member_id"))
                    member_name = (u.get("member_name") or "").strip() or None
                    member_key = (u.get("member_key") or "").strip() or member_name

                    if member_key:
                        if member_key in bearer_member_cache:
                            member_row_obj = bearer_member_cache[member_key]
                        else:
                            member_row_obj = TicketMemberRow.objects.create(
                                ticket=ticket,
                                section_type="OFFICE_BEARER",
                                member_data_id=member_data_id,
                                member_name=member_name,
                                father_husband_name=u.get("father_husband_name"),
                                nationality=u.get("nationality"),
                                occupation=u.get("occupation"),
                                post_in_association=u.get("post_in_association"),
                                relationship_with_other_member=u.get("relationship_with_other_member"),
                                address=u.get("address"),
                                mobile=u.get("mobile"),
                                pan_no=u.get("pan_no"),
                                created_by=request.user,
                            )
                            bearer_member_cache[member_key] = member_row_obj

                try:
                    assigned_user = User.objects.get(id=user_id)
                except User.DoesNotExist:
                    continue

                ticket.assigned_users.add(assigned_user)

                assign_obj, _ = AssignUsersCategory.objects.get_or_create(
                    ticket=ticket,
                    section_type=db_section,
                    member_row=member_row_obj if db_section == "OFFICE_BEARER" else None,
                    user=assigned_user,
                    category=category_obj,
                )

                for q_item in questions:
                    q_obj = None
                    try:
                        q_obj = Question.objects.filter(pk=int(q_item)).first()
                    except Exception:
                        pass

                    if q_obj:
                        TicketAnswer.objects.get_or_create(
                            ticket=ticket,
                            question=q_obj,
                            answered_by=assigned_user
                        )
                        assign_obj.questions.add(q_obj)

            TicketActionTrace.objects.create(
                ticket=ticket,
                action_key="TICKET_UPDATED",
                is_success=True,
                status_text="Ticket updated successfully",
                performed_by=request.user,
                office_file_no=file_no,
                meta={
                    "ticket_id": ticket.id,
                    "old_user_ids": list(old_user_ids),
                    "new_user_ids": list(new_user_ids),
                    "removed_user_ids": list(removed_user_ids),
                }
            )

            return JsonResponse({
                "responseCode": 200,
                "responseMessage": "Ticket updated successfully!",
                "ticket_id": ticket.id,
            })

    except Exception as e:
        logger.exception("Ticket update failed")
        return JsonResponse({
            "responseCode": 400,
            "responseMessage": f"Error updating ticket: {str(e)}"
        }, status=400)
    


@login_required(login_url="login")
def uo_upload(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)

    if request.method != "POST":
        return redirect("tickets:ticket_detail", pk=ticket.id)

    try:
        uo_file = request.FILES.get("uo_file")

        if not uo_file:
            messages.error(request, "Please choose a UO file.")
            return redirect("tickets:ticket_detail", pk=ticket.id)

        rcn = (request.POST.get("rcn") or "").strip() or None
        association_name = (request.POST.get("association_name") or "").strip() or None

        document_category = (request.POST.get("document_category") or "uo").strip().lower()
        if document_category not in ["uo", "report"]:
            document_category = "uo"

        ext = uo_file.name.rsplit(".", 1)[-1].lower() if "." in uo_file.name else ""

        if ext == "pdf":
            doc_type = "pdf"
        elif ext in ["xls", "xlsx"]:
            doc_type = "excel"
        elif ext == "csv":
            doc_type = "csv"
        elif ext in ["jpg", "jpeg", "png", "webp"]:
            doc_type = "image"
        else:
            doc_type = "other"

        base_dir = getattr(settings, "NGO_UO_DOCUMENT_ROOT", None) or settings.MEDIA_ROOT

        filename = f"uo_{uuid.uuid4().hex}.{ext}" if ext else f"uo_{uuid.uuid4().hex}"
        relative_path = os.path.join("uo_documents", filename).replace("\\", "/")
        full_path = os.path.join(base_dir, relative_path)

        os.makedirs(os.path.dirname(full_path), exist_ok=True)

        with open(full_path, "wb+") as dest:
            for chunk in uo_file.chunks():
                dest.write(chunk)

        now = timezone.now()

        NgoUoDocument.objects.using("secondary").create(
            rcn=rcn,
            association_name=association_name,
            document_category=document_category,
            doc_type=doc_type,
            original_name=uo_file.name,
            file_path=relative_path,
            mime_type=uo_file.content_type or None,
            file_ext=ext or None,
            file_size=uo_file.size,
            created_at=now,
            updated_at=now,
        )

        messages.success(request, "UO file uploaded successfully.")

    except Exception as e:
        logger.exception("UO upload failed for ticket_id=%s", ticket.id)
        messages.error(request, f"UO upload failed: {str(e)}")

    return redirect("tickets:ticket_detail", pk=ticket.id)




@login_required(login_url="login")
@require_POST
def creator_action_taken(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)

    if not (request.user.is_superuser or ticket.created_by_id == request.user.id):
        return JsonResponse({
            "status": "error",
            "message": "Only creator can mark this."
        }, status=403)

    already_done = TicketActionTrace.objects.filter(
        ticket=ticket,
        action_key="CREATOR_ACTION_TAKEN",
        performed_by=request.user,
        is_success=True,
    ).exists()

    if not already_done:
        log_ticket_action(
            ticket=ticket,
            action_key="CREATOR_ACTION_TAKEN",
            user=request.user,
            success=True,
            status_text="Creator checkbox marked",
            meta={"ticket_id": ticket.id, "source": "ticket_list_creator_checkbox"},
        )

    return JsonResponse({
        "status": "success",
        "message": "Marked successfully."
    })